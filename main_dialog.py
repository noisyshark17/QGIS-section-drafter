"""
QGIS Polyline Creator with Raster Draping, Slope Profiler, and Intersection Points
Combined tool for creating draped polylines, extracting elevation profiles, and finding intersection points
Version: 2.5 - Fixed Excel export crashes and improved intersection point display
"""

from qgis.core import (
    QgsVectorLayer, QgsFeature, QgsGeometry, QgsPointXY, QgsPoint,
    QgsProject, QgsFields, QgsField, QgsWkbTypes,
    QgsCoordinateReferenceSystem, QgsCoordinateTransform,
    QgsRasterLayer, QgsLineString, QgsVectorFileWriter,
    QgsMapLayerProxyModel, QgsRaster, QgsMapLayer,
    QgsMarkerSymbol, QgsMessageLog, Qgis,
    QgsFeatureRequest, QgsSpatialIndex, NULL
)
from qgis.gui import QgsMapToolEmitPoint, QgsRubberBand
from PyQt5.QtCore import QVariant, Qt, pyqtSignal, QCoreApplication, QSettings
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
    QLineEdit, QPushButton, QComboBox, QCheckBox,
    QDialogButtonBox, QMessageBox, QGroupBox,
    QSpinBox, QDoubleSpinBox, QListWidget,
    QRadioButton, QButtonGroup, QFileDialog,
    QProgressDialog, QTabWidget, QWidget,
    QTextEdit, QAbstractItemView, QTableWidget,
    QTableWidgetItem, QHeaderView, QColorDialog,
    QScrollArea, QSizePolicy, QApplication, QFrame
)
from PyQt5.QtGui import QColor, QFont, QPixmap
import processing
import os
from pathlib import Path
import math
import traceback
import csv
import gc

# Import core modules
from .core.map_tools import PolylineMapTool
from .core.excel_utils import SafeExcelWriter, EXCEL_ENGINE
from .gui.professional_theme import ProfessionalTheme

import re
from typing import Optional, List, Dict, Any, Tuple
import geopandas as gpd

# Try to import pandas for data processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Pandas not available for data processing")

# Try to import matplotlib for plotting
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.backends.backend_pdf import PdfPages
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Matplotlib not available for plotting")


class PlotDisplayWidget(QWidget):
    """Widget to display matplotlib plots in the dialog"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        
        # Create scroll area for plots
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # Content widget that goes inside scroll area
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(5, 5, 5, 5)
        self.content_layout.setSpacing(10)
        
        self.scroll_area.setWidget(self.content_widget)
        self.layout.addWidget(self.scroll_area)
        self.setLayout(self.layout)
        self.canvases = []
        
        # Set minimum size for the widget
        self.setMinimumSize(600, 400)
        
    def add_plot(self, figure):
        """Add a matplotlib figure to the display with full size"""
        # Set figure size to ensure good quality display
        figure.set_size_inches(12, 8)  # Larger figure size
        figure.set_dpi(100)  # Good resolution
        
        # Create canvas
        canvas = FigureCanvas(figure)
        
        # Set canvas size to display the full plot
        canvas.setMinimumHeight(600)  # Increased from 400
        canvas.setMinimumWidth(800)   # Set minimum width
        canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        # Add some spacing and a separator line
        if self.canvases:  # Add separator if not the first plot
            separator = QFrame()
            separator.setFrameShape(QFrame.HLine)
            separator.setFrameShadow(QFrame.Sunken)
            separator.setStyleSheet("QFrame { color: #CCCCCC; }")
            self.content_layout.addWidget(separator)
        
        # Add plot title label
        plot_title = figure.axes[0].get_title() if figure.axes else "Plot"
        title_label = QLabel(f"<b>{plot_title}</b>")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("QLabel { font-size: 14px; margin: 5px; }")
        self.content_layout.addWidget(title_label)
        
        # Add the canvas
        self.content_layout.addWidget(canvas)
        self.canvases.append(canvas)
        
        # Ensure the canvas displays the plot properly
        canvas.draw()
        
    def clear_plots(self):
        """Clear all plots"""
        # Clear all widgets from the layout
        while self.content_layout.count():
            child = self.content_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.canvases.clear()

class PlotConfigDialog(QDialog):
    """Dialog for configuring plot parameters"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Plot Configuration")
        self.setModal(True)
        # Make dialog fit on screen better
        self.setMinimumSize(450, 500)
        self.resize(500, 650)
        self.init_ui()
        
        # Initialize Excel file display
        self.initialize_excel_file_display()
        
    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setSpacing(4)
        main_layout.setContentsMargins(8, 8, 8, 8)
        
        # Create scroll area for all content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # Content widget that goes inside scroll area
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)
        layout.setSpacing(6)
        layout.setContentsMargins(4, 4, 4, 4)
        
        # Colors group - more compact
        colors_group = QGroupBox("Colors")
        colors_layout = QVBoxLayout()
        colors_layout.setSpacing(3)
        
        self.color_settings = {
            'main_line': {'label': 'Main Line', 'default': '#000000'},
            'feature_line': {'label': 'Feature/Secondary Lines', 'default': '#0000FF'},
            'slope_line': {'label': 'Slope Analysis Lines', 'default': '#FF0000'},
            'runout_line': {'label': 'Runout Analysis Lines', 'default': '#008000'},
            'reference_line': {'label': 'Reference Angle Lines', 'default': '#800080'},
            'marker_1': {'label': 'Marker 1', 'default': '#0070FF'},
            'marker_2': {'label': 'Marker 2', 'default': '#0000FF'}
        }
        
        self.color_buttons = {}
        # Create a more compact grid layout for colors
        for i, (key, settings) in enumerate(self.color_settings.items()):
            h_layout = QHBoxLayout()
            h_layout.setSpacing(4)
            
            label = QLabel(settings['label'] + ':')
            label.setMinimumWidth(120)
            h_layout.addWidget(label)
            
            color_btn = QPushButton()
            color_btn.setFixedSize(40, 25)
            color_btn.setStyleSheet(f"background-color: {settings['default']}; border: 1px solid #999;")
            color_btn.clicked.connect(lambda checked, k=key: self.choose_color(k))
            self.color_buttons[key] = color_btn
            
            h_layout.addWidget(color_btn)
            h_layout.addStretch()
            colors_layout.addLayout(h_layout)
            
        colors_group.setLayout(colors_layout)
        layout.addWidget(colors_group)
        
        # Font sizes group - more compact
        font_group = QGroupBox("Font Sizes")
        font_layout = QVBoxLayout()
        font_layout.setSpacing(3)
        
        self.font_spinboxes = {}
        font_settings = [
            ('title_font', 'Title', 14),
            ('label_font', 'Axis Labels', 12),
            ('legend_font', 'Legend', 10),
            ('text_font', 'Text', 10)
        ]
        
        for key, label, default in font_settings:
            h_layout = QHBoxLayout()
            h_layout.setSpacing(4)
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(120)
            h_layout.addWidget(label_widget)
            
            spinbox = QSpinBox()
            spinbox.setRange(6, 24)
            spinbox.setValue(default)
            spinbox.setMaximumWidth(60)
            self.font_spinboxes[key] = spinbox
            
            h_layout.addWidget(spinbox)
            h_layout.addStretch()
            font_layout.addLayout(h_layout)
            
        font_group.setLayout(font_layout)
        layout.addWidget(font_group)
        
        # Marker styles and sizes combined group
        marker_group = QGroupBox("Marker Styles & Sizes")
        marker_layout = QVBoxLayout()
        marker_layout.setSpacing(3)
        
        self.marker_settings = {}
        self.size_settings = {}
        
        marker_types = [
            ('marker_1', 'Marker 1 (Excel Col A)', 'o', 8),
            ('marker_2', 'Marker 2 (Excel Col B)', 's', 8)
        ]
        
        # Available matplotlib marker styles (reduced list)
        marker_options = [
            ('o', 'Circle'), ('s', 'Square'), ('^', 'Triangle Up'), ('v', 'Triangle Down'),
            ('D', 'Diamond'), ('*', 'Star'), ('+', 'Plus'), ('x', 'X'), ('.', 'Point'),
            ('|', 'Vertical Line'), ('_', 'Horizontal Line'), ('1', 'Tri Down'), ('2', 'Tri Up')
        ]
        
        for key, label, default_style, default_size in marker_types:
            h_layout = QHBoxLayout()
            h_layout.setSpacing(4)
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(120)
            h_layout.addWidget(label_widget)
            
            # Style combo
            combo = QComboBox()
            combo.setMaximumWidth(100)
            for marker_code, marker_name in marker_options:
                combo.addItem(f"{marker_name}", marker_code)
                if marker_code == default_style:
                    combo.setCurrentIndex(combo.count() - 1)
            
            self.marker_settings[key] = combo
            h_layout.addWidget(combo)
            
            # Size spinbox
            size_spinbox = QSpinBox()
            size_spinbox.setRange(1, 20)
            size_spinbox.setValue(default_size)
            size_spinbox.setMaximumWidth(50)
            self.size_settings[key + '_size'] = size_spinbox
            
            h_layout.addWidget(QLabel('Size:'))
            h_layout.addWidget(size_spinbox)
            h_layout.addStretch()
            marker_layout.addLayout(h_layout)
            
        marker_group.setLayout(marker_layout)
        layout.addWidget(marker_group)
        
        # Line thickness group
        thickness_group = QGroupBox("Line Thickness")
        thickness_layout = QVBoxLayout()
        thickness_layout.setSpacing(3)
        
        self.thickness_settings = {}
        
        thickness_options = [
            ('main_line', 'Main Profile Line', 2.0),
            ('feature_line', 'Feature/Secondary Lines', 1.5),
            ('slope_line', 'Slope Analysis Lines', 2.0),
            ('runout_line', 'Runout Analysis Lines', 2.0),
            ('reference_line', 'Reference Angle Lines', 1.5),
            ('grid_lines', 'Grid Lines', 0.5)
        ]
        
        for key, label, default_thickness in thickness_options:
            h_layout = QHBoxLayout()
            h_layout.setSpacing(4)
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(120)
            h_layout.addWidget(label_widget)
            
            # Thickness spin box with decimal precision
            thickness_spinbox = QDoubleSpinBox()
            thickness_spinbox.setRange(0.1, 10.0)
            thickness_spinbox.setValue(default_thickness)
            thickness_spinbox.setSingleStep(0.1)
            thickness_spinbox.setDecimals(1)
            thickness_spinbox.setMaximumWidth(70)
            thickness_spinbox.setSuffix(' pt')
            self.thickness_settings[key + '_thickness'] = thickness_spinbox
            
            h_layout.addWidget(thickness_spinbox)
            h_layout.addStretch()
            thickness_layout.addLayout(h_layout)
            
        thickness_group.setLayout(thickness_layout)
        layout.addWidget(thickness_group)
        
        # Excel-Based Auto-Configuration group
        auto_config_group = QGroupBox("Excel-Based Auto-Configuration")
        auto_config_layout = QVBoxLayout()
        auto_config_layout.setSpacing(6)
        
        # Information display area
        info_layout = QVBoxLayout()
        info_layout.setSpacing(4)
        
        self.excel_file_label = QLabel("Excel File: <i>Not selected</i>")
        self.excel_file_label.setWordWrap(True)
        info_layout.addWidget(self.excel_file_label)
        
        self.detected_profiles_label = QLabel("Detected Profiles: <i>None</i>")
        self.detected_profiles_label.setWordWrap(True)
        info_layout.addWidget(self.detected_profiles_label)
        
        self.detected_markers_label = QLabel("Detected Markers: <i>None</i>")
        self.detected_markers_label.setWordWrap(True)
        info_layout.addWidget(self.detected_markers_label)
        
        auto_config_layout.addLayout(info_layout)
        
        # Auto-detect button
        auto_detect_layout = QHBoxLayout()
        self.auto_detect_btn = QPushButton("Auto-Detect from Excel")
        self.auto_detect_btn.setMinimumHeight(32)
        self.auto_detect_btn.setToolTip("Automatically configure plot settings based on Excel file structure")
        self.auto_detect_btn.clicked.connect(self.auto_detect_configuration)
        auto_detect_layout.addWidget(self.auto_detect_btn)
        auto_detect_layout.addStretch()
        
        auto_config_layout.addLayout(auto_detect_layout)
        
        # Status indicator
        self.auto_config_status = QLabel("Status: <i>Ready for auto-detection</i>")
        self.auto_config_status.setStyleSheet("color: #666; font-style: italic;")
        auto_config_layout.addWidget(self.auto_config_status)
        
        auto_config_group.setLayout(auto_config_layout)
        layout.addWidget(auto_config_group)
        
        # Legend labels group - more compact
        legend_group = QGroupBox("Legend Labels")
        legend_layout = QVBoxLayout()
        legend_layout.setSpacing(3)
        
        self.legend_edits = {}
        legend_defaults = [
            ('main_profile', 'Main Profile', 'Existing Ground Profile'),
            ('markers_1', 'Markers 1', 'Markers 1'),
            ('markers_2', 'Markers 2', 'Markers 2'),
            ('feature', 'Feature', 'Feature')
        ]
        
        for key, label, default in legend_defaults:
            h_layout = QHBoxLayout()
            h_layout.setSpacing(4)
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(120)
            h_layout.addWidget(label_widget)
            
            edit = QLineEdit(default)
            edit.setMaximumWidth(200)
            self.legend_edits[key] = edit
            
            h_layout.addWidget(edit)
            h_layout.addStretch()
            legend_layout.addLayout(h_layout)
            
        legend_group.setLayout(legend_layout)
        layout.addWidget(legend_group)
        
        # Excel Column Mapping group
        mapping_group = QGroupBox("Excel Column Mapping Reference")
        mapping_layout = QVBoxLayout()
        mapping_layout.setSpacing(6)
        
        # Information text
        info_text = QLabel(
            "<b>Plot Markers correspond to Excel columns as follows:</b><br>"
            "• <b>Marker 1</b> → Excel Column A (First data column)<br>"
            "• <b>Marker 2</b> → Excel Column B (Second data column)<br>"
            "<br><i>Note: Configure marker styles and colors above. "
            "Use 'Browse Excel File' to auto-configure based on your data structure.</i>"
        )
        info_text.setWordWrap(True)
        info_text.setStyleSheet("padding: 8px; background-color: #f0f0f0; border: 1px solid #ccc; border-radius: 4px;")
        mapping_layout.addWidget(info_text)
        
        # Browse Excel button
        browse_layout = QHBoxLayout()
        browse_layout.setSpacing(8)
        
        browse_excel_btn = QPushButton("Browse Excel File for Column Structure")
        browse_excel_btn.setMinimumHeight(30)
        browse_excel_btn.clicked.connect(self.browse_excel_columns)
        browse_layout.addWidget(browse_excel_btn)
        browse_layout.addStretch()
        
        mapping_layout.addLayout(browse_layout)
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        # Set the content widget to the scroll area
        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area)
        
        # Buttons at bottom (not scrolled)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)
        
        ok_btn = QPushButton("OK")
        ok_btn.setMinimumHeight(30)
        ok_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setMinimumHeight(30)
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
        
    def choose_color(self, key):
        """Open color dialog for button"""
        color = QColorDialog.getColor()
        if color.isValid():
            self.color_buttons[key].setStyleSheet(f"background-color: {color.name()}; border: 1px solid #999;")
            
    def get_config(self):
        """Get configuration as dictionary"""
        config = {}
        
        # Get colors
        for key, btn in self.color_buttons.items():
            style = btn.styleSheet()
            if 'background-color:' in style:
                color = style.split('background-color:')[1].split(';')[0].strip()
                config[key + '_color'] = color
            else:
                config[key + '_color'] = self.color_settings[key]['default']
            
        # Get font sizes
        for key, spinbox in self.font_spinboxes.items():
            config[key + '_size'] = str(spinbox.value())
            
        # Get legend labels
        for key, edit in self.legend_edits.items():
            config['legend_' + key] = edit.text()
            
        # Get marker styles
        for key, combo in self.marker_settings.items():
            config[key + '_marker_style'] = combo.currentData()
            
        # Get marker sizes
        for key, spinbox in self.size_settings.items():
            config[key] = str(spinbox.value())
        
        # Get line thickness settings
        for key, spinbox in self.thickness_settings.items():
            config[key] = str(spinbox.value())
        
        # Add backward compatibility defaults
        config['marker_size'] = config.get('marker_1_size', '8')
        config['marker_style'] = config.get('marker_1_marker_style', 'o')
        config['figure_size'] = '10'
        
        # Add missing legend labels for backward compatibility
        if 'legend_angular_elevation' not in config:
            config['legend_angular_elevation'] = 'Angular Elevation'
        
        return config
    
    def auto_detect_configuration(self):
        """Auto-detect plot configuration from Excel file"""
        try:
            # Get Excel file from parent dialog if available
            excel_file = self.get_excel_file_path()
            if not excel_file:
                self.auto_config_status.setText("Status: <span style='color: #d32f2f;'>No Excel file selected</span>")
                return
                
            # Update status
            self.auto_config_status.setText("Status: <span style='color: #1976d2;'>Analyzing Excel file...</span>")
            QApplication.processEvents()
            
            # Analyze Excel file
            analysis_result = self.analyze_excel_file(excel_file)
            
            if analysis_result:
                # Apply intelligent configuration
                self.apply_auto_configuration(analysis_result)
                
                # Update display labels
                self.update_detection_display(analysis_result)
                
                self.auto_config_status.setText("Status: <span style='color: #388e3c;'>Auto-configuration completed successfully!</span>")
            else:
                self.auto_config_status.setText("Status: <span style='color: #d32f2f;'>Could not analyze Excel file structure</span>")
                
        except Exception as e:
            self.auto_config_status.setText(f"Status: <span style='color: #d32f2f;'>Error: {str(e)}</span>")
    
    def get_excel_file_path(self):
        """Get Excel file path from parent dialog or prompt user"""
        # Try to get from parent dialog first
        if hasattr(self.parent(), 'excel_file_edit') and self.parent().excel_file_edit.text():
            return self.parent().excel_file_edit.text()
        
        # If not available, prompt user to select
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Select Excel File for Auto-Configuration",
            "",
            "Excel files (*.xlsx *.xls)"
        )
        return file_path if file_path else None
    
    def analyze_excel_file(self, excel_file):
        """Analyze Excel file structure to detect profiles and markers"""
        try:
            import pandas as pd
            
            # Read Excel file
            df = pd.read_excel(excel_file)
            columns = [col.lower() for col in df.columns]
            
            # Initialize analysis result
            analysis = {
                'file_path': excel_file,
                'total_columns': len(df.columns),
                'profiles': [],
                'markers': [],
                'suggested_colors': {},
                'suggested_markers': {},
                'suggested_legends': {}
            }
            
            # Profile detection patterns
            profile_patterns = {
                'main_profile': ['distance', 'elevation', 'ground', 'existing', 'natural'],
                'slope_profile': ['slope', 'cut', 'designed', 'proposed'],
                'reference_profile': ['reference', 'ref', 'original', 'baseline'],
                'feature_profile': ['feature', 'structure', 'wall', 'barrier'],
                'runout_profile': ['runout', 'debris', 'flow', 'avalanche']
            }
            
            # Marker detection patterns
            marker_patterns = {
                'boreholes': ['borehole', 'bh', 'drill', 'boring'],
                'monitoring': ['monitor', 'instrument', 'sensor', 'gauge'],
                'structures': ['structure', 'building', 'wall', 'fence'],
                'features': ['feature', 'point', 'marker', 'location'],
                'incidents': ['incident', 'failure', 'event', 'damage']
            }
            
            # Analyze columns for profiles
            for profile_type, keywords in profile_patterns.items():
                matching_cols = []
                for col in df.columns:
                    col_lower = col.lower()
                    if any(keyword in col_lower for keyword in keywords):
                        matching_cols.append(col)
                
                if matching_cols:
                    analysis['profiles'].append({
                        'type': profile_type,
                        'columns': matching_cols,
                        'display_name': self.generate_profile_display_name(profile_type, matching_cols)
                    })
            
            # Analyze columns for markers
            for marker_type, keywords in marker_patterns.items():
                matching_cols = []
                for col in df.columns:
                    col_lower = col.lower()
                    if any(keyword in col_lower for keyword in keywords):
                        matching_cols.append(col)
                
                if matching_cols:
                    analysis['markers'].append({
                        'type': marker_type,
                        'columns': matching_cols,
                        'display_name': self.generate_marker_display_name(marker_type, matching_cols)
                    })
            
            # Generate intelligent suggestions
            self.generate_intelligent_suggestions(analysis)
            
            return analysis
            
        except Exception as e:
            print(f"Error analyzing Excel file: {str(e)}")
            return None
    
    def generate_profile_display_name(self, profile_type, columns):
        """Generate user-friendly display name for profile"""
        name_mapping = {
            'main_profile': 'Existing Ground Profile',
            'slope_profile': 'Slope/Cut Profile', 
            'reference_profile': 'Reference Profile',
            'feature_profile': 'Feature Profile',
            'runout_profile': 'Runout Profile'
        }
        
        base_name = name_mapping.get(profile_type, profile_type.replace('_', ' ').title())
        
        # Add specific column info if helpful
        if len(columns) == 1:
            return f"{base_name} ({columns[0]})"
        elif len(columns) <= 3:
            return f"{base_name} ({', '.join(columns)})"
        else:
            return f"{base_name} ({len(columns)} columns)"
    
    def generate_marker_display_name(self, marker_type, columns):
        """Generate user-friendly display name for markers"""
        name_mapping = {
            'boreholes': 'Borehole Locations',
            'monitoring': 'Monitoring Points',
            'structures': 'Structures',
            'features': 'Feature Points',
            'incidents': 'Incident Locations'
        }
        
        base_name = name_mapping.get(marker_type, marker_type.replace('_', ' ').title())
        
        if len(columns) == 1:
            return f"{base_name} ({columns[0]})"
        else:
            return f"{base_name} ({len(columns)} columns)"
    
    def generate_intelligent_suggestions(self, analysis):
        """Generate intelligent color, marker, and legend suggestions"""
        # Color scheme for different profile types
        profile_color_scheme = {
            'main_profile': '#000000',      # Black for existing ground
            'slope_profile': '#FF0000',     # Red for slopes/cuts
            'reference_profile': '#800080', # Purple for reference
            'feature_profile': '#0000FF',   # Blue for features
            'runout_profile': '#008000'     # Green for runout
        }
        
        # Marker styles for different marker types
        marker_style_scheme = {
            'boreholes': ('o', 8),          # Circle, size 8
            'monitoring': ('^', 6),         # Triangle up, size 6
            'structures': ('s', 8),         # Square, size 8
            'features': ('D', 6),           # Diamond, size 6
            'incidents': ('*', 10)          # Star, size 10
        }
        
        # Apply color suggestions for profiles
        for profile in analysis['profiles']:
            profile_type = profile['type']
            if profile_type in profile_color_scheme:
                analysis['suggested_colors'][profile_type] = profile_color_scheme[profile_type]
        
        # Apply marker suggestions
        for marker in analysis['markers']:
            marker_type = marker['type']
            if marker_type in marker_style_scheme:
                style, size = marker_style_scheme[marker_type]
                analysis['suggested_markers'][marker_type] = {'style': style, 'size': size}
        
        # Generate legend suggestions
        for profile in analysis['profiles']:
            analysis['suggested_legends'][profile['type']] = profile['display_name']
        
        for marker in analysis['markers']:
            analysis['suggested_legends'][marker['type']] = marker['display_name']
    
    def apply_auto_configuration(self, analysis):
        """Apply auto-detected configuration to dialog controls"""
        try:
            # Apply color suggestions
            for profile_type, color in analysis['suggested_colors'].items():
                if profile_type in self.color_buttons:
                    self.color_buttons[profile_type].setStyleSheet(
                        f"background-color: {color}; border: 1px solid #999;"
                    )
            
            # Apply marker style suggestions
            for marker_type, settings in analysis['suggested_markers'].items():
                # Update marker style
                if f"{marker_type}_marker" in self.marker_settings:
                    combo = self.marker_settings[f"{marker_type}_marker"]
                    for i in range(combo.count()):
                        if combo.itemData(i) == settings['style']:
                            combo.setCurrentIndex(i)
                            break
                
                # Update marker size
                if f"{marker_type}_marker_size" in self.size_settings:
                    self.size_settings[f"{marker_type}_marker_size"].setValue(settings['size'])
            
            # Apply legend suggestions
            for item_type, legend_text in analysis['suggested_legends'].items():
                # Map to legend edit fields
                legend_mapping = {
                    'main_profile': 'main_profile',
                    'slope_profile': 'feature',
                    'feature_profile': 'profile_3',
                    'boreholes': 'markers_1',
                    'monitoring': 'markers_2'
                }
                
                if item_type in legend_mapping:
                    legend_key = legend_mapping[item_type]
                    if legend_key in self.legend_edits:
                        self.legend_edits[legend_key].setText(legend_text)
            
            # Apply intelligent defaults for font sizes based on complexity
            profile_count = len(analysis['profiles'])
            marker_count = len(analysis['markers'])
            
            if profile_count + marker_count > 5:
                # More complex plot, use smaller fonts
                self.font_spinboxes['legend_font'].setValue(9)
                self.font_spinboxes['text_font'].setValue(9)
            else:
                # Simpler plot, use standard fonts
                self.font_spinboxes['legend_font'].setValue(10)
                self.font_spinboxes['text_font'].setValue(10)
                
        except Exception as e:
            print(f"Error applying auto-configuration: {str(e)}")
    
    def update_detection_display(self, analysis):
        """Update the detection display labels"""
        try:
            # Update file path
            import os
            file_name = os.path.basename(analysis['file_path'])
            self.excel_file_label.setText(f"Excel File: <b>{file_name}</b>")
            
            # Update detected profiles
            if analysis['profiles']:
                profile_text = "<br>".join([f"• {p['display_name']}" for p in analysis['profiles']])
                self.detected_profiles_label.setText(f"Detected Profiles: <br>{profile_text}")
            else:
                self.detected_profiles_label.setText("Detected Profiles: <i>None detected</i>")
            
            # Update detected markers
            if analysis['markers']:
                marker_text = "<br>".join([f"• {m['display_name']}" for m in analysis['markers']])
                self.detected_markers_label.setText(f"Detected Markers: <br>{marker_text}")
            else:
                self.detected_markers_label.setText("Detected Markers: <i>None detected</i>")
                
        except Exception as e:
            print(f"Error updating detection display: {str(e)}")
    
    def initialize_excel_file_display(self):
        """Initialize Excel file display from parent dialog"""
        try:
            excel_file = self.get_excel_file_path()
            if excel_file:
                import os
                file_name = os.path.basename(excel_file)
                self.excel_file_label.setText(f"Excel File: <b>{file_name}</b>")
                self.auto_config_status.setText("Status: <i>Excel file detected - ready for auto-detection</i>")
            else:
                self.excel_file_label.setText("Excel File: <i>Not selected</i>")
                self.auto_config_status.setText("Status: <i>No Excel file - auto-detection will prompt for file selection</i>")
        except Exception as e:
            self.excel_file_label.setText("Excel File: <i>Error detecting file</i>")
            self.auto_config_status.setText("Status: <i>Ready for auto-detection</i>")

    def browse_excel_columns(self):
        """Browse Excel file to show column structure"""
        from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTextEdit, QVBoxLayout, QDialog, QPushButton
        
        # Get Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Browse Excel File for Column Information",
            "",
            "Excel files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
            
        try:
            import pandas as pd
            
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Create information dialog
            info_dialog = QDialog(self)
            info_dialog.setWindowTitle("Excel Column Structure")
            info_dialog.setModal(True)
            info_dialog.resize(600, 400)
            
            layout = QVBoxLayout()
            
            # Create text display
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            
            # Build column information
            info_text = f"<h3>Excel File: {file_path.split('/')[-1]}</h3>"
            info_text += f"<p><b>Total Columns:</b> {len(df.columns)}</p>"
            info_text += f"<p><b>Total Rows:</b> {len(df)}</p>"
            info_text += "<hr>"
            info_text += "<h4>Column Mapping for Plot Configuration:</h4>"
            info_text += "<ul>"
            
            for i, col in enumerate(df.columns):
                if i == 0:
                    info_text += f"<li><b>Column A (Marker 1):</b> {col}</li>"
                elif i == 1:
                    info_text += f"<li><b>Column B (Marker 2):</b> {col}</li>"
                else:
                    info_text += f"<li><b>Column {chr(65+i)}:</b> {col}</li>"
            
            info_text += "</ul>"
            info_text += "<hr>"
            info_text += "<h4>Sample Data (First 5 rows):</h4>"
            info_text += "<table border='1' style='border-collapse: collapse;'>"
            
            # Add header
            info_text += "<tr>"
            for col in df.columns:
                info_text += f"<th style='padding: 4px; background-color: #f0f0f0;'>{col}</th>"
            info_text += "</tr>"
            
            # Add first 5 rows
            for idx, row in df.head(5).iterrows():
                info_text += "<tr>"
                for col in df.columns:
                    value = str(row[col]) if pd.notna(row[col]) else "N/A"
                    info_text += f"<td style='padding: 4px;'>{value}</td>"
                info_text += "</tr>"
            
            info_text += "</table>"
            
            text_edit.setHtml(info_text)
            layout.addWidget(text_edit)
            
            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(info_dialog.close)
            layout.addWidget(close_btn)
            
            info_dialog.setLayout(layout)
            info_dialog.exec_()
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not read Excel file:\n{str(e)}")

class ColumnConfigDialog(QDialog):
    """Dialog for configuring output columns"""
    
    def __init__(self, current_config, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configure Output Columns")
        self.setModal(True)
        self.setMinimumSize(600, 500)
        self.resize(700, 600)
        
        self.config = current_config.copy()
        self.init_ui()
        
    def init_ui(self):
        main_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel("""
        <b>Configure Output Columns:</b><br>
        • Check/uncheck columns to enable/disable them<br>
        • Edit column names by clicking in the 'Column Name' field<br>
        • Use 'Add Column' to create new columns<br>
        • Use 'Remove' to delete selected columns<br>
        • Distance columns will be filled with calculated distances<br>
        • Elevation columns will be filled with DEM values<br>
        • Marker columns remain empty for manual data entry
        """)
        instructions.setWordWrap(True)
        instructions.setStyleSheet(ProfessionalTheme.get_status_card_style('info'))
        main_layout.addWidget(instructions)
        
        # Create table for column configuration
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Active", "Column Name", "Type", "Actions"])
        
        # Set column widths
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        
        self.table.setColumnWidth(0, 60)   # Active checkbox
        self.table.setColumnWidth(2, 150)  # Type dropdown
        self.table.setColumnWidth(3, 80)   # Remove button
        
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        main_layout.addWidget(self.table)
        
        # Populate table
        self.populate_table()
        
        # Buttons for adding/managing columns with professional styling
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Add Column")
        add_btn.clicked.connect(self.add_column)
        add_btn.setStyleSheet(ProfessionalTheme.get_button_style('primary', 'medium'))
        button_layout.addWidget(add_btn)
        
        # Preset buttons
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("Quick Presets:"))
        
        basic_btn = QPushButton("Basic (2 cols)")
        basic_btn.clicked.connect(lambda: self.apply_preset('basic'))
        basic_btn.setToolTip("Distance_01, Height_01 only")
        preset_layout.addWidget(basic_btn)
        
        standard_btn = QPushButton("Standard (4 cols)")
        standard_btn.clicked.connect(lambda: self.apply_preset('standard'))
        standard_btn.setToolTip("Distance_01, Height_01, Distance_02, Height_02")
        preset_layout.addWidget(standard_btn)
        
        full_btn = QPushButton("Full (8 cols)")
        full_btn.clicked.connect(lambda: self.apply_preset('full'))
        full_btn.setToolTip("All 4 distance/height pairs (01-04)")
        preset_layout.addWidget(full_btn)
        
        markers_btn = QPushButton("With Markers")
        markers_btn.clicked.connect(lambda: self.apply_preset('markers'))
        markers_btn.setToolTip("Include marker columns")
        preset_layout.addWidget(markers_btn)
        
        button_layout.addLayout(preset_layout)
        button_layout.addStretch()
        
        main_layout.addLayout(button_layout)
        
        # Dialog buttons
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dialog_buttons.accepted.connect(self.accept)
        dialog_buttons.rejected.connect(self.reject)
        main_layout.addWidget(dialog_buttons)
        
        self.setLayout(main_layout)
        
    def populate_table(self):
        """Populate the table with current configuration"""
        self.table.setRowCount(len(self.config['columns']))
        
        for row, col_config in enumerate(self.config['columns']):
            # Active checkbox
            checkbox = QCheckBox()
            checkbox.setChecked(col_config['active'])
            self.table.setCellWidget(row, 0, checkbox)
            
            # Column name
            name_item = QTableWidgetItem(col_config['name'])
            self.table.setItem(row, 1, name_item)
            
            # Type dropdown
            type_combo = QComboBox()
            type_combo.addItems([
                "distance", "elevation", "marker_distance", 
                "marker_elevation", "custom", "empty"
            ])
            type_combo.setCurrentText(col_config['type'])
            self.table.setCellWidget(row, 2, type_combo)
            
            # Remove button with professional styling
            remove_btn = QPushButton("🗑️")
            remove_btn.clicked.connect(lambda checked, r=row: self.remove_column(r))
            remove_btn.setStyleSheet(ProfessionalTheme.get_button_style('danger', 'small'))
            self.table.setCellWidget(row, 3, remove_btn)
            
    def add_column(self):
        """Add a new column"""
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        
        # Active checkbox
        checkbox = QCheckBox()
        checkbox.setChecked(True)
        self.table.setCellWidget(row_count, 0, checkbox)
        
        # Column name
        name_item = QTableWidgetItem(f"Column_{row_count + 1}")
        self.table.setItem(row_count, 1, name_item)
        
        # Type dropdown
        type_combo = QComboBox()
        type_combo.addItems([
            "distance", "elevation", "marker_distance", 
            "marker_elevation", "custom", "empty"
        ])
        type_combo.setCurrentText("custom")
        self.table.setCellWidget(row_count, 2, type_combo)
        
        # Remove button with professional styling
        remove_btn = QPushButton("🗑️")
        remove_btn.clicked.connect(lambda checked, r=row_count: self.remove_column(r))
        remove_btn.setStyleSheet(ProfessionalTheme.get_button_style('danger', 'small'))
        self.table.setCellWidget(row_count, 3, remove_btn)
        
    def remove_column(self, row):
        """Remove a column"""
        if self.table.rowCount() <= 1:
            QMessageBox.warning(self, "Cannot Remove", "At least one column must remain.")
            return
            
        reply = QMessageBox.question(
            self, "Confirm Removal", 
            f"Remove column at row {row + 1}?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.table.removeRow(row)
            # Reconnect remove buttons (row indices changed)
            self.reconnect_remove_buttons()
            
    def reconnect_remove_buttons(self):
        """Reconnect remove buttons after row removal"""
        for row in range(self.table.rowCount()):
            widget = self.table.cellWidget(row, 3)
            if isinstance(widget, QPushButton):
                widget.clicked.disconnect()
                widget.clicked.connect(lambda checked, r=row: self.remove_column(r))
                
    def apply_preset(self, preset_type):
        """Apply a column preset"""
        presets = {
            'basic': [
                {'name': 'Distance_01', 'type': 'distance', 'active': True},
                {'name': 'Height_01', 'type': 'elevation', 'active': True}
            ],
            'standard': [
                {'name': 'Distance_01', 'type': 'distance', 'active': True},
                {'name': 'Height_01', 'type': 'elevation', 'active': True},
                {'name': 'Distance_02', 'type': 'distance', 'active': True},
                {'name': 'Height_02', 'type': 'elevation', 'active': True}
            ],
            'full': [
                {'name': 'Distance_01', 'type': 'distance', 'active': True},
                {'name': 'Height_01', 'type': 'elevation', 'active': True},
                {'name': 'Distance_02', 'type': 'distance', 'active': True},
                {'name': 'Height_02', 'type': 'elevation', 'active': True},
                {'name': 'Distance_03', 'type': 'distance', 'active': True},
                {'name': 'Height_03', 'type': 'elevation', 'active': True},
                {'name': 'Distance_04', 'type': 'distance', 'active': True},
                {'name': 'Height_04', 'type': 'elevation', 'active': True}
            ],
            'markers': [
                {'name': 'Distance_01', 'type': 'distance', 'active': True},
                {'name': 'Height_01', 'type': 'elevation', 'active': True},
                {'name': 'Marker_Distance', 'type': 'marker_distance', 'active': True},
                {'name': 'Marker_Height', 'type': 'marker_elevation', 'active': True},
                {'name': 'Marker_Distance_2', 'type': 'marker_distance', 'active': True},
                {'name': 'Marker_Height_2', 'type': 'marker_elevation', 'active': True}
            ]
        }
        
        if preset_type in presets:
            # Clear current table
            self.table.setRowCount(0)
            
            # Apply preset
            self.config['columns'] = presets[preset_type]
            self.populate_table()
            
    def get_config(self):
        """Get the current configuration from the table"""
        columns = []
        
        for row in range(self.table.rowCount()):
            # Get checkbox state
            checkbox = self.table.cellWidget(row, 0)
            active = checkbox.isChecked() if checkbox else False
            
            # Get column name
            name_item = self.table.item(row, 1)
            name = name_item.text() if name_item else f"Column_{row + 1}"
            
            # Get type
            type_combo = self.table.cellWidget(row, 2)
            col_type = type_combo.currentText() if type_combo else "custom"
            
            columns.append({
                'name': name,
                'type': col_type,
                'active': active
            })
            
        return {'columns': columns}

class MergeColumnConfigDialog(QDialog):
    """Dialog for configuring merge workflow column names"""
    
    def __init__(self, current_config, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configure Merge Column Names")
        self.setModal(True)
        self.setMinimumSize(600, 500)
        self.resize(700, 600)
        
        self.config = current_config.copy()
        self.init_ui()
        
    def init_ui(self):
        main_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel("""
        <b>Configure Column Names for Excel Merge:</b><br>
        • Set custom column names for intersection data processing<br>
        • Set custom column names for profile data integration<br>
        • These names will be used when reading and writing Excel files<br>
        • Make sure the names match your actual Excel file column headers
        """)
        instructions.setWordWrap(True)
        instructions.setStyleSheet(ProfessionalTheme.get_status_card_style('info'))
        main_layout.addWidget(instructions)
        
        # Create tabs for different column types
        tab_widget = QTabWidget()
        
        # Intersection columns tab
        intersection_tab = QWidget()
        intersection_layout = QVBoxLayout()
        
        intersection_layout.addWidget(QLabel("<b>Intersection Data Column Names:</b>"))
        self.intersection_edits = {}
        
        intersection_mappings = [
            ('line_name', 'Line Name Column', 'Column containing line/feature names'),
            ('distance', 'Distance Column', 'Column containing distance along line'),
            ('elevation', 'Elevation Column', 'Column containing elevation/height values'),
            ('id', 'ID Column', 'Column containing feature IDs'),
            ('layer', 'Layer Column', 'Column containing layer information'),
            ('angle', 'Angle Column', 'Column containing angle measurements')
        ]
        
        for key, label, tooltip in intersection_mappings:
            h_layout = QHBoxLayout()
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(140)
            label_widget.setToolTip(tooltip)
            h_layout.addWidget(label_widget)
            
            edit = QLineEdit(self.config['intersection_columns'].get(key, key))
            edit.setToolTip(tooltip)
            self.intersection_edits[key] = edit
            h_layout.addWidget(edit)
            
            intersection_layout.addLayout(h_layout)
        
        intersection_tab.setLayout(intersection_layout)
        tab_widget.addTab(intersection_tab, "Intersection Columns")
        
        # Profile columns tab
        profile_tab = QWidget()
        profile_layout = QVBoxLayout()
        
        profile_layout.addWidget(QLabel("<b>Profile Data Column Names:</b>"))
        self.profile_edits = {}
        
        profile_mappings = [
            ('distance_01', 'Distance 01 Column', 'First distance column'),
            ('height_01', 'Height 01 Column', 'First height/elevation column'),
            ('distance_02', 'Distance 02 Column', 'Second distance column'),
            ('height_02', 'Height 02 Column', 'Second height/elevation column'),
            ('marker_distance', 'Marker Distance Column', 'First marker distance column'),
            ('marker_height', 'Marker Height Column', 'First marker height column'),
            ('marker_distance_2', 'Marker Distance 2 Column', 'Second marker distance column'),
            ('marker_height_2', 'Marker Height 2 Column', 'Second marker height column')
        ]
        
        for key, label, tooltip in profile_mappings:
            h_layout = QHBoxLayout()
            
            label_widget = QLabel(label + ':')
            label_widget.setMinimumWidth(140)
            label_widget.setToolTip(tooltip)
            h_layout.addWidget(label_widget)
            
            edit = QLineEdit(self.config['profile_columns'].get(key, key))
            edit.setToolTip(tooltip)
            self.profile_edits[key] = edit
            h_layout.addWidget(edit)
            
            profile_layout.addLayout(h_layout)
        
        profile_tab.setLayout(profile_layout)
        tab_widget.addTab(profile_tab, "Profile Columns")
        
        main_layout.addWidget(tab_widget)
        
        # Preset buttons
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("Quick Presets:"))
        
        default_btn = QPushButton("Default Names")
        default_btn.clicked.connect(lambda: self.apply_preset('default'))
        default_btn.setToolTip("Reset to default column names")
        preset_layout.addWidget(default_btn)
        
        custom_btn = QPushButton("Custom Template")
        custom_btn.clicked.connect(lambda: self.apply_preset('custom'))
        custom_btn.setToolTip("Apply custom template with common variations")
        preset_layout.addWidget(custom_btn)
        
        preset_layout.addStretch()
        main_layout.addLayout(preset_layout)
        
        # Dialog buttons
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dialog_buttons.accepted.connect(self.accept)
        dialog_buttons.rejected.connect(self.reject)
        main_layout.addWidget(dialog_buttons)
        
        self.setLayout(main_layout)
        
    def apply_preset(self, preset_type):
        """Apply a preset configuration"""
        if preset_type == 'default':
            # Reset to default values
            default_config = {
                'intersection_columns': {
                    'line_name': 'line_name',
                    'distance': 'distance', 
                    'elevation': 'elevation',
                    'id': 'id',
                    'layer': 'layer',
                    'angle': 'angle'
                },
                'profile_columns': {
                    'distance_01': 'Distance_01',
                    'height_01': 'Height_01',
                    'distance_02': 'Distance_02', 
                    'height_02': 'Height_02',
                    'marker_distance': 'Marker_Distance',
                    'marker_height': 'Marker_Height',
                    'marker_distance_2': 'Marker_Distance_2',
                    'marker_height_2': 'Marker_Height_2'
                }
            }
        elif preset_type == 'custom':
            # Alternative common naming convention
            default_config = {
                'intersection_columns': {
                    'line_name': 'Line_Name',
                    'distance': 'Distance_m', 
                    'elevation': 'Elevation_m',
                    'id': 'Feature_ID',
                    'layer': 'Source_Layer',
                    'angle': 'Slope_Angle'
                },
                'profile_columns': {
                    'distance_01': 'Dist_1',
                    'height_01': 'Elev_1',
                    'distance_02': 'Dist_2', 
                    'height_02': 'Elev_2',
                    'marker_distance': 'Mark_Dist_1',
                    'marker_height': 'Mark_Elev_1',
                    'marker_distance_2': 'Mark_Dist_2',
                    'marker_height_2': 'Mark_Elev_2'
                }
            }
        
        # Update the text fields
        for key, value in default_config['intersection_columns'].items():
            if key in self.intersection_edits:
                self.intersection_edits[key].setText(value)
                
        for key, value in default_config['profile_columns'].items():
            if key in self.profile_edits:
                self.profile_edits[key].setText(value)
        
    def get_config(self):
        """Get the current configuration from the dialog"""
        config = {
            'intersection_columns': {},
            'profile_columns': {}
        }
        
        # Get intersection column names
        for key, edit in self.intersection_edits.items():
            config['intersection_columns'][key] = edit.text().strip()
            
        # Get profile column names  
        for key, edit in self.profile_edits.items():
            config['profile_columns'][key] = edit.text().strip()
            
        return config

class SheetParametersDialog(QDialog):
    """Dedicated dialog for viewing and editing sheet parameters"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sheet Parameters Editor")
        self.setModal(False)  # Allow working with main dialog simultaneously
        self.setMinimumSize(800, 600)
        self.resize(1000, 700)
        
        # Store reference to parent's sheet table
        self.parent_dialog = parent
        self.sheet_data = {}
        
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Info section
        info_label = QLabel("""
        <b>Sheet Parameters Editor</b><br>
        Edit parameters for each sheet in your Excel file. Changes are automatically saved back to the main dialog.<br>
        <i>Tip: You can keep this window open while working with the main plotting dialog.</i>
        """)
        info_label.setWordWrap(True)
        info_label.setStyleSheet("""
            QLabel {
                background-color: #e3f2fd;
                padding: 10px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(info_label)
        
        # Large, detailed table
        self.params_table = QTableWidget()
        self.params_table.setColumnCount(7)
        self.params_table.setHorizontalHeaderLabels([
            "Sheet Name", "Plot", "Start Slope (m)", "End Slope (m)",
            "Runout Start (m)", "Runout End (m)", "Reference Angle (°)"
        ])
        
        # Enhanced table styling for better visibility
        self.params_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #2196F3;
                background-color: white;
                alternate-background-color: #f8f9fa;
                selection-background-color: #2196F3;
                font-size: 12px;
                border: 2px solid #2196F3;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #e0e0e0;
                min-height: 20px;
            }
            QTableWidget::item:selected {
                background-color: #2196F3;
                color: white;
            }
            QHeaderView::section {
                background-color: #2196F3;
                color: white;
                padding: 12px;
                border: 1px solid #1976d2;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        
        # Better column sizing
        header = self.params_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # Sheet name
        header.setSectionResizeMode(1, QHeaderView.Fixed)    # Plot checkbox
        self.params_table.setColumnWidth(1, 80)
        for i in range(2, 7):  # Parameter columns
            header.setSectionResizeMode(i, QHeaderView.Fixed)
            self.params_table.setColumnWidth(i, 120)
        
        # Larger row height
        self.params_table.verticalHeader().setDefaultSectionSize(50)
        self.params_table.setAlternatingRowColors(True)
        
        layout.addWidget(self.params_table)
        
        # Action buttons
        button_layout = QHBoxLayout()
        
        load_btn = QPushButton("🔄 Load from Main Dialog")
        load_btn.clicked.connect(self.load_from_parent)
        load_btn.setToolTip("Load current sheet parameters from the main dialog")
        button_layout.addWidget(load_btn)
        
        save_btn = QPushButton("💾 Save to Main Dialog")
        save_btn.clicked.connect(self.save_to_parent)
        save_btn.setToolTip("Save changes back to the main dialog")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
            }
        """)
        button_layout.addWidget(save_btn)
        
        button_layout.addStretch()
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        # Status bar
        self.status_label = QLabel("Ready - Load parameters from main dialog to begin editing")
        self.status_label.setStyleSheet("""
            QLabel {
                background-color: #f5f5f5;
                padding: 8px;
                border-top: 1px solid #ddd;
                color: #666;
            }
        """)
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
        
        # Connect table changes to auto-save
        self.params_table.cellChanged.connect(self.on_cell_changed)
        
    def load_from_parent(self):
        """Load parameters from the parent dialog's sheet table"""
        if not self.parent_dialog or not hasattr(self.parent_dialog, 'sheet_table'):
            QMessageBox.warning(self, "Error", "Cannot access main dialog sheet parameters")
            return
            
        parent_table = self.parent_dialog.sheet_table
        
        # Clear current table
        self.params_table.setRowCount(0)
        self.sheet_data.clear()
        
        # Copy data from parent table
        for row in range(parent_table.rowCount()):
            self.params_table.insertRow(row)
            
            # Sheet name (column 0)
            sheet_name_item = parent_table.item(row, 0)
            if sheet_name_item:
                sheet_name = sheet_name_item.text()
                name_item = QTableWidgetItem(sheet_name)
                name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                name_item.setBackground(QColor(230, 247, 255))
                self.params_table.setItem(row, 0, name_item)
                
                # Store sheet data
                self.sheet_data[sheet_name] = {}
            
            # Plot checkbox (column 1)
            parent_widget = parent_table.cellWidget(row, 1)
            if parent_widget:
                parent_checkbox = parent_widget.findChild(QCheckBox)
                if parent_checkbox:
                    checkbox = QCheckBox()
                    checkbox.setChecked(parent_checkbox.isChecked())
                    checkbox.setStyleSheet("QCheckBox { margin: 15px; }")
                    
                    checkbox_widget = QWidget()
                    checkbox_layout = QHBoxLayout(checkbox_widget)
                    checkbox_layout.addWidget(checkbox)
                    checkbox_layout.setAlignment(Qt.AlignCenter)
                    checkbox_layout.setContentsMargins(0, 0, 0, 0)
                    
                    self.params_table.setCellWidget(row, 1, checkbox_widget)
                    
                    if sheet_name in self.sheet_data:
                        self.sheet_data[sheet_name]['plot'] = checkbox.isChecked()
            
            # Parameter columns (2-6)
            for col in range(2, 7):
                parent_item = parent_table.item(row, col)
                if parent_item:
                    value = parent_item.text()
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignCenter)
                    
                    # Add detailed tooltips
                    tooltips = [
                        "Distance where slope analysis starts (meters)",
                        "Distance where slope analysis ends (meters)", 
                        "Distance where runout analysis starts (meters)",
                        "Distance where runout analysis ends (meters)",
                        "Reference angle in degrees (measured from horizontal)"
                    ]
                    item.setToolTip(tooltips[col-2])
                    self.params_table.setItem(row, col, item)
        
        self.status_label.setText(f"Loaded {parent_table.rowCount()} sheets from main dialog")
        
    def save_to_parent(self):
        """Save parameters back to the parent dialog's sheet table"""
        if not self.parent_dialog or not hasattr(self.parent_dialog, 'sheet_table'):
            QMessageBox.warning(self, "Error", "Cannot access main dialog sheet parameters")
            return
            
        parent_table = self.parent_dialog.sheet_table
        changes_made = 0
        
        # Update parent table with changes
        for row in range(min(self.params_table.rowCount(), parent_table.rowCount())):
            # Update checkbox
            widget = self.params_table.cellWidget(row, 1)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    parent_widget = parent_table.cellWidget(row, 1)
                    if parent_widget:
                        parent_checkbox = parent_widget.findChild(QCheckBox)
                        if parent_checkbox and parent_checkbox.isChecked() != checkbox.isChecked():
                            parent_checkbox.setChecked(checkbox.isChecked())
                            changes_made += 1
            
            # Update parameter values
            for col in range(2, 7):
                item = self.params_table.item(row, col)
                parent_item = parent_table.item(row, col)
                
                if item and parent_item:
                    if item.text() != parent_item.text():
                        parent_item.setText(item.text())
                        changes_made += 1
        
        # Trigger parent updates
        if hasattr(self.parent_dialog, 'update_parameters_summary'):
            self.parent_dialog.update_parameters_summary()
        
        self.status_label.setText(f"Saved {changes_made} changes to main dialog")
        
        if changes_made > 0:
            QMessageBox.information(self, "Success", 
                f"Successfully saved {changes_made} parameter changes to the main dialog!")
        
    def on_cell_changed(self, row, column):
        """Handle cell changes with validation"""
        if column < 2:  # Skip sheet name and checkbox columns
            return
            
        item = self.params_table.item(row, column)
        if not item:
            return
            
        try:
            value = float(item.text())
            
            # Validate ranges
            if column in [2, 3, 4, 5] and value < 0:  # Distance parameters
                item.setBackground(QColor(255, 200, 200))
                item.setToolTip("Distance values cannot be negative")
            else:
                item.setBackground(QColor(255, 255, 255))
                
        except ValueError:
            item.setBackground(QColor(255, 200, 200))
            item.setToolTip("Please enter a valid number")

class CombinedGeospatialToolDialog(QDialog):
    """Main dialog combining all geospatial tools"""
    
    def __init__(self, iface, parent=None):
        super().__init__(parent)
        self.iface = iface  # Store the QGIS interface reference
        self.setWindowTitle("QGIS section drafter")
        self.setModal(False)
        
        # Apply professional theme
        self.apply_professional_theme()
        
        # More manageable window sizing - smaller initial size but expandable
        self.setMinimumSize(1000, 650)  # Reduced minimum size
        self.resize(1200, 750)  # Smaller initial size, still good for plot viewing
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Set global processing settings to skip invalid features
        from qgis.core import QgsSettings
        settings = QgsSettings()
        settings.setValue("Processing/Configuration/INVALID_FEATURES_FILTERING", 1)  # Skip invalid features
        
        # Initialize column configuration
        self.column_config = self.load_default_column_config()
        
        # Initialize merge column configuration
        self.merge_column_config = self.load_default_merge_column_config()
        
        # Center the window on screen
        self.center_on_screen()
        
        self.map_tool = None
        self.temp_polylines = []
        self.settings = QSettings("QGIS", "QgisSectionDrafter")
        self.last_directory = self.settings.value("lastDirectory", os.path.expanduser("~"))
        self.plot_config = self.load_plot_config()
        self.plot_display_widget = None
        self.sheet_params_dialog = None  # Initialize sheet parameters dialog
        self.init_ui()
        
    def apply_professional_theme(self):
        """Apply professional styling to the dialog"""
        # Set the dialog's base style with professional theme
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {ProfessionalTheme.COLORS['background']};
                color: {ProfessionalTheme.COLORS['on_surface']};
                font-family: "Segoe UI", "Roboto", "Helvetica Neue", Arial, sans-serif;
            }}
            
            {ProfessionalTheme.get_button_style('primary')}
            {ProfessionalTheme.get_input_style()}
            {ProfessionalTheme.get_tab_style()}
            {ProfessionalTheme.get_group_box_style()}
            {ProfessionalTheme.get_progress_style()}
            
            /* Enhanced scrollbar styling */
            QScrollBar:vertical {{
                background-color: {ProfessionalTheme.COLORS['surface_variant']};
                width: 12px;
                border-radius: 6px;
                margin: 0;
            }}
            QScrollBar::handle:vertical {{
                background-color: {ProfessionalTheme.COLORS['on_surface_variant']};
                border-radius: 6px;
                min-height: 20px;
                margin: 2px;
            }}
            QScrollBar::handle:vertical:hover {{
                background-color: {ProfessionalTheme.COLORS['primary']};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            
            /* Enhanced text editing */
            QTextEdit {{
                border: 2px solid {ProfessionalTheme.COLORS['outline']};
                border-radius: {ProfessionalTheme.RADIUS['md']}px;
                background-color: {ProfessionalTheme.COLORS['surface']};
                color: {ProfessionalTheme.COLORS['on_surface']};
                padding: {ProfessionalTheme.SPACING['sm']}px;
            }}
            QTextEdit:focus {{
                border-color: {ProfessionalTheme.COLORS['primary']};
            }}
            
            /* Enhanced table styling */
            QTableWidget {{
                border: 1px solid {ProfessionalTheme.COLORS['outline']};
                border-radius: {ProfessionalTheme.RADIUS['md']}px;
                background-color: {ProfessionalTheme.COLORS['surface']};
                gridline-color: {ProfessionalTheme.COLORS['outline']};
                selection-background-color: {ProfessionalTheme.COLORS['success_bg']};
            }}
            QTableWidget::item {{
                padding: {ProfessionalTheme.SPACING['sm']}px;
                border-bottom: 1px solid {ProfessionalTheme.COLORS['outline']};
            }}
            QTableWidget::item:selected {{
                background-color: {ProfessionalTheme.COLORS['primary']};
                color: {ProfessionalTheme.COLORS['on_primary']};
            }}
            QHeaderView::section {{
                background-color: {ProfessionalTheme.COLORS['surface_variant']};
                color: {ProfessionalTheme.COLORS['on_surface']};
                padding: {ProfessionalTheme.SPACING['sm']}px;
                border: 1px solid {ProfessionalTheme.COLORS['outline']};
                font-weight: 600;
            }}
        """)

    def center_on_screen(self):
        """Center the dialog on the screen"""
        try:
            from PyQt5.QtWidgets import QDesktopWidget
            screen = QDesktopWidget().screenGeometry()
            size = self.geometry()
            self.move(
                (screen.width() - size.width()) // 2,
                (screen.height() - size.height()) // 2
            )
        except:
            pass  # Fallback if centering fails
            
    def load_default_column_config(self):
        """Load default column configuration"""
        return {
            'columns': [
                {'name': 'Distance_01', 'type': 'distance', 'active': True},
                {'name': 'Height_01', 'type': 'elevation', 'active': True},
                {'name': 'Distance_02', 'type': 'distance', 'active': False},
                {'name': 'Height_02', 'type': 'elevation', 'active': False},
                {'name': 'Distance_03', 'type': 'distance', 'active': False},
                {'name': 'Height_03', 'type': 'elevation', 'active': False},
                {'name': 'Distance_04', 'type': 'distance', 'active': False},
                {'name': 'Height_04', 'type': 'elevation', 'active': False},
                {'name': 'Marker_Distance', 'type': 'marker_distance', 'active': True},
                {'name': 'Marker_Height', 'type': 'marker_elevation', 'active': True},
                {'name': 'Marker_Distance_2', 'type': 'marker_distance', 'active': True},
                {'name': 'Marker_Height_2', 'type': 'marker_elevation', 'active': True}
            ]
        }

    def update_column_config_display(self):
        """Update the display showing current column configuration"""
        active_columns = [col for col in self.column_config['columns'] if col['active']]
        if active_columns:
            column_names = [col['name'] for col in active_columns]
            display_text = f"Active Columns ({len(active_columns)}): {', '.join(column_names)}"
        else:
            display_text = "No active columns configured"
        
        self.column_config_display.setText(display_text)

    def configure_output_columns(self):
        """Open column configuration dialog"""
        dialog = ColumnConfigDialog(self.column_config, self)
        if dialog.exec_():
            self.column_config = dialog.get_config()
            self.update_column_config_display()
            QMessageBox.information(self, "Success", "Column configuration updated successfully!")

    def load_default_merge_column_config(self):
        """Load default column configuration for merge workflow"""
        return {
            'intersection_columns': {
                'line_name': 'line_name',
                'distance': 'distance', 
                'elevation': 'elevation',
                'id': 'id',
                'layer': 'layer',
                'angle': 'angle'
            },
            'profile_columns': {
                'distance_01': 'Distance_01',
                'height_01': 'Height_01',
                'distance_02': 'Distance_02', 
                'height_02': 'Height_02',
                'marker_distance': 'Marker_Distance',
                'marker_height': 'Marker_Height',
                'marker_distance_2': 'Marker_Distance_2',
                'marker_height_2': 'Marker_Height_2'
            }
        }

    def update_merge_column_display(self):
        """Update the display showing current merge column configuration"""
        intersection_cols = list(self.merge_column_config['intersection_columns'].values())
        profile_cols = list(self.merge_column_config['profile_columns'].values())
        
        display_text = f"Intersection Columns: {', '.join(intersection_cols[:3])}...\n"
        display_text += f"Profile Columns: {', '.join(profile_cols[:4])}..."
        
        self.merge_column_display.setText(display_text)

    def configure_merge_column_names(self):
        """Open merge column configuration dialog"""
        dialog = MergeColumnConfigDialog(self.merge_column_config, self)
        if dialog.exec_():
            self.merge_column_config = dialog.get_config()
            self.update_merge_column_display()
            QMessageBox.information(self, "Success", "Merge column configuration updated successfully!")
        
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(ProfessionalTheme.SPACING['md'])  # Professional spacing
        main_layout.setContentsMargins(
            ProfessionalTheme.SPACING['lg'], 
            ProfessionalTheme.SPACING['lg'], 
            ProfessionalTheme.SPACING['lg'], 
            ProfessionalTheme.SPACING['lg']
        )  # Professional margins
        
        # Add warning banner with professional styling
        warning_group = QGroupBox("⚠️ Important - Save Your Work First!")
        warning_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        warning_group.setStyleSheet(ProfessionalTheme.get_status_card_style('warning'))
        warning_layout = QVBoxLayout()
        warning_layout.setSpacing(ProfessionalTheme.SPACING['sm'])  # Professional spacing
        warning_layout.setContentsMargins(
            ProfessionalTheme.SPACING['md'], 
            ProfessionalTheme.SPACING['md'], 
            ProfessionalTheme.SPACING['md'], 
            ProfessionalTheme.SPACING['md']
        )  # Professional margins
        
        warning_text = QTextEdit()
        warning_text.setReadOnly(True)
        warning_text.setMinimumHeight(80)  # Reduced from 100
        warning_text.setMaximumHeight(120)  # Reduced from 150
        warning_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        warning_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: transparent;
                border: none;
                font-size: {ProfessionalTheme.FONTS['body_large']['size']}px;
                color: {ProfessionalTheme.COLORS['on_surface']};
            }}
        """)
        warning_text.setHtml(f"""
            <b>Critical Steps:</b><br>
            1. <span style='color: {ProfessionalTheme.COLORS['error']}; font-weight: bold;'>SAVE PROJECT</span> (Ctrl+S)<br>
            2. Ensure layers are saved<br>
            3. Check DEM/raster is loaded<br>
            4. Test with small datasets first
        """)
        warning_layout.addWidget(warning_text)
        
        # Add compact save project button with professional styling
        save_project_btn = QPushButton("💾 Save Project")
        save_project_btn.setStyleSheet(ProfessionalTheme.get_button_style('danger', 'medium'))
        save_project_btn.clicked.connect(self.save_project)
        warning_layout.addWidget(save_project_btn)
        
        warning_group.setLayout(warning_layout)
        main_layout.addWidget(warning_group)
        
        # Create tab widget with improved sizing
        self.tabs = QTabWidget()
        self.tabs.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Set tab position and add scroll if needed
        self.tabs.setTabPosition(QTabWidget.North)
        self.tabs.setMovable(False)
        self.tabs.setUsesScrollButtons(True)
        
        # Tab 1: Polyline Creation
        self.polyline_tab = QWidget()
        self.setup_polyline_tab()
        self.polyline_tab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tabs.addTab(self.polyline_tab, "1. Create Polylines")
        
        # Tab 2: Profile Extraction
        self.profile_tab = QWidget()
        self.setup_profile_tab()
        self.profile_tab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tabs.addTab(self.profile_tab, "2. Extract Profiles")
        
        # Tab 3: Intersection Points
        self.intersection_tab = QWidget()
        self.setup_intersection_tab()
        self.intersection_tab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tabs.addTab(self.intersection_tab, "3. Intersection Points")
        
        # Tab 4: Excel Merge Workflow
        self.excel_merge_tab = QWidget()
        self.setup_excel_merge_tab()
        self.excel_merge_tab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tabs.addTab(self.excel_merge_tab, "4. Excel Merge Workflow")
        
        # Tab 5: Profile Plotting (moved to last position)
        if MATPLOTLIB_AVAILABLE:
            self.plotting_tab = QWidget()
            self.setup_plotting_tab()
            self.plotting_tab.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.tabs.addTab(self.plotting_tab, "5. Plot Profiles")
        
        main_layout.addWidget(self.tabs)
        
        # Professional button layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(ProfessionalTheme.SPACING['md'])  # Professional spacing
        
        # Add a help button with professional styling
        help_btn = QPushButton("❓ Help")
        help_btn.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        help_btn.clicked.connect(self.show_help)
        help_btn.setStyleSheet(ProfessionalTheme.get_button_style('text', 'medium'))
        button_layout.addWidget(help_btn)
        
        button_layout.addStretch()
        
        # Professional close button
        close_btn = QPushButton("✖️ Close")
        close_btn.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        close_btn.clicked.connect(self.close)
        close_btn.setStyleSheet(ProfessionalTheme.get_button_style('secondary', 'medium'))
        button_layout.addWidget(close_btn)
        
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
        
    def setup_excel_merge_tab(self):
        """Setup the Excel merge workflow tab (formerly IGT_excel_merge.py)"""
        layout = QVBoxLayout()

        info_label = QLabel("<b>Excel Merge Workflow:</b><br>\n"
            "1. Separate intersection points by line name into individual sheets.<br>\n"
            "2. Optionally merge intersection data into elevation profiles.<br>\n"
            "3. Customize column names for output files.<br>\n"
            "<i>All dialogs and file selection use PyQt5 for consistency.</i>")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # Column Configuration Section
        column_config_group = QGroupBox("Column Name Configuration")
        column_config_layout = QVBoxLayout()
        
        # Button to configure column names with professional styling
        self.config_merge_columns_btn = QPushButton("⚙️ Configure Column Names")
        self.config_merge_columns_btn.clicked.connect(self.configure_merge_column_names)
        self.config_merge_columns_btn.setStyleSheet(ProfessionalTheme.get_button_style('primary', 'medium'))
        column_config_layout.addWidget(self.config_merge_columns_btn)
        
        # Display current column configuration
        self.merge_column_display = QLabel()
        self.merge_column_display.setWordWrap(True)
        self.merge_column_display.setStyleSheet("""
            QLabel {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-family: monospace;
                font-size: 10px;
            }
        """)
        column_config_layout.addWidget(self.merge_column_display)
        
        column_config_group.setLayout(column_config_layout)
        layout.addWidget(column_config_group)

        # File Selection Section
        file_group = QGroupBox("File Selection")
        file_layout = QVBoxLayout()
        
        # Intersection file selection
        file_layout.addWidget(QLabel("Intersection Points File:"))
        self.merge_intersection_edit = QLineEdit()
        self.merge_intersection_edit.setReadOnly(True)
        browse_intersection_btn = QPushButton("Browse Intersection Points File")
        browse_intersection_btn.clicked.connect(self.browse_merge_intersection_file)
        file_layout.addWidget(self.merge_intersection_edit)
        file_layout.addWidget(browse_intersection_btn)

        # Elevation profiles file selection
        file_layout.addWidget(QLabel("Elevation Profiles File:"))
        self.merge_profiles_edit = QLineEdit()
        self.merge_profiles_edit.setReadOnly(True)
        browse_profiles_btn = QPushButton("Browse Elevation Profiles File")
        browse_profiles_btn.clicked.connect(self.browse_merge_profiles_file)
        file_layout.addWidget(self.merge_profiles_edit)
        file_layout.addWidget(browse_profiles_btn)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # Run workflow button with professional styling
        run_workflow_btn = QPushButton("🚀 Run Excel Merge Workflow")
        run_workflow_btn.clicked.connect(self.run_excel_merge_workflow)
        run_workflow_btn.setStyleSheet(ProfessionalTheme.get_button_style('primary', 'large'))
        layout.addWidget(run_workflow_btn)

        # Wrap in scroll area
        _scroll = QScrollArea()
        _scroll.setWidgetResizable(True)
        _scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _content = QWidget()
        _content.setLayout(layout)
        _scroll.setWidget(_content)
        _outer = QVBoxLayout()
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.addWidget(_scroll)
        self.excel_merge_tab.setLayout(_outer)
        
        # Initialize the column display
        self.update_merge_column_display()

    def browse_merge_intersection_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Intersection Points Excel File",
            self.last_directory,
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        if file_path:
            self.merge_intersection_edit.setText(file_path)
            self.last_directory = os.path.dirname(file_path)
            self.settings.setValue("lastDirectory", self.last_directory)

    def browse_merge_profiles_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Elevation Profiles Excel File",
            self.last_directory,
            "Excel Files (*.xlsx *.xls);;All Files (*.*)"
        )
        if file_path:
            self.merge_profiles_edit.setText(file_path)
            self.last_directory = os.path.dirname(file_path)
            self.settings.setValue("lastDirectory", self.last_directory)

    def run_excel_merge_workflow(self):
        """Run the merged Excel workflow using custom column names"""
        try:
            intersection_file = self.merge_intersection_edit.text()
            profiles_file = self.merge_profiles_edit.text()
            
            if not intersection_file:
                QMessageBox.warning(self, "No File", "Please select the intersection points Excel file.")
                return
                
            # Get configured column names
            col_config = self.merge_column_config
            line_name_col = col_config['intersection_columns']['line_name']
            distance_col = col_config['intersection_columns']['distance']
            elevation_col = col_config['intersection_columns']['elevation']
            
            # Read intersection data using configured column names
            df = pd.read_excel(intersection_file)
            
            if line_name_col not in df.columns:
                QMessageBox.critical(self, "Column Error", 
                    f"The file must contain a '{line_name_col}' column.\n\n"
                    f"Current file columns: {', '.join(df.columns)}\n\n"
                    f"Use 'Configure Column Names' to set the correct column name.")
                return
                
            # Check for other required columns
            missing_cols = []
            for col_key, col_name in col_config['intersection_columns'].items():
                if col_name and col_name not in df.columns:
                    missing_cols.append(f"{col_key} ('{col_name}')")
            
            if missing_cols:
                reply = QMessageBox.question(
                    self, "Missing Columns",
                    f"Some configured columns were not found:\n{', '.join(missing_cols)}\n\n"
                    f"Continue anyway? (Missing columns will be skipped)",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return
            
            # Process data using configured column names
            unique_names = df[line_name_col].dropna().unique()
            sorted_names = sorted(unique_names, key=lambda x: int(re.findall(r'\d+', str(x))[0]) if re.findall(r'\d+', str(x)) else 0)
            polyline_data = {name: df[df[line_name_col] == name].copy() for name in sorted_names}
            
            # Save separated data
            output_dir = Path(intersection_file).parent
            base_name = Path(intersection_file).stem
            separated_file = output_dir / f"{base_name}_by_line_name.xlsx"
            
            with pd.ExcelWriter(separated_file, engine='openpyxl') as writer:
                for name in sorted_names:
                    sheet_name = str(name).replace('/', '_').replace('\\', '_')[:31]
                    polyline_data[name].to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Continue with profile integration if requested
            reply = QMessageBox.question(
                self, "Continue Process",
                f"Step 1 Complete: Data separated into {len(sorted_names)} sheets.\n\n"
                f"Do you want to continue and add intersection data to elevation profiles?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                QMessageBox.information(self, "Complete", f"Intersection data separated and saved to {separated_file}.")
                return
                
            if not profiles_file:
                QMessageBox.warning(self, "Partial Complete", 
                    "Elevation profiles file not selected. Intersection data has been separated but not added to profiles.")
                return
            
            # Process profiles using configured column names
            profile_col_config = col_config['profile_columns']
            distance_01_col = profile_col_config['distance_01']
            marker_dist_col = profile_col_config['marker_distance']
            marker_height_col = profile_col_config['marker_height']
            marker_dist_2_col = profile_col_config['marker_distance_2']
            marker_height_2_col = profile_col_config['marker_height_2']
            
            profile_sheets = pd.read_excel(profiles_file, sheet_name=None)
            processed_count = 0
            
            # Statistics tracking
            total_intersection_points = 0
            total_placed_points = 0
            total_skipped_points = 0
            
            for name in sorted_names:
                sheet_name = str(name).replace('/', '_').replace('\\', '_')[:31]
                
                if sheet_name in profile_sheets:
                    intersection_data = polyline_data[name]
                    profile_data = profile_sheets[sheet_name].copy()
                    
                    # Check for required columns in intersection data
                    if distance_col not in intersection_data.columns or elevation_col not in intersection_data.columns:
                        continue
                        
                    # Check for required columns in profile data
                    if distance_01_col not in profile_data.columns:
                        QMessageBox.warning(self, "Column Missing", 
                            f"Profile sheet '{sheet_name}' missing '{distance_01_col}' column. Skipping.")
                        continue
                    
                    # Add marker columns if they don't exist and initialize them properly
                    for col in [marker_dist_col, marker_height_col, marker_dist_2_col, marker_height_2_col]:
                        if col not in profile_data.columns:
                            profile_data[col] = pd.NA  # Use pandas NA instead of np.nan
                    
                    # ENHANCED: Comprehensive duplicate prevention system
                    used_profile_rows = set()  # Track which profile rows have been used
                    intersection_point_assignments = {}  # Track what was assigned where
                    
                    # Sort intersection points by distance for consistent processing
                    intersection_data_sorted = intersection_data.sort_values(by=distance_col)
                    
                    QgsMessageLog.logMessage(
                        f"Processing {len(intersection_data_sorted)} intersection points for sheet '{sheet_name}'", 
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                    
                    for idx, row in intersection_data_sorted.iterrows():
                        distance = row.get(distance_col)
                        elevation = row.get(elevation_col)
                        total_intersection_points += 1
                        
                        # Comprehensive validation of intersection point data
                        if pd.isna(distance) or pd.isna(elevation) or distance <= 0 or elevation <= 0:
                            QgsMessageLog.logMessage(
                                f"Skipping invalid intersection point: distance={distance}, elevation={elevation}", 
                                "IntegratedProfileAnalyzer", Qgis.Warning
                            )
                            total_skipped_points += 1
                            continue
                        
                        # Get valid profile distances
                        profile_distances = profile_data[distance_01_col].dropna()
                        if len(profile_distances) == 0:
                            QgsMessageLog.logMessage(
                                f"No valid profile distances found in '{distance_01_col}' column", 
                                "IntegratedProfileAnalyzer", Qgis.Warning
                            )
                            total_skipped_points += 1
                            continue
                        
                        # Find closest profile row
                        differences = np.abs(profile_distances - distance)
                        closest_idx = differences.idxmin()
                        closest_distance_diff = differences.min()
                        
                        # ENHANCED: Check if this profile row has already been used
                        if closest_idx in used_profile_rows:
                            # Find alternative unused rows within tolerance
                            tolerance = 2.0  # 2 meter tolerance for alternative rows
                            sorted_diffs = differences.sort_values()
                            alternative_found = False
                            
                            for alt_idx in sorted_diffs.index:
                                if (alt_idx not in used_profile_rows and 
                                    sorted_diffs[alt_idx] <= (closest_distance_diff + tolerance)):
                                    closest_idx = alt_idx
                                    closest_distance_diff = sorted_diffs[alt_idx]
                                    alternative_found = True
                                    QgsMessageLog.logMessage(
                                        f"Using alternative profile row {alt_idx} (distance diff: {closest_distance_diff:.2f}m)", 
                                        "IntegratedProfileAnalyzer", Qgis.Info
                                    )
                                    break
                            
                            if not alternative_found:
                                QgsMessageLog.logMessage(
                                    f"Skipping intersection point at distance {distance}m - no unused profile row available within tolerance", 
                                    "IntegratedProfileAnalyzer", Qgis.Warning
                                )
                                total_skipped_points += 1
                                continue
                        
                        # ENHANCED: Robust empty value detection for marker columns
                        use_marker_dist = None
                        use_marker_height = None
                        
                        # Check first marker column pair with enhanced empty detection
                        marker_1_dist_val = profile_data.loc[closest_idx, marker_dist_col]
                        marker_1_height_val = profile_data.loc[closest_idx, marker_height_col]
                        
                        # More robust checking for empty values
                        marker_1_empty = (pd.isna(marker_1_dist_val) or 
                                         marker_1_dist_val == '' or 
                                         marker_1_dist_val == 0 or
                                         str(marker_1_dist_val).strip() == '')
                        
                        if marker_1_empty:
                            use_marker_dist = marker_dist_col
                            use_marker_height = marker_height_col
                            marker_pair = "1"
                        else:
                            # Check second marker column pair with enhanced empty detection
                            marker_2_dist_val = profile_data.loc[closest_idx, marker_dist_2_col]
                            marker_2_height_val = profile_data.loc[closest_idx, marker_height_2_col]
                            
                            marker_2_empty = (pd.isna(marker_2_dist_val) or 
                                             marker_2_dist_val == '' or 
                                             marker_2_dist_val == 0 or
                                             str(marker_2_dist_val).strip() == '')
                            
                            if marker_2_empty:
                                use_marker_dist = marker_dist_2_col
                                use_marker_height = marker_height_2_col
                                marker_pair = "2"
                            else:
                                # Both marker columns are occupied - skip this intersection point
                                QgsMessageLog.logMessage(
                                    f"Skipping intersection point at distance {distance}m - profile row {closest_idx} has both marker columns occupied (Marker1: {marker_1_dist_val}, Marker2: {marker_2_dist_val})", 
                                    "IntegratedProfileAnalyzer", Qgis.Warning
                                )
                                total_skipped_points += 1
                                continue
                        
                        # Place the intersection point only if we found available columns
                        if use_marker_dist and use_marker_height:
                            profile_data.loc[closest_idx, use_marker_dist] = distance
                            profile_data.loc[closest_idx, use_marker_height] = elevation
                            used_profile_rows.add(closest_idx)  # Mark this row as used
                            total_placed_points += 1
                            
                            # Track assignment for debugging
                            assignment_key = f"row_{closest_idx}_marker_{marker_pair}"
                            intersection_point_assignments[assignment_key] = {
                                'distance': distance,
                                'elevation': elevation,
                                'distance_diff': closest_distance_diff
                            }
                            
                            QgsMessageLog.logMessage(
                                f"✓ Placed intersection point: distance={distance}m, elevation={elevation}m → profile row {closest_idx}, marker pair {marker_pair} (diff: {closest_distance_diff:.2f}m)", 
                                "IntegratedProfileAnalyzer", Qgis.Info
                            )
                    
                    profile_sheets[sheet_name] = profile_data
                    processed_count += 1
                    
                    # Log summary for this sheet
                    sheet_placed = len([k for k in intersection_point_assignments.keys() if k.startswith(f'row_')])
                    QgsMessageLog.logMessage(
                        f"Sheet '{sheet_name}': {sheet_placed} intersection points placed, {len(used_profile_rows)} profile rows used", 
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
            
            # Save updated profiles
            if processed_count > 0:
                with pd.ExcelWriter(profiles_file, engine='openpyxl') as writer:
                    for sheet_name, data in profile_sheets.items():
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Enhanced completion summary with comprehensive statistics
            summary = (f"Excel Merge Workflow Complete!\n\n"
                      f"📊 Summary Statistics:\n"
                      f"✓ Intersection data separated into {len(sorted_names)} sheets\n"
                      f"✓ {processed_count} elevation profiles updated\n"
                      f"✓ {total_placed_points} intersection points placed successfully\n"
                      f"⚠ {total_skipped_points} intersection points skipped\n"
                      f"📋 {total_intersection_points} total intersection points processed\n\n"
                      f"📁 Files saved:\n"
                      f"- Intersection data: {separated_file.name}\n"
                      f"- Updated profiles: {Path(profiles_file).name}\n\n"
                      f"⚙️ Column configuration used:\n"
                      f"- Line name: '{line_name_col}'\n"
                      f"- Distance: '{distance_col}'\n"
                      f"- Elevation: '{elevation_col}'\n\n"
                      f"� Duplicate Prevention Status: ACTIVE\n"
                      f"Each intersection point is placed only once in the closest available profile row.\n"
                      f"No duplicates should appear in marker columns.\n\n"
                      f"💡 Processing Details:\n"
                      f"- Used enhanced empty value detection\n"
                      f"- Applied 2.0m tolerance for alternative row selection\n"
                      f"- Sorted intersection points by distance for consistency\n"
                      f"- Prevented any overwriting of existing marker data")
            
            QMessageBox.information(self, "Complete", summary)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
            QgsMessageLog.logMessage(f"Excel merge workflow error: {str(e)}", "IntegratedProfileAnalyzer", Qgis.Warning)
            import traceback
            QgsMessageLog.logMessage(traceback.format_exc(), "IntegratedProfileAnalyzer", Qgis.Warning)
        
    def load_plot_config(self):
        """Load plotting configuration"""
        config = {
            'marker_size': '8',
            'marker_style': 'o',
            'figure_size': '10',
            'slope_line_color': '#FF0000',
            'reference_line_color': '#800080',
            'main_line_color': '#000000',
            'feature_line_color': '#0000FF',
            'runout_line_color': '#008000',
            'marker_1_color': '#0070FF',
            'marker_2_color': '#0000FF',
            # New marker style configurations
            'marker_1_marker_style': 'o',
            'marker_2_marker_style': 's',
            'main_profile_marker_style': 'o',
            'marker_1_size': '8',
            'marker_2_size': '8',
            'main_profile_size': '4',
            'title_font_size': '14',
            'label_font_size': '12',
            'legend_font_size': '10',
            'text_font_size': '10',
            'legend_main_profile': 'Existing Ground Profile',
            'legend_markers_1': 'Markers 1',
            'legend_markers_2': 'Markers 2',
            'legend_feature': 'Feature',
            'legend_angular_elevation': 'Angular Elevation'
        }
        return config
        
    def setup_plotting_tab(self):
        """Setup the profile plotting tab"""
        main_layout = QVBoxLayout()
        
        # Create horizontal splitter for controls and display
        h_layout = QHBoxLayout()
        
        # Left side - controls
        controls_widget = QWidget()
        controls_layout = QVBoxLayout()
        controls_layout.setSpacing(4)
        controls_widget.setLayout(controls_layout)
        controls_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        controls_widget.setMinimumWidth(320)
        # Ensure labels inside HBox rows don't steal space from combos/edits
        controls_widget.setStyleSheet(
            "QGroupBox { font-weight: bold; } "
            "QLabel { min-width: 0px; } "
            "QComboBox { min-width: 80px; } "
        )
        
        # File selection
        file_group = QGroupBox("Excel File Selection")
        file_layout = QVBoxLayout()
        
        # Excel file selection
        excel_layout = QHBoxLayout()
        excel_layout.addWidget(QLabel("Excel File:"))
        self.plot_excel_edit = QLineEdit()
        self.plot_excel_edit.setReadOnly(True)
        excel_layout.addWidget(self.plot_excel_edit)
        
        browse_excel_btn = QPushButton("Browse...")
        browse_excel_btn.clicked.connect(self.browse_plot_excel)
        excel_layout.addWidget(browse_excel_btn)
        file_layout.addLayout(excel_layout)
        
        file_group.setLayout(file_layout)
        controls_layout.addWidget(file_group)
        
        # Sheet parameters
        sheet_group = QGroupBox("Sheet Parameters")
        sheet_layout = QVBoxLayout()
        
        # Add refresh sheets button
        refresh_sheets_btn = QPushButton("Load/Refresh Sheets")
        refresh_sheets_btn.clicked.connect(self.load_excel_sheets)
        sheet_layout.addWidget(refresh_sheets_btn)
        
        # Create table for sheet parameters with enhanced visibility
        self.sheet_table = QTableWidget()
        self.sheet_table.setColumnCount(7)
        self.sheet_table.setHorizontalHeaderLabels([
            "Sheet Name", "Plot", "Start Slope (m)", "End Slope (m)",
            "Runout Start (m)", "Runout End (m)", "Reference Angle (°)"
        ])
        
        # Enhanced table visibility and sizing
        self.sheet_table.horizontalHeader().setStretchLastSection(True)
        self.sheet_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.sheet_table.setMinimumHeight(200)  # Increased from 120
        self.sheet_table.setMaximumHeight(400)  # Allow more space
        self.sheet_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.sheet_table.setAlternatingRowColors(True)

        # Enhanced styling for better visibility
        self.sheet_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #8faadb;
                background-color: white;
                alternate-background-color: #f0f8ff;
                selection-background-color: #4472c4;
                font-size: 11px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ddd;
            }
            QTableWidget::item:selected {
                background-color: #4472c4;
                color: white;
            }
            QHeaderView::section {
                background-color: #4472c4;
                color: white;
                padding: 8px;
                border: 1px solid #2e5bb8;
                font-weight: bold;
            }
        """)

        # Set column widths for better visibility
        header = self.sheet_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # Sheet name stretches
        header.setSectionResizeMode(1, QHeaderView.Fixed)    # Plot checkbox fixed
        self.sheet_table.setColumnWidth(1, 60)  # Plot column
        for i in range(2, 7):  # Parameter columns
            header.setSectionResizeMode(i, QHeaderView.Fixed)
            self.sheet_table.setColumnWidth(i, 100)

        # Row height for better readability
        self.sheet_table.verticalHeader().setDefaultSectionSize(35)
        self.sheet_table.verticalHeader().setVisible(True)
        
        sheet_layout.addWidget(self.sheet_table)
        
        # Add summary display below table
        self.sheet_summary_label = QLabel("No sheets loaded")
        self.sheet_summary_label.setStyleSheet("""
            QLabel {
                background-color: #e8f4fd;
                padding: 8px;
                border: 1px solid #4472c4;
                border-radius: 4px;
                font-weight: bold;
                color: #2e5bb8;
            }
        """)
        sheet_layout.addWidget(self.sheet_summary_label)
        
        # Add button to open dedicated sheet parameters dialog
        open_params_btn = QPushButton("📋 Open Sheet Parameters Editor")
        open_params_btn.clicked.connect(self.open_sheet_parameters_dialog)
        open_params_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        open_params_btn.setToolTip("Open a dedicated window for viewing and editing all sheet parameters with better visibility")
        sheet_layout.addWidget(open_params_btn)
        
        sheet_group.setLayout(sheet_layout)
        controls_layout.addWidget(sheet_group)

        # ── Drill Hole Overlay ─────────────────────────────────────────────
        dh_group = QGroupBox("Drill Hole Overlay")
        dh_layout = QVBoxLayout()

        # Enable checkbox
        self.dh_enable_chk = QCheckBox("Plot drill holes on cross-sections")
        self.dh_enable_chk.setToolTip(
            "When enabled, drill hole collars and traces will be projected\n"
            "onto each cross-section and plotted."
        )
        dh_layout.addWidget(self.dh_enable_chk)

        # Drill log folder picker
        dh_file_layout = QHBoxLayout()
        dh_file_layout.addWidget(QLabel("Drill Log Folder:"))
        self.dh_folder_edit = QLineEdit()
        self.dh_folder_edit.setReadOnly(True)
        self.dh_folder_edit.setPlaceholderText("Folder containing per-hole .xlsx drill log files")
        dh_file_layout.addWidget(self.dh_folder_edit)
        dh_browse_btn = QPushButton("Browse…")
        dh_browse_btn.clicked.connect(self.browse_drillhole_folder)
        dh_file_layout.addWidget(dh_browse_btn)
        dh_layout.addLayout(dh_file_layout)

        # Collar coordinate CRS
        dh_crs_layout = QHBoxLayout()
        dh_crs_layout.addWidget(QLabel("Collar CRS:"))
        self.dh_crs_combo = QComboBox()
        self.dh_crs_combo.setToolTip(
            "Coordinate reference system of the Easting/Northing values in the CSV.\n"
            "Must match the projection used when recording drill hole collars."
        )
        for label, epsg in [
            ("MGA Zone 50 – GDA94  (EPSG:28350)", "EPSG:28350"),
            ("MGA Zone 51 – GDA94  (EPSG:28351)", "EPSG:28351"),
            ("UTM Zone 50S – WGS84 (EPSG:32750)", "EPSG:32750"),
            ("UTM Zone 51S – WGS84 (EPSG:32751)", "EPSG:32751"),
            ("GDA2020 / MGA Zone 50 (EPSG:7850)", "EPSG:7850"),
            ("GDA2020 / MGA Zone 51 (EPSG:7851)", "EPSG:7851"),
        ]:
            self.dh_crs_combo.addItem(label, epsg)
        dh_crs_layout.addWidget(self.dh_crs_combo)
        dh_layout.addLayout(dh_crs_layout)

        # LiDAR / DEM raster picker for automatic collar RL
        dh_raster_layout = QHBoxLayout()
        dh_raster_layout.addWidget(QLabel("LiDAR / DEM:"))
        self.dh_raster_combo = QComboBox()
        self.dh_raster_combo.setToolTip(
            "Select the LiDAR/DEM raster to use for collar RL (elevation).\n"
            "The collar's Easting/Northing will be sampled from this raster.\n"
            "If the CSV already has an RL column it will be used unless\n"
            "'Always use raster RL' is ticked."
        )
        self.refresh_dh_raster_list()
        dh_raster_layout.addWidget(self.dh_raster_combo)
        dh_raster_refresh_btn = QPushButton("↻")
        dh_raster_refresh_btn.setFixedWidth(28)
        dh_raster_refresh_btn.setToolTip("Refresh raster list")
        dh_raster_refresh_btn.clicked.connect(self.refresh_dh_raster_list)
        dh_raster_layout.addWidget(dh_raster_refresh_btn)
        dh_layout.addLayout(dh_raster_layout)

        self.dh_raster_override_chk = QCheckBox("Always use raster RL (ignore CSV RL column if present)")
        self.dh_raster_override_chk.setChecked(True)
        self.dh_raster_override_chk.setToolTip(
            "When ticked, the collar elevation is always taken from the selected\n"
            "LiDAR/DEM raster, even if the CSV contains an RL column."
        )
        dh_layout.addWidget(self.dh_raster_override_chk)

        # Section polyline layer picker
        dh_layer_layout = QHBoxLayout()
        dh_layer_layout.addWidget(QLabel("Section line layer:"))
        self.dh_layer_combo = QComboBox()
        self.dh_layer_combo.setToolTip(
            "Select the polyline layer whose features define the section lines.\n"
            "The plugin will use the first vertex and last vertex of each feature\n"
            "to determine the section azimuth."
        )
        self.refresh_dh_layer_list()
        dh_layer_layout.addWidget(self.dh_layer_combo)
        dh_refresh_btn = QPushButton("↻")
        dh_refresh_btn.setFixedWidth(28)
        dh_refresh_btn.setToolTip("Refresh layer list")
        dh_refresh_btn.clicked.connect(self.refresh_dh_layer_list)
        dh_layer_layout.addWidget(dh_refresh_btn)
        dh_layout.addLayout(dh_layer_layout)

        # Corridor width
        corridor_layout = QHBoxLayout()
        corridor_layout.addWidget(QLabel("Corridor width (±m):"))
        self.dh_corridor_spin = QDoubleSpinBox()
        self.dh_corridor_spin.setRange(1, 5000)
        self.dh_corridor_spin.setValue(200)
        self.dh_corridor_spin.setSuffix(" m")
        self.dh_corridor_spin.setToolTip(
            "Drill holes within this perpendicular distance from the section\n"
            "line will be projected and plotted."
        )
        corridor_layout.addWidget(self.dh_corridor_spin)
        corridor_layout.addStretch()
        dh_layout.addLayout(corridor_layout)

        # Label options
        self.dh_label_id_chk = QCheckBox("Label collar with Hole ID")
        self.dh_label_id_chk.setChecked(True)
        dh_layout.addWidget(self.dh_label_id_chk)

        self.dh_label_td_chk = QCheckBox("Label trace end with Total Depth (m)")
        self.dh_label_td_chk.setChecked(True)
        dh_layout.addWidget(self.dh_label_td_chk)

        dh_group.setLayout(dh_layout)
        controls_layout.addWidget(dh_group)
        # ──────────────────────────────────────────────────────────────────

        # ── Drill Trace Geology ───────────────────────────────────────────
        dh_geo_group = QGroupBox("Drill Trace Geology")
        dh_geo_layout = QVBoxLayout()

        self.dh_geo_chk = QCheckBox("Plot geology on drill traces")
        self.dh_geo_chk.setToolTip(
            "Reads the geology sheet from each drill log .xlsx and colours\n"
            "segments of the drill trace by lithology code."
        )
        dh_geo_layout.addWidget(self.dh_geo_chk)

        # Sheet name
        dh_geo_sheet_row = QHBoxLayout()
        dh_geo_sheet_row.addWidget(QLabel("Sheet name:"))
        self.dh_geo_sheet_edit = QLineEdit("Geology")
        self.dh_geo_sheet_edit.setToolTip("Name of the geology/lithology sheet in the .xlsx files")
        dh_geo_sheet_row.addWidget(self.dh_geo_sheet_edit)
        dh_geo_layout.addLayout(dh_geo_sheet_row)

        # From / To columns
        dh_geo_ft_row = QHBoxLayout()
        dh_geo_ft_row.addWidget(QLabel("From col:"))
        self.dh_geo_from_edit = QLineEdit("mFrom")
        self.dh_geo_from_edit.setMaximumWidth(70)
        self.dh_geo_from_edit.setToolTip("Column name for interval start depth")
        dh_geo_ft_row.addWidget(self.dh_geo_from_edit)
        dh_geo_ft_row.addSpacing(6)
        dh_geo_ft_row.addWidget(QLabel("To col:"))
        self.dh_geo_to_edit = QLineEdit("mTo")
        self.dh_geo_to_edit.setMaximumWidth(70)
        self.dh_geo_to_edit.setToolTip("Column name for interval end depth")
        dh_geo_ft_row.addWidget(self.dh_geo_to_edit)
        dh_geo_layout.addLayout(dh_geo_ft_row)

        # Code column
        dh_geo_code_row = QHBoxLayout()
        dh_geo_code_row.addWidget(QLabel("Code column:"))
        self.dh_geo_code_edit = QLineEdit("Lith1_Code")
        self.dh_geo_code_edit.setToolTip(
            "Column name for the lithology/rock code used for colouring"
        )
        dh_geo_code_row.addWidget(self.dh_geo_code_edit)
        dh_geo_layout.addLayout(dh_geo_code_row)

        dh_geo_group.setLayout(dh_geo_layout)
        controls_layout.addWidget(dh_geo_group)
        # ──────────────────────────────────────────────────────────────────

        # ── Planned Holes ─────────────────────────────────────────────────
        planned_group = QGroupBox("Planned Holes")
        planned_layout = QVBoxLayout()

        self.planned_enable_chk = QCheckBox("Plot planned holes on cross-sections")
        self.planned_enable_chk.setToolTip(
            "Reads a CSV of planned collar locations and projects straight traces\n"
            "onto each cross-section using the hole azimuth, dip and TD."
        )
        planned_layout.addWidget(self.planned_enable_chk)

        # CSV file picker
        planned_file_layout = QHBoxLayout()
        planned_file_layout.addWidget(QLabel("Collars CSV:"))
        self.planned_csv_edit = QLineEdit()
        self.planned_csv_edit.setReadOnly(True)
        self.planned_csv_edit.setPlaceholderText(
            "CSV with Hole_ID, Easting, Northing, Azimuth, Dip, TD columns"
        )
        planned_file_layout.addWidget(self.planned_csv_edit)
        planned_browse_btn = QPushButton("Browse\u2026")
        planned_browse_btn.clicked.connect(self.browse_planned_holes_csv)
        planned_file_layout.addWidget(planned_browse_btn)
        planned_layout.addLayout(planned_file_layout)

        # Corridor
        planned_corr_layout = QHBoxLayout()
        planned_corr_layout.addWidget(QLabel("Corridor (\u00b1m):"))
        self.planned_corridor_spin = QDoubleSpinBox()
        self.planned_corridor_spin.setRange(1, 5000)
        self.planned_corridor_spin.setValue(200)
        self.planned_corridor_spin.setSuffix(" m")
        self.planned_corridor_spin.setToolTip(
            "Maximum perpendicular distance from the section line for a planned\n"
            "hole to be included on that cross-section."
        )
        planned_corr_layout.addWidget(self.planned_corridor_spin)
        planned_corr_layout.addStretch()
        planned_layout.addLayout(planned_corr_layout)

        self.planned_label_chk = QCheckBox("Label planned holes with Hole ID")
        self.planned_label_chk.setChecked(True)
        planned_layout.addWidget(self.planned_label_chk)

        planned_group.setLayout(planned_layout)
        controls_layout.addWidget(planned_group)
        # ──────────────────────────────────────────────────────────────────

        # Geology Bands group
        geo_group = QGroupBox("Geology Bands")
        geo_layout = QVBoxLayout()

        self.geo_band_chk = QCheckBox("Plot geology bands on cross-sections")
        self.geo_band_chk.setToolTip(
            "Intersects the section line with the selected polygon layer and draws\n"
            "a colored strip at the top of each plot showing the geology along the surface."
        )
        geo_layout.addWidget(self.geo_band_chk)

        # Section line layer picker (independent of Drill Hole Overlay)
        geo_sec_layout = QHBoxLayout()
        geo_sec_layout.addWidget(QLabel("Section line layer:"))
        self.geo_section_line_combo = QComboBox()
        self.geo_section_line_combo.setToolTip(
            "Select the polyline layer whose features define the cross-section lines.\n"
            "Each feature must have a name field that matches the Excel sheet name exactly.\n"
            "This is required for geology bands to know where to intersect the geology polygons."
        )
        self.refresh_geo_section_line_layers()
        geo_sec_layout.addWidget(self.geo_section_line_combo)
        geo_sec_refresh_btn = QPushButton("↻")
        geo_sec_refresh_btn.setFixedWidth(28)
        geo_sec_refresh_btn.setToolTip("Refresh section line layer list")
        geo_sec_refresh_btn.clicked.connect(self.refresh_geo_section_line_layers)
        geo_sec_layout.addWidget(geo_sec_refresh_btn)
        geo_layout.addLayout(geo_sec_layout)

        # Polygon layer picker
        geo_layer_layout = QHBoxLayout()
        geo_layer_layout.addWidget(QLabel("Geology layer:"))
        self.geo_band_layer_combo = QComboBox()
        self.geo_band_layer_combo.setToolTip("Select the polygon layer containing geology units")
        self.refresh_geo_band_layers()
        geo_layer_layout.addWidget(self.geo_band_layer_combo)
        geo_refresh_btn = QPushButton("↻")
        geo_refresh_btn.setFixedWidth(28)
        geo_refresh_btn.setToolTip("Refresh layer list")
        geo_refresh_btn.clicked.connect(self.refresh_geo_band_layers)
        geo_layer_layout.addWidget(geo_refresh_btn)
        geo_layout.addLayout(geo_layer_layout)

        # Label field picker (populated when layer changes)
        geo_field_layout = QHBoxLayout()
        geo_field_layout.addWidget(QLabel("Label field:"))
        self.geo_band_field_combo = QComboBox()
        self.geo_band_field_combo.setToolTip("Field in the geology layer to use as the band label")
        geo_field_layout.addWidget(self.geo_band_field_combo)
        geo_layout.addLayout(geo_field_layout)
        self.geo_band_layer_combo.currentIndexChanged.connect(self.refresh_geo_band_fields)

        # Band thickness as % of plot height
        geo_thick_layout = QHBoxLayout()
        geo_thick_layout.addWidget(QLabel("Band height (%):"))
        self.geo_band_pct_spin = QSpinBox()
        self.geo_band_pct_spin.setRange(2, 25)
        self.geo_band_pct_spin.setValue(8)
        self.geo_band_pct_spin.setToolTip("Height of the geology strip as a percentage of the total plot height")
        geo_thick_layout.addWidget(self.geo_band_pct_spin)
        geo_thick_layout.addStretch()
        geo_layout.addLayout(geo_thick_layout)

        self.geo_band_label_chk = QCheckBox("Show geology labels in band")
        self.geo_band_label_chk.setChecked(True)
        geo_layout.addWidget(self.geo_band_label_chk)

        geo_group.setLayout(geo_layout)
        controls_layout.addWidget(geo_group)
        # ──────────────────────────────────────────────────────────────────

        # ── Interpreted Pegmatites group ──────────────────────────────────
        peg_group = QGroupBox("Interpreted Pegmatites")
        peg_layout = QVBoxLayout()

        self.peg_band_chk = QCheckBox("Plot interpreted pegmatites on cross-sections")
        self.peg_band_chk.setToolTip(
            "Intersects the section line with the selected polygon layer and draws\n"
            "hatched vertical bands showing the pegmatite extents along each section."
        )
        peg_layout.addWidget(self.peg_band_chk)

        # Polygon layer picker
        peg_layer_layout = QHBoxLayout()
        peg_layer_layout.addWidget(QLabel("Pegmatite layer:"))
        self.peg_band_layer_combo = QComboBox()
        self.peg_band_layer_combo.setToolTip(
            "Select the polygon layer containing interpreted pegmatite outlines\n"
            "(e.g. the study_areas layer from Interpreted_pegmatites.gpkg)"
        )
        self.refresh_peg_layers()
        peg_layer_layout.addWidget(self.peg_band_layer_combo)
        peg_refresh_btn = QPushButton("↻")
        peg_refresh_btn.setFixedWidth(28)
        peg_refresh_btn.setToolTip("Refresh layer list")
        peg_refresh_btn.clicked.connect(self.refresh_peg_layers)
        peg_layer_layout.addWidget(peg_refresh_btn)
        peg_layout.addLayout(peg_layer_layout)

        self.peg_labeled_only_chk = QCheckBox("Only show labeled features (filter out null labels)")
        self.peg_labeled_only_chk.setChecked(True)
        self.peg_labeled_only_chk.setToolTip(
            "When checked, only features with a non-empty 'label' field value\n"
            "are included (e.g. Int_peg_01 – Int_peg_06). Unlabeled/draft polygons\n"
            "are excluded."
        )
        peg_layout.addWidget(self.peg_labeled_only_chk)

        self.peg_band_label_chk = QCheckBox("Show pegmatite labels on plot")
        self.peg_band_label_chk.setChecked(True)
        peg_layout.addWidget(self.peg_band_label_chk)

        peg_group.setLayout(peg_layout)
        controls_layout.addWidget(peg_group)
        # ──────────────────────────────────────────────────────────────────

        # Structural Measurements group
        struct_group = QGroupBox("Structural Measurements")
        struct_layout = QVBoxLayout()

        self.struct_enable_chk = QCheckBox("Plot structural dip ticks on cross-sections")
        self.struct_enable_chk.setToolTip(
            "Projects point features from the selected layer onto each section\n"
            "and draws a dip tick mark (5 m long) at the surface."
        )
        struct_layout.addWidget(self.struct_enable_chk)

        # Point layer picker
        struct_layer_layout = QHBoxLayout()
        struct_layer_layout.addWidget(QLabel("Structure layer:"))
        self.struct_layer_combo = QComboBox()
        self.struct_layer_combo.setToolTip("Point layer containing structural measurements")
        self.refresh_struct_layers()
        struct_layer_layout.addWidget(self.struct_layer_combo)
        struct_refresh_btn = QPushButton("↻")
        struct_refresh_btn.setFixedWidth(28)
        struct_refresh_btn.setToolTip("Refresh layer list")
        struct_refresh_btn.clicked.connect(self.refresh_struct_layers)
        struct_layer_layout.addWidget(struct_refresh_btn)
        struct_layout.addLayout(struct_layer_layout)

        # Dip field picker
        struct_dip_layout = QHBoxLayout()
        struct_dip_layout.addWidget(QLabel("Dip field:"))
        self.struct_dip_combo = QComboBox()
        self.struct_dip_combo.setToolTip("Field containing the dip value (degrees from horizontal)")
        struct_dip_layout.addWidget(self.struct_dip_combo)
        struct_layout.addLayout(struct_dip_layout)

        # Dip direction field picker
        struct_dir_layout = QHBoxLayout()
        struct_dir_layout.addWidget(QLabel("Dip direction field:"))
        self.struct_dir_combo = QComboBox()
        self.struct_dir_combo.setToolTip("Field containing the dip direction / azimuth (degrees from north)")
        struct_dir_layout.addWidget(self.struct_dir_combo)
        struct_layout.addLayout(struct_dir_layout)

        self.struct_layer_combo.currentIndexChanged.connect(self.refresh_struct_fields)

        # Tick length
        struct_tick_layout = QHBoxLayout()
        struct_tick_layout.addWidget(QLabel("Tick length (m):"))
        self.struct_tick_spin = QDoubleSpinBox()
        self.struct_tick_spin.setRange(1, 100)
        self.struct_tick_spin.setValue(5.0)
        self.struct_tick_spin.setDecimals(1)
        self.struct_tick_spin.setToolTip("Length of the dip tick line in metres")
        struct_tick_layout.addWidget(self.struct_tick_spin)
        struct_tick_layout.addStretch()
        struct_layout.addLayout(struct_tick_layout)

        # Structural corridor — independent of drill hole corridor
        struct_corr_layout = QHBoxLayout()
        struct_corr_layout.addWidget(QLabel("Corridor (±m):"))
        self.struct_corridor_spin = QDoubleSpinBox()
        self.struct_corridor_spin.setRange(1, 10000)
        self.struct_corridor_spin.setValue(500)
        self.struct_corridor_spin.setDecimals(0)
        self.struct_corridor_spin.setSuffix(" m")
        self.struct_corridor_spin.setToolTip(
            "Maximum perpendicular distance from the section line\n"
            "for a structural measurement to be included."
        )
        struct_corr_layout.addWidget(self.struct_corridor_spin)
        struct_corr_layout.addStretch()
        struct_layout.addLayout(struct_corr_layout)

        self.struct_label_chk = QCheckBox("Label tick with dip value")
        self.struct_label_chk.setChecked(True)
        struct_layout.addWidget(self.struct_label_chk)

        struct_group.setLayout(struct_layout)
        controls_layout.addWidget(struct_group)
        # ──────────────────────────────────────────────────────────────────

        # ── Plate Locations group ──────────────────────────────────────────
        plate_group = QGroupBox("Plate Locations")
        plate_layout = QVBoxLayout()

        self.plate_enable_chk = QCheckBox("Plot plate locations on cross-sections")
        self.plate_enable_chk.setToolTip(
            "Projects photo plate locations onto each section line and marks\n"
            "them at the top of the plot with a downward arrow and label.\n"
            "Markers sit outside the data area so nothing is obscured."
        )
        plate_layout.addWidget(self.plate_enable_chk)

        plate_layer_layout = QHBoxLayout()
        plate_layer_layout.addWidget(QLabel("Plate layer:"))
        self.plate_layer_combo = QComboBox()
        self.plate_layer_combo.setToolTip(
            "Point layer containing plate/photo locations\n"
            "(e.g. Plate_Locations from Plates_photo_locations.shp)"
        )
        self.refresh_plate_layers()
        plate_layer_layout.addWidget(self.plate_layer_combo)
        plate_refresh_btn = QPushButton("↻")
        plate_refresh_btn.setFixedWidth(28)
        plate_refresh_btn.setToolTip("Refresh layer list")
        plate_refresh_btn.clicked.connect(self.refresh_plate_layers)
        plate_layer_layout.addWidget(plate_refresh_btn)
        plate_layout.addLayout(plate_layer_layout)

        plate_corr_layout = QHBoxLayout()
        plate_corr_layout.addWidget(QLabel("Corridor (±m):"))
        self.plate_corridor_spin = QDoubleSpinBox()
        self.plate_corridor_spin.setRange(1, 10000)
        self.plate_corridor_spin.setValue(500)
        self.plate_corridor_spin.setDecimals(0)
        self.plate_corridor_spin.setSuffix(" m")
        self.plate_corridor_spin.setToolTip(
            "Maximum perpendicular distance from the section line\n"
            "for a plate location to be included on that cross-section."
        )
        plate_corr_layout.addWidget(self.plate_corridor_spin)
        plate_corr_layout.addStretch()
        plate_layout.addLayout(plate_corr_layout)

        self.plate_label_chk = QCheckBox("Show plate name label")
        self.plate_label_chk.setChecked(True)
        plate_layout.addWidget(self.plate_label_chk)

        plate_group.setLayout(plate_layout)
        controls_layout.addWidget(plate_group)
        # ──────────────────────────────────────────────────────────────────

        # ── Plot Options ─────────────────────────────────────────────────
        plot_opt_group = QGroupBox("Plot Options")
        plot_opt_layout = QVBoxLayout()

        # Marker size row
        mk_size_row = QHBoxLayout()
        mk_size_row.addWidget(QLabel("Marker size (pt):"))
        self.plot_marker_size_spin = QSpinBox()
        self.plot_marker_size_spin.setRange(0, 30)
        self.plot_marker_size_spin.setValue(6)
        self.plot_marker_size_spin.setToolTip(
            "Size of the marker symbols on the profile plots.\n"
            "Set to 0 to hide all markers."
        )
        mk_size_row.addWidget(self.plot_marker_size_spin)
        mk_size_row.addStretch()
        plot_opt_layout.addLayout(mk_size_row)

        # Y-axis extra headroom
        y_extra_row = QHBoxLayout()
        y_extra_row.addWidget(QLabel("Extra Y headroom (m):"))
        self.plot_y_extra_spin = QSpinBox()
        self.plot_y_extra_spin.setRange(0, 2000)
        self.plot_y_extra_spin.setValue(0)
        self.plot_y_extra_spin.setSingleStep(10)
        self.plot_y_extra_spin.setToolTip(
            "Extra metres to add above and below the auto-calculated Y-axis limits.\n"
            "Useful to see more vertical context while keeping 1:1 scale."
        )
        y_extra_row.addWidget(self.plot_y_extra_spin)
        y_extra_row.addStretch()
        plot_opt_layout.addLayout(y_extra_row)

        plot_opt_group.setLayout(plot_opt_layout)
        controls_layout.addWidget(plot_opt_group)
        # ──────────────────────────────────────────────────────────────────

        # Plot settings
        settings_group = QGroupBox("Plot Settings")
        settings_layout = QVBoxLayout()
        
        # Configuration button
        config_btn = QPushButton("Configure Plot Appearance")
        config_btn.clicked.connect(self.configure_plot_appearance)
        settings_layout.addWidget(config_btn)
        
        # Output options
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("Save plots as:"))
        self.plot_output_combo = QComboBox()
        self.plot_output_combo.addItems(["Display Only", "PNG", "PDF", "Both PDF and PNG"])
        output_layout.addWidget(self.plot_output_combo)
        settings_layout.addLayout(output_layout)
        
        # Display options - Always display in dialog
        display_label = QLabel("📊 Plots will be displayed in the right panel")
        display_label.setStyleSheet("QLabel { color: #1976D2; font-weight: bold; }")
        settings_layout.addWidget(display_label)
        
        settings_group.setLayout(settings_layout)
        controls_layout.addWidget(settings_group)
        
        # Plot button
        self.plot_btn = QPushButton("Generate Plots")
        self.plot_btn.clicked.connect(self.generate_plots)
        self.plot_btn.setEnabled(False)
        self.plot_btn.setStyleSheet("QPushButton { background-color: #9C27B0; color: white; font-weight: bold; padding: 8px; }")
        controls_layout.addWidget(self.plot_btn)
        
        controls_layout.addStretch()
        
        # Right side - plot display (expanded area)
        self.plot_display_widget = PlotDisplayWidget()
        self.plot_display_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Create a group box for the plot area with a border
        plot_group = QGroupBox("📊 Plot Display Area")
        plot_group_layout = QVBoxLayout()
        plot_group_layout.setContentsMargins(5, 5, 5, 5)
        plot_group_layout.addWidget(self.plot_display_widget)
        plot_group.setLayout(plot_group_layout)
        plot_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #1976D2;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #1976D2;
            }
        """)
        
        # Add to horizontal layout with more space for plots
        # Wrap controls in a scroll area so all options are reachable
        controls_scroll = QScrollArea()
        controls_scroll.setWidgetResizable(True)
        controls_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        controls_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        controls_scroll.setMinimumWidth(340)
        controls_scroll.setMaximumWidth(480)
        controls_scroll.setWidget(controls_widget)
        h_layout.addWidget(controls_scroll, 2)  # Controls take 2 parts
        h_layout.addWidget(plot_group, 5)       # Plots take 5 parts (more space)
        
        main_layout.addLayout(h_layout)
        self.plotting_tab.setLayout(main_layout)
        
    def browse_plot_excel(self):
        """Browse for Excel file to plot"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            self.last_directory,
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        
        if file_path:
            self.plot_excel_edit.setText(file_path)
            self.last_directory = os.path.dirname(file_path)
            self.settings.setValue("lastDirectory", self.last_directory)
            self.load_excel_sheets()
            
    def load_excel_sheets(self):
        """Load sheets from Excel file with enhanced visibility"""
        if not self.plot_excel_edit.text():
            QMessageBox.warning(self, "No File", "Please select an Excel file first")
            return
            
        try:
            if PANDAS_AVAILABLE:
                excel_file = pd.ExcelFile(self.plot_excel_edit.text())
                sheet_names = excel_file.sheet_names
                
                # Clear existing table
                self.sheet_table.setRowCount(0)
                
                # Add rows for each sheet with enhanced defaults
                for i, sheet_name in enumerate(sheet_names):
                    row_position = self.sheet_table.rowCount()
                    self.sheet_table.insertRow(row_position)
                    
                    # Sheet name (non-editable)
                    name_item = QTableWidgetItem(sheet_name)
                    name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                    name_item.setBackground(QColor(240, 248, 255))  # Light blue background
                    self.sheet_table.setItem(row_position, 0, name_item)
                    
                    # Plot checkbox with better styling
                    plot_checkbox = QCheckBox()
                    plot_checkbox.setChecked(True)
                    plot_checkbox.setStyleSheet("""
                        QCheckBox {
                            margin: 5px;
                        }
                        QCheckBox::indicator {
                            width: 18px;
                            height: 18px;
                        }
                    """)
                    
                    # Center the checkbox
                    checkbox_widget = QWidget()
                    checkbox_layout = QHBoxLayout(checkbox_widget)
                    checkbox_layout.addWidget(plot_checkbox)
                    checkbox_layout.setAlignment(Qt.AlignCenter)
                    checkbox_layout.setContentsMargins(0, 0, 0, 0)
                    
                    self.sheet_table.setCellWidget(row_position, 1, checkbox_widget)
                    
                    # Default values with better formatting
                    default_values = ["0.0", "0.0", "0.0", "0.0", "0.0"]
                    for col, default in enumerate(default_values, start=2):
                        item = QTableWidgetItem(default)
                        item.setTextAlignment(Qt.AlignCenter)
                        # Add tooltips for parameter columns
                        tooltips = [
                            "Distance where slope analysis starts",
                            "Distance where slope analysis ends", 
                            "Distance where runout analysis starts",
                            "Distance where runout analysis ends",
                            "Reference angle in degrees"
                        ]
                        item.setToolTip(tooltips[col-2])
                        self.sheet_table.setItem(row_position, col, item)
                    
                    # Connect checkbox change to summary update
                    plot_checkbox.stateChanged.connect(self.update_parameters_summary)
                        
                # Update summary
                checked_count = len([name for name in sheet_names])  # All start checked
                self.sheet_summary_label.setText(
                    f"📊 Loaded {len(sheet_names)} sheets | "
                    f"✓ {checked_count} selected for plotting | "
                    f"📋 Click cells to edit parameters"
                )
                
                # Connect table changes to summary updates
                self.sheet_table.cellChanged.connect(self.update_parameters_summary)
                            
                self.plot_btn.setEnabled(True)
                
                QMessageBox.information(self, "Success", 
                    f"Loaded {len(sheet_names)} sheets\n\n"
                    f"📋 Edit parameters directly in the table\n"
                    f"✓ Check/uncheck Plot column to include/exclude sheets\n"
                    f"💡 Hover over parameter cells for help tooltips")
            else:
                QMessageBox.warning(self, "Not Available", "Pandas is required to read Excel files")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel file:\n{str(e)}")
    
    def update_parameters_summary(self):
        """Update the summary display when parameters change"""
        try:
            selected_count = 0
            total_sheets = self.sheet_table.rowCount()
            
            for row in range(total_sheets):
                plot_widget = self.sheet_table.cellWidget(row, 1)
                checkbox = plot_widget.findChild(QCheckBox) if plot_widget else None
                if checkbox and checkbox.isChecked():
                    selected_count += 1
            
            if total_sheets > 0:
                self.sheet_summary_label.setText(
                    f"📊 {total_sheets} sheets loaded | "
                    f"✓ {selected_count} selected for plotting | "
                    f"📋 {total_sheets - selected_count} deselected"
                )
            else:
                self.sheet_summary_label.setText("No sheets loaded")
                
        except Exception as e:
            self.sheet_summary_label.setText("Error updating summary")
    
    def validate_sheet_parameters(self):
        """Validate all sheet parameters before plotting"""
        errors = []
        
        for row in range(self.sheet_table.rowCount()):
            sheet_name = self.sheet_table.item(row, 0).text()
            plot_widget = self.sheet_table.cellWidget(row, 1)
            
            # Check if sheet is selected for plotting
            checkbox = plot_widget.findChild(QCheckBox) if plot_widget else None
            if not (checkbox and checkbox.isChecked()):
                continue
                
            # Validate numeric parameters
            for col in range(2, 7):
                item = self.sheet_table.item(row, col)
                if item:
                    try:
                        value = float(item.text())
                        if col in [2, 3, 4, 5] and value < 0:  # Distance parameters
                            errors.append(f"Sheet '{sheet_name}': Distance values cannot be negative")
                    except ValueError:
                        param_names = ["Start Slope", "End Slope", "Runout Start", "Runout End", "Reference Angle"]
                        errors.append(f"Sheet '{sheet_name}': Invalid {param_names[col-2]} value")
        
        return errors
            
            
    def configure_plot_appearance(self):
        """Open plot configuration dialog"""
        dialog = PlotConfigDialog(self)
        if dialog.exec_():
            self.plot_config = dialog.get_config()
            QMessageBox.information(self, "Success", "Plot configuration updated")
    
    def open_sheet_parameters_dialog(self):
        """Open the dedicated sheet parameters dialog"""
        if not hasattr(self, 'sheet_params_dialog') or not self.sheet_params_dialog:
            self.sheet_params_dialog = SheetParametersDialog(self)
        
        self.sheet_params_dialog.show()
        self.sheet_params_dialog.raise_()
        self.sheet_params_dialog.activateWindow()
        
        # Auto-load current parameters
        self.sheet_params_dialog.load_from_parent()
            
    def generate_plots(self):
        """Generate plots from Excel data"""
        if not MATPLOTLIB_AVAILABLE:
            QMessageBox.warning(self, "Not Available", 
                "Matplotlib is required for plotting. Please install matplotlib.")
            return
            
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, "Not Available", 
                "Pandas is required to read Excel files. Please install pandas.")
            return
        
        # Validate sheet parameters before plotting
        validation_errors = self.validate_sheet_parameters()
        if validation_errors:
            error_message = "Please fix the following parameter errors:\n\n" + "\n".join(validation_errors)
            QMessageBox.warning(self, "Parameter Validation Failed", error_message)
            return
            
        try:
            excel_path = self.plot_excel_edit.text()
            if not excel_path:
                QMessageBox.warning(self, "No File", "Please select an Excel file first")
                return
                
            # Collect sheet parameters
            sheet_params = {}
            for row in range(self.sheet_table.rowCount()):
                sheet_name = self.sheet_table.item(row, 0).text()
                plot_widget = self.sheet_table.cellWidget(row, 1)
                
                # Find the actual checkbox inside the widget
                checkbox = plot_widget.findChild(QCheckBox) if plot_widget else None
                
                if checkbox and checkbox.isChecked():
                    try:
                        params = {
                            'start_slope': float(self.sheet_table.item(row, 2).text()),
                            'end_slope': float(self.sheet_table.item(row, 3).text()),
                            'runout_start': float(self.sheet_table.item(row, 4).text()),
                            'runout_end': float(self.sheet_table.item(row, 5).text()),
                            'ref_angle': float(self.sheet_table.item(row, 6).text())
                        }
                        sheet_params[sheet_name] = params
                    except ValueError:
                        QMessageBox.warning(self, "Invalid Input", 
                            f"Invalid numeric values for sheet: {sheet_name}")
                        return
                        
            if not sheet_params:
                QMessageBox.warning(self, "No Sheets", "No sheets selected for plotting")
                return
                
            # Clear previous plots and always display in dialog
            if self.plot_display_widget:
                self.plot_display_widget.clear_plots()
                
            # Generate plots
            figures = []
            progress = QProgressDialog("Generating plots...", "Cancel", 0, len(sheet_params), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            for idx, (sheet_name, params) in enumerate(sheet_params.items()):
                progress.setValue(idx)
                progress.setLabelText(f"Processing {sheet_name}")
                QCoreApplication.processEvents()
                
                if progress.wasCanceled():
                    break
                    
                try:
                    # Read data
                    df = pd.read_excel(excel_path, sheet_name=sheet_name)
                    
                    # Create plot
                    fig = self.plot_single_profile(df, sheet_name, params, None)
                    if fig:
                        figures.append(fig)
                        
                        # Always display in dialog
                        if self.plot_display_widget:
                            self.plot_display_widget.add_plot(fig)
                            
                except Exception as e:
                    QgsMessageLog.logMessage(f"Error plotting {sheet_name}: {str(e)}", 
                                           "IntegratedProfileAnalyzer", Qgis.Warning)
                    
            progress.close()
            
            if figures:
                # Save plots if requested
                output_type = self.plot_output_combo.currentText()
                
                if output_type != "Display Only":
                    base_path = os.path.splitext(excel_path)[0]
                    
                    if output_type in ["PDF", "Both PDF and PNG"]:
                        pdf_path = base_path + "_plots.pdf"
                        with PdfPages(pdf_path) as pdf:
                            for fig in figures:
                                pdf.savefig(fig, orientation='landscape')
                        QMessageBox.information(self, "Success", f"PDF saved to:\n{pdf_path}")
                        
                    if output_type in ["PNG", "Both PDF and PNG"]:
                        png_dir = base_path + "_plots"
                        os.makedirs(png_dir, exist_ok=True)
                        for idx, fig in enumerate(figures):
                            png_path = os.path.join(png_dir, f"plot_{idx+1}.png")
                            fig.savefig(png_path, dpi=300, bbox_inches='tight')
                        QMessageBox.information(self, "Success", f"PNG files saved to:\n{png_dir}")
                
                # Plots are always displayed in the dialog panel
                # No need for separate matplotlib windows
                
                # Clean up figures after saving (keep them for display)
                # Figures will be managed by the PlotDisplayWidget
                gc.collect()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate plots:\n{str(e)}")
    
    def _detect_profile_columns(self, df: pd.DataFrame) -> List[Tuple[str, str]]:
        """Detect all distance/elevation column pairs in the dataframe using regex patterns"""
        import re
        
        column_pairs = []
        available_columns = list(df.columns)
        
        QgsMessageLog.logMessage(
            f"Available columns for detection: {available_columns}",
            "IntegratedProfileAnalyzer", Qgis.Info
        )
        
        # Common patterns for distance and elevation columns with regex
        pattern_sets = [
            # Numbered patterns: Distance_01/Height_01, Distance_02/Height_02, etc.
            (r'(?i)distance[-_]?(\d{1,2})', r'(?i)height[-_]?\1'),
            (r'(?i)distance[-_]?(\d{1,2})', r'(?i)elev(?:ation)?[-_]?\1'),
            (r'(?i)dist[-_]?(\d{1,2})', r'(?i)height[-_]?\1'),
            (r'(?i)dist[-_]?(\d{1,2})', r'(?i)elev(?:ation)?[-_]?\1'),
            
            # Prefixed numbered patterns: Geology_distance_02/Geology_height_02, feature_Distance_03/Feature_Height_03
            (r'(?i)(.+?)[-_]?distance[-_]?(\d{1,2})', r'(?i)\1[-_]?height[-_]?\2'),
            (r'(?i)(.+?)[-_]?distance[-_]?(\d{1,2})', r'(?i)\1[-_]?elev(?:ation)?[-_]?\2'),
            (r'(?i)(.+?)[-_]?dist[-_]?(\d{1,2})', r'(?i)\1[-_]?height[-_]?\2'),
            (r'(?i)(.+?)[-_]?dist[-_]?(\d{1,2})', r'(?i)\1[-_]?elev(?:ation)?[-_]?\2'),
            
            # Letter patterns: Distance_A/Height_A, Distance_B/Height_B, etc.
            (r'(?i)distance[-_]?([A-Z])', r'(?i)height[-_]?\1'),
            (r'(?i)distance[-_]?([A-Z])', r'(?i)elev(?:ation)?[-_]?\1'),
            (r'(?i)dist[-_]?([A-Z])', r'(?i)height[-_]?\1'),
            (r'(?i)dist[-_]?([A-Z])', r'(?i)elev(?:ation)?[-_]?\1'),
            
            # Prefixed letter patterns: Site_Distance_A/Site_Height_A
            (r'(?i)(.+?)[-_]?distance[-_]?([A-Z])', r'(?i)\1[-_]?height[-_]?\2'),
            (r'(?i)(.+?)[-_]?distance[-_]?([A-Z])', r'(?i)\1[-_]?elev(?:ation)?[-_]?\2'),
            
            # Simple patterns: Distance/Height (no suffix)
            (r'(?i)^distance$', r'(?i)^height$'),
            (r'(?i)^distance$', r'(?i)^elevation$'),
            (r'(?i)^dist$', r'(?i)^height$'),
            (r'(?i)^dist$', r'(?i)^elev$'),
        ]
        
        used_columns = set()
        
        # Try pattern matching
        for dist_pattern, elev_pattern in pattern_sets:
            for col in available_columns:
                if col in used_columns or 'marker' in col.lower():
                    continue
                    
                dist_match = re.match(dist_pattern, col)
                if dist_match:
                    # Found distance column, look for matching elevation column
                    if dist_match.groups():
                        # Handle different numbers of capture groups
                        if len(dist_match.groups()) == 1:
                            # Single group (suffix only): Distance_01 -> Height_01
                            group = dist_match.group(1)
                            target_elev_pattern = elev_pattern.replace(r'\1', group)
                        elif len(dist_match.groups()) == 2:
                            # Two groups (prefix + suffix): Geology_distance_02 -> Geology_height_02
                            prefix = dist_match.group(1)
                            suffix = dist_match.group(2)
                            target_elev_pattern = elev_pattern.replace(r'\1', prefix).replace(r'\2', suffix)
                        else:
                            target_elev_pattern = elev_pattern
                    else:
                        target_elev_pattern = elev_pattern
                    
                    QgsMessageLog.logMessage(
                        f"Testing distance column '{col}' - looking for elevation pattern: {target_elev_pattern}",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                    
                    for elev_col in available_columns:
                        if elev_col in used_columns or 'marker' in elev_col.lower():
                            continue
                        if re.match(target_elev_pattern, elev_col):
                            column_pairs.append((col, elev_col))
                            used_columns.add(col)
                            used_columns.add(elev_col)
                            QgsMessageLog.logMessage(
                                f"✓ Detected column pair: {col} -> {elev_col}",
                                "IntegratedProfileAnalyzer", Qgis.Info
                            )
                            break
        
        # If no pattern matches found, try explicit common pairs
        if not column_pairs:
            QgsMessageLog.logMessage(
                "No regex matches found, trying explicit common pairs...",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            common_pairs = [
                ('Distance_01', 'Height_01'),
                ('Distance_1', 'Height_1'),
                ('Dist_01', 'Elev_01'),
                ('Distance_02', 'Height_02'),
                ('Distance_2', 'Height_B'),
                ('Distance_03', 'Height_03'),
                ('Distance_3', 'Height_C'),
                ('Distance_04', 'Height_04'),
                ('Distance_4', 'Height_D'),
                ('Distance_A', 'Height_A'),
                ('Distance_B', 'Height_B'),
                ('Distance_C', 'Height_C'),
                ('Distance_D', 'Height_D'),
                ('DISTANCE_01', 'HEIGHT_01'),
                ('distance_01', 'height_01'),
                # Add some more flexible patterns
                ('Distance', 'Height'),
                ('Distance', 'Elevation'),
                ('Dist', 'Height'),
                ('Dist', 'Elev'),
            ]
            
            for dist_col, elev_col in common_pairs:
                if dist_col in available_columns and elev_col in available_columns:
                    if dist_col not in used_columns and elev_col not in used_columns:
                        column_pairs.append((dist_col, elev_col))
                        used_columns.add(dist_col)
                        used_columns.add(elev_col)
                        QgsMessageLog.logMessage(
                            f"✓ Found explicit pair: {dist_col} -> {elev_col}",
                            "IntegratedProfileAnalyzer", Qgis.Info
                        )
        
        # If still no matches, try a more flexible approach - look for any column containing "distance" or "dist" 
        # and pair with any column containing "height", "elevation", or "elev"
        if not column_pairs:
            QgsMessageLog.logMessage(
                "No explicit matches found, trying flexible keyword matching...",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            # Find all potential distance columns
            distance_cols = []
            elevation_cols = []
            
            for col in available_columns:
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in ['distance', 'dist']):
                    if 'marker' not in col_lower:  # Exclude marker columns
                        distance_cols.append(col)
                elif any(keyword in col_lower for keyword in ['height', 'elevation', 'elev']):
                    if 'marker' not in col_lower:  # Exclude marker columns
                        elevation_cols.append(col)
            
            QgsMessageLog.logMessage(
                f"Found potential distance columns: {distance_cols}",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            QgsMessageLog.logMessage(
                f"Found potential elevation columns: {elevation_cols}",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            # Try to pair them intelligently
            for i, dist_col in enumerate(distance_cols):
                if i < len(elevation_cols):
                    elev_col = elevation_cols[i]
                    column_pairs.append((dist_col, elev_col))
                    QgsMessageLog.logMessage(
                        f"✓ Paired by keywords: {dist_col} -> {elev_col}",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
        
        QgsMessageLog.logMessage(
            f"Final result: {len(column_pairs)} column pairs detected: {column_pairs}",
            "IntegratedProfileAnalyzer", Qgis.Info
        )
        
        return column_pairs
    
    def _create_legend_label(self, dist_col: str, elev_col: str, index: int) -> str:
        """Create legend label using configuration first, then fallback to column-based names"""
        import re
        
        # Try to extract meaningful part from column names for pattern matching
        clean_dist = re.sub(r'(?i)distance[-_]?', '', dist_col)
        clean_elev = re.sub(r'(?i)(height|elevation|elev)[-_]?', '', elev_col)
        
        # PRIORITY 1: Use configured legend labels based on profile type/index
        if index == 0:
            # First profile - use main_profile config
            return self.plot_config.get('legend_main_profile', 'Main Profile')
        elif index == 1:
            # Second profile - check if it's a feature or use generic name
            if 'feature' in elev_col.lower() or 'feature' in dist_col.lower():
                return self.plot_config.get('legend_feature', 'Feature Profile')
            else:
                return f'Profile {index + 1}'
        elif index == 2:
            # Third profile - use generic name
            return f'Profile {index + 1}'
        elif index == 3:
            # Fourth profile - use generic name
            return f'Profile {index + 1}'
        
        # PRIORITY 2: Check column content for specific types
        if 'feature' in elev_col.lower() or 'feature' in dist_col.lower():
            return self.plot_config.get('legend_feature', 'Feature Profile')
        elif 'geology' in elev_col.lower() or 'geo' in elev_col.lower():
            return 'Geology Profile'
        elif 'runout' in elev_col.lower() or 'runout' in dist_col.lower():
            return "Runout Profile"
        elif 'slope' in elev_col.lower() or 'slope' in dist_col.lower():
            return "Slope Profile"
        
        # PRIORITY 3: Fallback to original column-based logic with config integration
        if clean_elev:
            if clean_elev.isdigit():
                if clean_elev == '01' or clean_elev == '1':
                    return self.plot_config.get('legend_main_profile', 'Main Profile')
                elif clean_elev == '02' or clean_elev == '2':
                    return self.plot_config.get('legend_feature', 'Feature Profile')
                elif clean_elev == '03' or clean_elev == '3':
                    return f'Profile 3'
                elif clean_elev == '04' or clean_elev == '4':
                    return f'Profile 4'
                else:
                    return f"Profile {clean_elev}"
            else:
                # Handle letter suffixes
                if clean_elev.upper() == 'A':
                    return self.plot_config.get('legend_main_profile', 'Main Profile')
                elif clean_elev.upper() == 'B':
                    return self.plot_config.get('legend_feature', 'Feature Profile')
                elif clean_elev.upper() == 'C':
                    return f'Profile C'
                elif clean_elev.upper() == 'D':
                    return f'Profile D'
                else:
                    return f"Profile {clean_elev.upper()}"
        elif clean_dist:
            if clean_dist.isdigit():
                if clean_dist == '01' or clean_dist == '1':
                    return self.plot_config.get('legend_main_profile', 'Main Profile')
                else:
                    return f"Profile {clean_dist}"
            else:
                return f"Profile {clean_dist.upper()}"
        else:
            # Final fallback based on index with configuration
            if index == 0:
                return self.plot_config.get('legend_main_profile', 'Main Profile')
            else:
                return f"Profile {index + 1}"
    
    def _plot_marker_columns(self, ax, df: pd.DataFrame):
        """Automatically detect and plot all marker columns"""
        import re
        
        marker_patterns = [
            # Marker patterns with optional numbers
            (r'(?i)marker[-_]?distance[-_]?(\d*)', r'(?i)marker[-_]?height[-_]?\1'),
            (r'(?i)marker[-_]?distance[-_]?(\d*)', r'(?i)marker[-_]?elev(?:ation)?[-_]?\1'),
            (r'(?i)marker[-_]?dist[-_]?(\d*)', r'(?i)marker[-_]?height[-_]?\1'),
            (r'(?i)mark[-_]?dist[-_]?(\d*)', r'(?i)mark[-_]?height[-_]?\1'),
            (r'(?i)mark[-_]?distance[-_]?(\d*)', r'(?i)mark[-_]?height[-_]?\1'),
        ]
        
        marker_colors = [
            self.plot_config.get('marker_1_color', '#0070FF'),
            self.plot_config.get('marker_2_color', '#FF0000'),
            '#FF9800',  # Orange
            '#4CAF50',  # Green
            '#9C27B0',  # Purple
            '#E91E63',  # Pink
        ]
        
        marker_styles = [
            self.plot_config.get('marker_1_marker_style', 'o'),
            self.plot_config.get('marker_2_marker_style', 's'),
            '^', 'D', 'v', '<'  # Additional default styles for extra markers
        ]
        # Honour the Plot Options spinbox if present (0 = hide markers)
        _mk_override = (getattr(self, 'plot_marker_size_spin', None) and
                        self.plot_marker_size_spin.value())
        _mk_sz = self.plot_marker_size_spin.value() if getattr(self, 'plot_marker_size_spin', None) else None
        if _mk_sz is not None:
            marker_sizes = [_mk_sz] * 6
        else:
            marker_sizes = [
                int(self.plot_config.get('marker_1_size', '8')),
                int(self.plot_config.get('marker_2_size', '8')),
                6, 6, 8, 6  # Additional default sizes for extra markers
            ]
        
        available_columns = list(df.columns)
        used_columns = set()
        marker_pairs = []
        
        # Detect marker column pairs using regex
        for dist_pattern, elev_pattern in marker_patterns:
            for col in available_columns:
                if col in used_columns:
                    continue
                    
                dist_match = re.match(dist_pattern, col)
                if dist_match:
                    if dist_match.groups():
                        group = dist_match.group(1)
                        target_elev_pattern = elev_pattern.replace(r'\1', group)
                    else:
                        target_elev_pattern = elev_pattern
                    
                    for elev_col in available_columns:
                        if elev_col in used_columns:
                            continue
                        if re.match(target_elev_pattern, elev_col):
                            marker_pairs.append((col, elev_col))
                            used_columns.add(col)
                            used_columns.add(elev_col)
                            break
        
        # Also check for explicit marker columns
        explicit_marker_pairs = [
            ('Marker_Distance', 'Marker_Height'),
            ('Marker_Distance_2', 'Marker_Height_2'),
            ('Marker_Dist', 'Marker_Elev'),
            ('Marker_Dist_2', 'Marker_Elev_2'),
            ('MARKER_DISTANCE', 'MARKER_HEIGHT'),
            ('marker_distance', 'marker_height'),
        ]
        
        for dist_col, elev_col in explicit_marker_pairs:
            if (dist_col in available_columns and elev_col in available_columns and
                dist_col not in used_columns and elev_col not in used_columns):
                marker_pairs.append((dist_col, elev_col))
                used_columns.add(dist_col)
                used_columns.add(elev_col)
        
        # Plot marker pairs
        for idx, (dist_col, elev_col) in enumerate(marker_pairs):
            mask = df[dist_col].notna() & df[elev_col].notna()
            if mask.any():
                color = marker_colors[idx % len(marker_colors)]
                style = marker_styles[idx % len(marker_styles)]
                size = marker_sizes[idx % len(marker_sizes)]
                
                # FIXED: Always use configuration for marker labels with consistent priority
                if idx == 0:
                    # First marker pair - use markers_1 config
                    label = self.plot_config.get('legend_markers_1', 'Markers 1')
                elif idx == 1:
                    # Second marker pair - use markers_2 config
                    label = self.plot_config.get('legend_markers_2', 'Markers 2')
                elif '2' in dist_col or '_2' in dist_col.lower():
                    # Check for "2" indicators in column name
                    label = self.plot_config.get('legend_markers_2', 'Markers 2')
                elif '1' in dist_col or '_1' in dist_col.lower() or dist_col.lower() == 'marker_distance':
                    # Check for "1" indicators or base marker column
                    label = self.plot_config.get('legend_markers_1', 'Markers 1')
                else:
                    # Fallback for additional marker pairs beyond configured ones
                    label = f"Markers {idx + 1}"
                
                if size == 0:
                    continue  # marker size 0 = hide this marker series
                ax.plot(df[dist_col][mask], df[elev_col][mask],
                       color=color,
                       marker=style,
                       markersize=size,
                       linestyle='none',
                       label=label,
                       markeredgecolor='black',
                       markeredgewidth=0.5,
                       zorder=4)
                
                QgsMessageLog.logMessage(
                    f"Plotted {mask.sum()} marker points from {dist_col}/{elev_col}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )

    # =========================================================================
    # DRILL HOLE UI HELPERS
    # =========================================================================

    def refresh_dh_layer_list(self):
        """Populate the section-line layer combo with all polyline layers in the project."""
        self.dh_layer_combo.clear()
        self.dh_layer_combo.addItem("— select section line layer —", None)
        for lyr in QgsProject.instance().mapLayers().values():
            if lyr.type() == QgsMapLayer.VectorLayer and lyr.geometryType() == QgsWkbTypes.LineGeometry:
                self.dh_layer_combo.addItem(lyr.name(), lyr.id())
        # Keep the geology bands section line combo in sync
        self.refresh_geo_section_line_layers()

    def refresh_geo_section_line_layers(self):
        """Populate the geology bands section line layer combo."""
        if not hasattr(self, 'geo_section_line_combo'):
            return
        current_id = self.geo_section_line_combo.currentData()
        self.geo_section_line_combo.blockSignals(True)
        self.geo_section_line_combo.clear()
        self.geo_section_line_combo.addItem("— select section line layer —", None)
        restore_idx = 0
        for lyr in QgsProject.instance().mapLayers().values():
            if (lyr.type() == QgsMapLayer.VectorLayer and
                    lyr.geometryType() == QgsWkbTypes.LineGeometry):
                self.geo_section_line_combo.addItem(lyr.name(), lyr.id())
                if lyr.id() == current_id:
                    restore_idx = self.geo_section_line_combo.count() - 1
        self.geo_section_line_combo.blockSignals(False)
        self.geo_section_line_combo.setCurrentIndex(restore_idx)

    def refresh_geo_band_layers(self):
        """Populate the geology polygon layer combo."""
        if not hasattr(self, 'geo_band_layer_combo'):
            return
        current_id = self.geo_band_layer_combo.currentData()
        self.geo_band_layer_combo.blockSignals(True)
        self.geo_band_layer_combo.clear()
        self.geo_band_layer_combo.addItem("— select geology polygon layer —", None)
        restore_idx = 0
        for lyr in QgsProject.instance().mapLayers().values():
            if (lyr.type() == QgsMapLayer.VectorLayer and
                    lyr.geometryType() == QgsWkbTypes.PolygonGeometry):
                self.geo_band_layer_combo.addItem(lyr.name(), lyr.id())
                if lyr.id() == current_id:
                    restore_idx = self.geo_band_layer_combo.count() - 1
        self.geo_band_layer_combo.blockSignals(False)
        self.geo_band_layer_combo.setCurrentIndex(restore_idx)
        self.refresh_geo_band_fields()

    def refresh_geo_band_fields(self):
        """Populate the label-field combo from the currently selected geology layer."""
        if not hasattr(self, 'geo_band_field_combo'):
            return
        self.geo_band_field_combo.clear()
        layer_id = self.geo_band_layer_combo.currentData() if hasattr(self, 'geo_band_layer_combo') else None
        if not layer_id:
            return
        layer = QgsProject.instance().mapLayer(layer_id)
        if not layer:
            return
        for field in layer.fields():
            self.geo_band_field_combo.addItem(field.name())

    def refresh_plate_layers(self):
        """Populate the plate locations point layer combo."""
        if not hasattr(self, 'plate_layer_combo'):
            return
        current_id = self.plate_layer_combo.currentData()
        self.plate_layer_combo.blockSignals(True)
        self.plate_layer_combo.clear()
        self.plate_layer_combo.addItem("\u2014 select plate locations layer \u2014", None)
        restore_idx = 0
        for lyr in QgsProject.instance().mapLayers().values():
            if (lyr.type() == QgsMapLayer.VectorLayer and
                    lyr.geometryType() == QgsWkbTypes.PointGeometry):
                self.plate_layer_combo.addItem(lyr.name(), lyr.id())
                if lyr.id() == current_id:
                    restore_idx = self.plate_layer_combo.count() - 1
        self.plate_layer_combo.blockSignals(False)
        self.plate_layer_combo.setCurrentIndex(restore_idx)

    def refresh_peg_layers(self):
        """Populate the interpreted pegmatites polygon layer combo."""
        if not hasattr(self, 'peg_band_layer_combo'):
            return
        current_id = self.peg_band_layer_combo.currentData()
        self.peg_band_layer_combo.blockSignals(True)
        self.peg_band_layer_combo.clear()
        self.peg_band_layer_combo.addItem("\u2014 select pegmatite polygon layer —", None)
        restore_idx = 0
        for lyr in QgsProject.instance().mapLayers().values():
            if (lyr.type() == QgsMapLayer.VectorLayer and
                    lyr.geometryType() == QgsWkbTypes.PolygonGeometry):
                self.peg_band_layer_combo.addItem(lyr.name(), lyr.id())
                if lyr.id() == current_id:
                    restore_idx = self.peg_band_layer_combo.count() - 1
        self.peg_band_layer_combo.blockSignals(False)
        self.peg_band_layer_combo.setCurrentIndex(restore_idx)

    def refresh_struct_layers(self):
        """Populate the structural measurements point layer combo."""
        if not hasattr(self, 'struct_layer_combo'):
            return
        current_id = self.struct_layer_combo.currentData()
        self.struct_layer_combo.blockSignals(True)
        self.struct_layer_combo.clear()
        self.struct_layer_combo.addItem('— select structure point layer —', None)
        restore_idx = 0
        for lyr in QgsProject.instance().mapLayers().values():
            if (lyr.type() == QgsMapLayer.VectorLayer and
                    lyr.geometryType() == QgsWkbTypes.PointGeometry):
                self.struct_layer_combo.addItem(lyr.name(), lyr.id())
                if lyr.id() == current_id:
                    restore_idx = self.struct_layer_combo.count() - 1
        self.struct_layer_combo.blockSignals(False)
        self.struct_layer_combo.setCurrentIndex(restore_idx)
        self.refresh_struct_fields()

    def refresh_struct_fields(self):
        """Populate the dip and dip-direction field combos from the selected structure layer."""
        if not hasattr(self, 'struct_dip_combo'):
            return
        self.struct_dip_combo.clear()
        self.struct_dir_combo.clear()
        layer_id = self.struct_layer_combo.currentData() if hasattr(self, 'struct_layer_combo') else None
        if not layer_id:
            return
        layer = QgsProject.instance().mapLayer(layer_id)
        if not layer:
            return
        import re as _re
        dip_pat = _re.compile(r'(?i)\bdip\b')
        dir_pat = _re.compile(r'(?i)(dip.?dir|dip.?az|strike|azimuth|bearing)')
        first_dip = first_dir = None
        for field in layer.fields():
            name = field.name()
            self.struct_dip_combo.addItem(name)
            self.struct_dir_combo.addItem(name)
            if first_dip is None and dip_pat.search(name):
                first_dip = name
            if first_dir is None and dir_pat.search(name):
                first_dir = name
        # Auto-select the best-matching fields
        if first_dip:
            idx = self.struct_dip_combo.findText(first_dip)
            if idx >= 0:
                self.struct_dip_combo.setCurrentIndex(idx)
        if first_dir:
            idx = self.struct_dir_combo.findText(first_dir)
            if idx >= 0:
                self.struct_dir_combo.setCurrentIndex(idx)

    def refresh_dh_raster_list(self):
        """
        Populate the LiDAR/DEM raster combo.

        Only single-band rasters are listed — multi-band imagery (e.g. Google
        Satellite, Mt_Elvire_2) is excluded because sampling Band 1 of an RGB
        image gives a pixel intensity (0-255) not an elevation.

        The combo also adds dem_clipped.tif as a direct file entry if it is
        not already present as a project layer, and auto-selects the best
        candidate so the user does not have to pick manually.
        """
        import os as _os

        self.dh_raster_combo.clear()
        self.dh_raster_combo.addItem("— select LiDAR / DEM raster —", None)

        dem_path = r"C:/QGIS/Mt Elvire/data/rasters/dem_clipped.tif"
        project_dem_ids = set()   # track which single-band layers are already listed
        best_idx = None           # index of preferred auto-selection

        # ── single-band project layers (potential DEMs only) ─────────────
        for lyr in QgsProject.instance().mapLayers().values():
            if lyr.type() != QgsMapLayer.RasterLayer:
                continue
            bands = lyr.bandCount()
            if bands != 1:
                # Skip multi-band imagery — it is NOT a DEM
                QgsMessageLog.logMessage(
                    f"DrillHoles: skipping '{lyr.name()}' ({bands} bands) – not a DEM",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                continue
            label = f"{lyr.name()}  (1 band)"
            idx = self.dh_raster_combo.count()
            self.dh_raster_combo.addItem(label, lyr.id())
            project_dem_ids.add(lyr.id())
            # Prefer a layer whose source matches dem_clipped.tif
            if best_idx is None or 'dem_clipped' in lyr.name().lower() or \
               dem_path.replace('\\', '/') in lyr.source().replace('\\', '/'):
                best_idx = idx

        # ── add local dem_clipped.tif if not already listed ──────────────
        if _os.path.isfile(dem_path):
            already_listed = any(
                dem_path.replace('\\', '/') in
                (QgsProject.instance().mapLayer(lid).source().replace('\\', '/')
                 if QgsProject.instance().mapLayer(lid) else '')
                for lid in project_dem_ids
            )
            if not already_listed:
                idx = self.dh_raster_combo.count()
                self.dh_raster_combo.addItem("dem_clipped.tif  (local file)", dem_path)
                if best_idx is None:
                    best_idx = idx

        # ── auto-select best candidate ────────────────────────────────────
        if best_idx is not None:
            self.dh_raster_combo.setCurrentIndex(best_idx)

    def _resolve_dem_raster(self):
        """
        Resolve and return the DEM QgsRasterLayer to use for collar RL sampling.

        Call this ONCE per plot operation and pass the result to
        _sample_raster_rl().  This avoids reopening the file for every collar.

        Guards against multi-band imagery being selected: if the chosen layer
        has more than 1 band it is rejected and the method falls back to the
        hardcoded dem_clipped.tif path.

        Returns QgsRasterLayer or None.
        """
        from qgis.core import QgsRasterLayer
        import os as _os

        fallback_path = r"C:/QGIS/Mt Elvire/data/rasters/dem_clipped.tif"
        raster_ref = self.dh_raster_combo.currentData()

        raster = None

        if isinstance(raster_ref, str) and raster_ref.endswith('.tif'):
            # Local file path entry
            raster = QgsRasterLayer(raster_ref, "_dh_dem")
            if not raster.isValid():
                raster = None
        elif raster_ref:
            # Project layer id
            raster = QgsProject.instance().mapLayer(raster_ref)

        # Guard: reject imagery (multi-band rasters)
        if raster and raster.bandCount() != 1:
            QgsMessageLog.logMessage(
                f"DrillHoles: selected raster '{raster.name()}' has "
                f"{raster.bandCount()} bands – this is likely imagery, not a DEM. "
                "Falling back to dem_clipped.tif.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            raster = None

        # Fallback to dem_clipped.tif if nothing usable is selected
        if raster is None:
            if _os.path.isfile(fallback_path):
                raster = QgsRasterLayer(fallback_path, "_dh_dem")
                if raster.isValid():
                    QgsMessageLog.logMessage(
                        "DrillHoles: no valid single-band raster selected – "
                        "automatically using dem_clipped.tif.",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                else:
                    raster = None

        return raster

    def _sample_raster_rl(self, easting: float, northing: float,
                          raster=None) -> float:
        """
        Sample *raster* at (easting, northing) and return the elevation.

        Parameters
        ----------
        easting, northing : float
            Collar coordinates in the CRS chosen by ``self.dh_crs_combo``.
        raster : QgsRasterLayer, optional
            Pre-resolved DEM raster. Pass the value returned by
            ``_resolve_dem_raster()`` so the file is not reopened per collar.
            When None the method resolves it inline (slower).

        Returns
        -------
        float or None
        """
        from qgis.core import (QgsPointXY, QgsCoordinateReferenceSystem,
                               QgsCoordinateTransform, QgsRaster)

        if raster is None:
            raster = self._resolve_dem_raster()
        if raster is None or not raster.isValid():
            return None

        # ── build coordinate transform ────────────────────────────────────
        collar_epsg = self.dh_crs_combo.currentData()   # e.g. "EPSG:28350"
        collar_crs  = QgsCoordinateReferenceSystem(collar_epsg)
        raster_crs  = raster.crs()

        pt = QgsPointXY(easting, northing)

        if collar_crs != raster_crs:
            xform = QgsCoordinateTransform(collar_crs, raster_crs, QgsProject.instance())
            try:
                pt = xform.transform(pt)
            except Exception as exc:
                QgsMessageLog.logMessage(
                    f"DrillHoles: coordinate transform failed – {exc}",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )
                return None

        # ── sample raster band 1 ─────────────────────────────────────────
        result = raster.dataProvider().identify(pt, QgsRaster.IdentifyFormatValue)
        if result.isValid():
            val = result.results().get(1)
            nodata = raster.dataProvider().sourceNoDataValue(1)
            # Guard: WMS/XYZ tiles return isValid=True but empty results dict
            if val is None:
                pass
            elif nodata is not None and not (nodata != nodata) and val == nodata:
                # nodata match (handles float nodata; nodata!=nodata catches NaN)
                pass
            else:
                return float(val)

        QgsMessageLog.logMessage(
            f"DrillHoles: no valid raster value at E={easting:.1f} N={northing:.1f} "
            f"(raster: {raster.name()}, CRS: {raster.crs().authid()})",
            "IntegratedProfileAnalyzer", Qgis.Warning
        )
        return None

    def browse_drillhole_folder(self):
        """Open a folder browser to choose the drill log folder."""
        folder = QFileDialog.getExistingDirectory(
            self, "Select Drill Log Folder", self.last_directory
        )
        if folder:
            self.dh_folder_edit.setText(folder)
            self.last_directory = folder
            self.settings.setValue("lastDirectory", self.last_directory)

    def browse_planned_holes_csv(self):
        """Open a file browser to choose the planned collars CSV."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Planned Collars CSV", self.last_directory,
            "CSV files (*.csv);;All files (*.*)"
        )
        if path:
            self.planned_csv_edit.setText(path)
            self.last_directory = os.path.dirname(path)
            self.settings.setValue("lastDirectory", self.last_directory)

    # =========================================================================
    # DRILL HOLE PROJECTION + PLOTTING
    # =========================================================================

    def _load_planned_collars_csv(self):
        """
        Load planned collar data from a CSV file.

        Expected columns (case-insensitive, flexible spacing):
            Hole_ID, Easting, Northing, Azimuth, Dip, TD

        Alternate header spellings accepted:
            Easting (m E) / Easting_mE / East
            Northing (m N) / Northing_mN / North
            Azimuth (deg) / Azimuth_deg / Azi
            Dip (deg) / Dip_deg
            TD (m) / TD_m / Total_Depth / MaxDepth / Max_Depth

        Returns a list of dicts: {id, east, north, azimuth, dip, td}
        """
        import csv as _csv
        import re as _re

        path = getattr(self, 'planned_csv_edit', None)
        path = path.text().strip() if path else ''
        if not path or not os.path.isfile(path):
            return []

        def _norm(h):
            return _re.sub(r'[^a-z0-9]', '', h.lower())

        def _find_col(headers_norm, patterns):
            for pat in patterns:
                for i, h in enumerate(headers_norm):
                    if _re.search(pat, h):
                        return i
            return None

        collars = []
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = _csv.reader(f)
                raw_headers = next(reader)
                hn = [_norm(h) for h in raw_headers]

                i_id  = _find_col(hn, [r'^holeid$', r'^hole_?id$', r'^id$'])
                i_e   = _find_col(hn, [r'^easting', r'^east'])
                i_n   = _find_col(hn, [r'^northing', r'^north'])
                i_azi = _find_col(hn, [r'^azimuth', r'^azi', r'^bearing'])
                i_dip = _find_col(hn, [r'^dip'])
                i_td  = _find_col(hn, [r'^tdm$', r'^td', r'^totaldepth', r'^maxdepth', r'^max_?depth'])

                missing = [n for n, i in [('Hole_ID', i_id), ('Easting', i_e),
                                           ('Northing', i_n), ('Azimuth', i_azi),
                                           ('Dip', i_dip), ('TD', i_td)] if i is None]
                if missing:
                    QgsMessageLog.logMessage(
                        f"PlannedHoles: CSV is missing required columns: {missing}. "
                        f"Headers found: {raw_headers}",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                    return []

                for row in reader:
                    if not row or all(c.strip() == '' for c in row):
                        continue
                    try:
                        collars.append({
                            'id':      str(row[i_id]).strip(),
                            'east':    float(row[i_e]),
                            'north':   float(row[i_n]),
                            'azimuth': float(row[i_azi]),
                            'dip':     float(row[i_dip]),
                            'td':      float(row[i_td]),
                        })
                    except (ValueError, IndexError):
                        continue
        except Exception as exc:
            QgsMessageLog.logMessage(
                f"PlannedHoles: error reading CSV '{path}': {exc}",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return []

        QgsMessageLog.logMessage(
            f"PlannedHoles: loaded {len(collars)} planned hole(s) from CSV.",
            "IntegratedProfileAnalyzer", Qgis.Info
        )
        return collars

    def _plot_planned_drillholes(self, ax, sheet_name: str):
        """
        Project planned (straight) drill holes onto the cross-section and draw
        them as dashed traces.

        Each planned hole is assumed straight (no deviation) from the collar
        in the direction of its Azimuth / Dip to Total Depth.

        Plotted style: dashed line, open-circle collar symbol, grey colour
        palette so they are visually distinct from drilled holes.
        """
        import math

        if not getattr(self, 'planned_enable_chk', None) or not self.planned_enable_chk.isChecked():
            return

        collars = self._load_planned_collars_csv()
        if not collars:
            return

        # ── section line ──────────────────────────────────────────────────
        sec_layer_id = (self.dh_layer_combo.currentData()
                        if getattr(self, 'dh_layer_combo', None) else None)
        if not sec_layer_id:
            return
        section = self._get_section_line_from_layer(sheet_name, sec_layer_id)
        if section is None:
            return

        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return
        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        corridor = (self.planned_corridor_spin.value()
                    if getattr(self, 'planned_corridor_spin', None) else 200.0)
        show_label = (getattr(self, 'planned_label_chk', None) and
                      self.planned_label_chk.isChecked())

        # ── CRS: planned CSV always in the same CRS as dh_crs_combo ───────
        collar_epsg = (self.dh_crs_combo.currentData()
                       if getattr(self, 'dh_crs_combo', None) else 'EPSG:28350')
        collar_crs  = QgsCoordinateReferenceSystem(collar_epsg)
        sec_lyr     = QgsProject.instance().mapLayer(sec_layer_id)
        section_crs = sec_lyr.crs() if sec_lyr else QgsProject.instance().crs()
        needs_xform = (collar_crs.isValid() and section_crs.isValid() and
                       collar_crs.authid() != section_crs.authid())
        crs_xform   = (QgsCoordinateTransform(collar_crs, section_crs, QgsProject.instance())
                       if needs_xform else None)

        # DEM raster for collar RL
        dem_raster = self._resolve_dem_raster()

        # Grey palette — distinct from the drilled hole palette
        _PLANNED_COLOURS = [
            '#555555', '#888888', '#AAAAAA', '#333333', '#666666',
            '#999999', '#BBBBBB', '#444444', '#777777', '#CCCCCC',
        ]

        plotted = 0
        for col_idx, collar in enumerate(collars):
            colour = _PLANNED_COLOURS[col_idx % len(_PLANNED_COLOURS)]

            east_proj, north_proj = collar['east'], collar['north']
            if crs_xform is not None:
                try:
                    pt = crs_xform.transform(QgsPointXY(east_proj, north_proj))
                    east_proj, north_proj = pt.x(), pt.y()
                except Exception:
                    continue

            # Corridor check
            vec_e = east_proj - E0
            vec_n = north_proj - N0
            chainage = vec_e * ux + vec_n * uy
            perp     = vec_e * (-uy) + vec_n * ux

            QgsMessageLog.logMessage(
                f"PlannedHoles: {collar['id']} chainage={chainage:.0f}m "
                f"perp={perp:.0f}m corridor=\u00b1{corridor:.0f}m "
                f"\u2192 {'IN' if abs(perp) <= corridor else 'OUT'}",
                "IntegratedProfileAnalyzer", Qgis.Info
            )

            if abs(perp) > corridor:
                continue

            # RL from DEM
            collar_rl = self._sample_raster_rl(collar['east'], collar['north'],
                                               raster=dem_raster)
            if collar_rl is None:
                QgsMessageLog.logMessage(
                    f"PlannedHoles: {collar['id']} – could not resolve RL; skipping.",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )
                continue

            # Straight-hole trace (single interval)
            azi_rad      = math.radians(collar['azimuth'])
            abs_dip_rad  = math.radians(abs(collar['dip']))
            td           = collar['td']

            horiz   = td * math.cos(abs_dip_rad)
            d_east  = horiz * math.sin(azi_rad)
            d_north = horiz * math.cos(azi_rad)
            d_down  = td   * math.sin(abs_dip_rad)

            toe_chainage = chainage + d_east * ux + d_north * uy
            toe_rl       = collar_rl - d_down

            # Plot dashed trace
            ax.plot([chainage, toe_chainage], [collar_rl, toe_rl],
                    color=colour,
                    linestyle='--',
                    linewidth=1.5,
                    zorder=4,
                    label=f'Planned: {collar["id"]}' if plotted == 0 else '_nolegend_')

            # Open-circle collar symbol
            ax.plot(chainage, collar_rl,
                    marker='o',
                    markersize=7,
                    color='white',
                    markeredgecolor=colour,
                    markeredgewidth=1.5,
                    zorder=5)

            # Collar label
            if show_label:
                ax.annotate(
                    collar['id'],
                    xy=(chainage, collar_rl),
                    xytext=(4, 4),
                    textcoords='offset points',
                    fontsize=7,
                    color=colour,
                    zorder=6,
                    clip_on=True
                )

            plotted += 1

        if plotted:
            QgsMessageLog.logMessage(
                f"PlannedHoles: drew {plotted} planned hole(s) on section '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Info
            )

    def _load_collar_data(self):
        """
        Load collar data from a folder of per-hole .xlsx drill log files.

        Each .xlsx must follow the RC log format with sheets:
          'Collar'   — one data row: Hole_ID, Easting, Northing, Max_Depth, (Orig_RL optional)
          'DHSurvey' — rows: Depth, Dip, Orig_Azimuth for downhole deviation survey

        Returns a list of collar dicts with keys:
            id, east, north, rl, td, surveys
        where 'surveys' is a sorted list of (depth_m, dip_deg, azimuth_deg).
        """
        import glob

        folder = getattr(self, 'dh_folder_edit', None)
        folder = folder.text().strip() if folder else ''
        if not folder or not os.path.isdir(folder):
            return []

        xlsx_files = sorted(glob.glob(os.path.join(folder, '*.xlsx')))
        if not xlsx_files:
            QgsMessageLog.logMessage(
                "DrillHoles: no .xlsx files found in the selected folder.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return []

        collars = []
        for fpath in xlsx_files:
            if os.path.basename(fpath).startswith('~'):
                continue  # skip temp/lock files
            try:
                collar = self._read_drill_log_xlsx(fpath)
                if collar:
                    collars.append(collar)
                    QgsMessageLog.logMessage(
                        f"DrillHoles: loaded {collar['id']} from {os.path.basename(fpath)}",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
            except Exception as exc:
                QgsMessageLog.logMessage(
                    f"DrillHoles: error reading {os.path.basename(fpath)} – {exc}",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )

        QgsMessageLog.logMessage(
            f"DrillHoles: loaded {len(collars)} hole(s) from folder.",
            "IntegratedProfileAnalyzer", Qgis.Info
        )
        return collars

    def _read_drill_log_xlsx(self, fpath):
        """
        Parse a single drill log .xlsx file and return a collar dict or None.

        Expected sheets:
          'Collar'   — first data row: Hole_ID, Easting, Northing, Max_Depth, (Orig_RL)
          'DHSurvey' — rows: Depth, Dip, Orig_Azimuth

        Returns dict with keys: id, east, north, rl, td, surveys
        where surveys = [(depth, dip, azimuth), ...]
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required to read .xlsx drill log files")

        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

        # ── Collar sheet ──────────────────────────────────────────────────
        if 'Collar' not in wb.sheetnames:
            return None

        ws_collar = wb['Collar']
        collar_rows = []
        for row in ws_collar.iter_rows(values_only=True, max_row=10):
            collar_rows.append(row)

        if len(collar_rows) < 2:
            return None

        headers  = [str(h).strip() if h is not None else '' for h in collar_rows[0]]
        data_row = collar_rows[1]

        # Case-insensitive header → column-index map (first occurrence wins)
        hmap = {}
        for i, h in enumerate(headers):
            key = h.lower().strip()
            if key and key not in hmap:
                hmap[key] = i

        def _get(col_names):
            for c in col_names:
                if c in hmap:
                    idx = hmap[c]
                    if idx < len(data_row):
                        return data_row[idx]
            return None

        hole_id = _get(['hole_id', 'holeid'])
        if hole_id is None:
            return None
        hole_id = str(hole_id).strip()
        if not hole_id:
            return None

        max_depth = _get(['max_depth', 'maxdepth', 'td', 'total_depth'])
        if max_depth is None:
            return None
        try:
            max_depth = float(max_depth)
        except (TypeError, ValueError):
            return None

        # Use final GPS Easting/Northing columns first; fall back to Orig_East/North
        east  = _get(['easting', 'east_', 'east', 'utm_easting', 'e'])
        north = _get(['northing', 'north_', 'north', 'utm_northing', 'n'])
        if east is None:
            east  = _get(['orig_east'])
        if north is None:
            north = _get(['orig_north'])

        if east is None or north is None:
            return None
        try:
            east, north = float(east), float(north)
        except (TypeError, ValueError):
            return None

        # RL — often absent; will be resolved from raster later
        rl = None
        rl_raw = _get(['orig_rl', 'nat_rl', 'local_rl', 'rl', 'elevation', 'elev'])
        if rl_raw is not None:
            try:
                rl = float(rl_raw)
            except (TypeError, ValueError):
                rl = None

        # ── DHSurvey sheet ────────────────────────────────────────────────
        surveys = []
        if 'DHSurvey' in wb.sheetnames:
            ws_surv  = wb['DHSurvey']
            surv_rows = list(ws_surv.iter_rows(values_only=True))
            if len(surv_rows) > 1:
                surv_hdrs = [
                    str(h).strip().lower() if h is not None else ''
                    for h in surv_rows[0]
                ]
                shmap = {h: i for i, h in enumerate(surv_hdrs) if h}

                seen = {}  # depth → (dip, azi) — first reading per depth
                for row in surv_rows[1:]:
                    if row is None or all(v is None for v in row):
                        continue
                    try:
                        depth = float(row[shmap['depth']])
                        dip   = float(row[shmap['dip']])
                        # Accept 'orig_azimuth' or plain 'azimuth'
                        azi_col = 'orig_azimuth' if 'orig_azimuth' in shmap else 'azimuth'
                        azi   = float(row[shmap[azi_col]])
                        if depth not in seen:
                            seen[depth] = (dip, azi)
                    except (KeyError, TypeError, ValueError):
                        continue

                surveys = sorted(
                    [(d, v[0], v[1]) for d, v in seen.items()],
                    key=lambda x: x[0]
                )

        wb.close()

        if not surveys:
            surveys = [(0.0, -90.0, 0.0)]  # assume vertical if no survey data

        return {
            'id':      hole_id,
            'east':    east,
            'north':   north,
            'rl':      rl,
            'td':      max_depth,
            'surveys': surveys,
            'fpath':   fpath,
        }

    def _read_drill_geology(self, fpath, sheet='Geology',
                            from_col='mFrom', to_col='mTo', code_col='Lith1_Code'):
        """
        Read lithology intervals from a drill log .xlsx file.

        Returns a list of (from_depth, to_depth, code) tuples sorted by from_depth.
        Returns [] if the sheet or columns are missing.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if sheet not in wb.sheetnames:
                wb.close()
                QgsMessageLog.logMessage(
                    f"DrillGeo: sheet '{sheet}' not found in {os.path.basename(fpath)}. "
                    f"Available: {wb.sheetnames}",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )
                return []

            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                return []

            headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
            hmap = {h: i for i, h in enumerate(headers) if h}

            from_idx = hmap.get(from_col.lower())
            to_idx   = hmap.get(to_col.lower())
            code_idx = hmap.get(code_col.lower())

            if from_idx is None or to_idx is None or code_idx is None:
                QgsMessageLog.logMessage(
                    f"DrillGeo: sheet '{sheet}' in {os.path.basename(fpath)} missing "
                    f"'{from_col}'/'{to_col}'/'{code_col}'. Available: {list(hmap.keys())}",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )
                return []

            intervals = []
            for row in rows[1:]:
                if row is None or len(row) <= max(from_idx, to_idx, code_idx):
                    continue
                try:
                    from_d = float(row[from_idx]) if row[from_idx] is not None else None
                    to_d   = float(row[to_idx])   if row[to_idx]   is not None else None
                    code   = str(row[code_idx]).strip() if row[code_idx] is not None else ''
                    if from_d is not None and to_d is not None and to_d > from_d:
                        intervals.append((from_d, to_d, code))
                except (TypeError, ValueError):
                    continue

            return sorted(intervals, key=lambda x: x[0])

        except Exception as exc:
            QgsMessageLog.logMessage(
                f"DrillGeo: error reading {os.path.basename(fpath)}: {exc}",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return []

    def _get_section_line_from_layer(self, sheet_name: str, layer_id: str):
        """
        Return (start_pt, end_pt) as (E, N) tuples for the section line that
        corresponds to *sheet_name* by searching *layer_id*.

        Matching strategy (in order):
          1. Case-insensitive exact match of sheet_name against every string
             field on every feature.
          2. Positional fallback: if sheet_name ends in a number N (e.g.
             "Polyline 1", "Section 3"), select the Nth feature (1-based) from
             the layer ordered by feature ID.  This handles the common case
             where the section line layer uses a different naming convention
             than the Excel sheet names (e.g. "Columba Section 01" vs
             "Polyline 1").

        Returns None if no match is found via either strategy.
        """
        import re as _re

        layer = QgsProject.instance().mapLayer(layer_id)
        if not layer:
            return None

        target = sheet_name.strip().lower()
        fields = layer.fields()
        # Indices of all string/text fields (we search these for the name match)
        str_field_indices = [
            i for i in range(fields.count())
            if fields.field(i).type() in (
                QVariant.String,  # Qt string type
            )
        ]
        # Always include field index 0 as a fallback search target
        if 0 not in str_field_indices:
            str_field_indices = [0] + str_field_indices

        # ── Strategy 1: exact string-field match ─────────────────────────
        matched_feature = None
        all_features = []   # collected once for potential positional fallback
        for feat in layer.getFeatures():
            all_features.append(feat)
            if matched_feature is not None:
                continue
            for idx in str_field_indices:
                try:
                    val = feat.attribute(idx)
                    if val is not None and str(val).strip().lower() == target:
                        matched_feature = feat
                        break
                except Exception:
                    continue

        # ── Strategy 2: positional fallback ──────────────────────────────
        if matched_feature is None:
            pos_m = _re.search(r'(\d+)\s*$', sheet_name.strip())
            if pos_m:
                one_based = int(pos_m.group(1))
                if 1 <= one_based <= len(all_features):
                    matched_feature = all_features[one_based - 1]
                    QgsMessageLog.logMessage(
                        f"SectionLine: no feature name matched '{sheet_name}' in "
                        f"layer '{layer.name()}'. Using positional fallback: "
                        f"feature #{one_based} → "
                        f"'{matched_feature.attribute(str_field_indices[0]) if str_field_indices else '?'}'.",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )

        if matched_feature is None:
            return None

        geom = matched_feature.geometry()
        if geom.isEmpty():
            return None

        # Get first and last vertex of the (possibly multi-part) line
        vertices = list(geom.vertices())
        if len(vertices) < 2:
            return None

        p0, p1 = vertices[0], vertices[-1]
        return (p0.x(), p0.y()), (p1.x(), p1.y())

    def _get_section_line(self, sheet_name: str):
        """
        Return (start_pt, end_pt) as (E, N) tuples for the section line that
        corresponds to *sheet_name*.

        Delegates to _get_section_line_from_layer using dh_layer_combo.
        Returns None if no suitable layer/feature is found.
        """
        layer_id = self.dh_layer_combo.currentData()
        if not layer_id:
            return None

        result = self._get_section_line_from_layer(sheet_name, layer_id)
        if result is None:
            layer = QgsProject.instance().mapLayer(layer_id)
            layer_name = layer.name() if layer else layer_id
            QgsMessageLog.logMessage(
                f"DrillHoles: no section line feature matches sheet name "
                f"'{sheet_name}' in layer '{layer_name}'. "
                f"Holes will not be plotted for this section. "
                f"Ensure a feature's name field matches the sheet name exactly.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
        return result

    def _project_collar_onto_section(self, collar, section):
        """
        Project a collar onto a section line.

        Parameters
        ----------
        collar  : dict  with keys 'east', 'north', 'rl', 'id', 'azi', 'dip', 'td'
        section : ((E0,N0), (E1,N1))  section line start / end

        Returns
        -------
        dict with:
            'chainage'  – distance along section line (metres)
            'perp'      – signed perpendicular distance from section (metres)
            'rl'        – collar elevation (pass-through)
            ...plus all original collar fields
        Returns None if the collar is outside the corridor.
        """
        import math

        (E0, N0), (E1, N1) = section
        dE, dN = E1 - E0, N1 - N0
        section_length = math.hypot(dE, dN)
        if section_length == 0:
            return None

        # Unit vector along section
        ux, uy = dE / section_length, dN / section_length

        # Vector from section start to collar
        vx = collar['east'] - E0
        vy = collar['north'] - N0

        # Chainage = dot product with unit vector
        chainage = vx * ux + vy * uy

        # Perpendicular distance = cross product magnitude (signed)
        perp = vx * (-uy) + vy * ux

        result = dict(collar)
        result['chainage'] = chainage
        result['perp'] = perp
        return result

    def _plot_drillholes(self, ax, sheet_name: str):
        """
        Project and draw deviated drill hole traces onto the cross-section axes.

        Each collar within the corridor is drawn as:
          • A filled circle at the projected collar position (chainage, RL)
          • A polyline trace of the deviated path projected onto the section
          • Optional Hole ID label at the collar
          • Optional TD label at the bottom-of-hole

        Deviated trace:
            The balanced-tangent method is used.  For each survey interval the
            average dip/azimuth of the two endpoints is used to compute a 3D
            displacement vector, which is then projected onto the section plane.

        Collar RL is resolved in priority order:
          1. If 'Always use raster RL' is ticked → sample LiDAR/DEM raster.
          2. Else if the xlsx file has an RL value → use that.
          3. Else                                  → sample LiDAR/DEM raster.
        """
        import math

        if not getattr(self, 'dh_enable_chk', None) or not self.dh_enable_chk.isChecked():
            return

        collars = self._load_collar_data()
        if not collars:
            QgsMessageLog.logMessage(
                "DrillHoles: _load_collar_data returned no collars – check folder path and .xlsx files.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        section = self._get_section_line(sheet_name)
        if section is None:
            QgsMessageLog.logMessage(
                f"DrillHoles: could not determine section line for sheet '{sheet_name}'. "
                "Check that a polyline layer is selected.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return

        # Unit vector along section (E, N components)
        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        corridor   = self.dh_corridor_spin.value()
        show_id    = self.dh_label_id_chk.isChecked()
        show_td    = self.dh_label_td_chk.isChecked()
        use_raster = (getattr(self, 'dh_raster_override_chk', None) and
                      self.dh_raster_override_chk.isChecked())

        # ── Drill trace geology config ────────────────────────────────────
        geo_trace_enabled = (getattr(self, 'dh_geo_chk', None) and
                             self.dh_geo_chk.isChecked())
        geo_sheet    = (self.dh_geo_sheet_edit.text().strip()
                        if getattr(self, 'dh_geo_sheet_edit', None) else 'Geology')
        geo_from_col = (self.dh_geo_from_edit.text().strip()
                        if getattr(self, 'dh_geo_from_edit', None) else 'mFrom')
        geo_to_col   = (self.dh_geo_to_edit.text().strip()
                        if getattr(self, 'dh_geo_to_edit', None) else 'mTo')
        geo_code_col = (self.dh_geo_code_edit.text().strip()
                        if getattr(self, 'dh_geo_code_edit', None) else 'Lith1_Code')

        # Colour palette for auto-assigning distinct colours to lithology codes
        _LITH_PALETTE = [
            '#e6194b', '#3cb44b', '#4363d8', '#f58231', '#911eb4',
            '#42d4f4', '#f032e6', '#bfef45', '#9A6324', '#469990',
            '#dcbeff', '#800000', '#aaffc3', '#808000', '#ffd8b1',
            '#000075', '#a9a9a9', '#ffe119', '#fabebe', '#008080',
        ]
        lith_colour_map = {}   # code → hex colour (shared across all holes this call)

        # ── CRS transform: collar CRS → section layer CRS ─────────────────
        collar_epsg = (self.dh_crs_combo.currentData()
                       if getattr(self, 'dh_crs_combo', None) else 'EPSG:28350')
        collar_crs  = QgsCoordinateReferenceSystem(collar_epsg)

        section_layer_id = (self.dh_layer_combo.currentData()
                            if getattr(self, 'dh_layer_combo', None) else None)
        if section_layer_id:
            sec_layer   = QgsProject.instance().mapLayer(section_layer_id)
            section_crs = sec_layer.crs() if sec_layer else QgsProject.instance().crs()
        else:
            section_crs = QgsProject.instance().crs()

        needs_xform = (collar_crs.isValid() and section_crs.isValid() and
                       collar_crs.authid() != section_crs.authid())
        crs_xform = (QgsCoordinateTransform(collar_crs, section_crs, QgsProject.instance())
                     if needs_xform else None)

        QgsMessageLog.logMessage(
            f"DrillHoles: {len(collars)} collar(s) loaded. "
            f"Section {sheet_name}: ({E0:.0f},{N0:.0f})→({E1:.0f},{N1:.1f}) "
            f"len={sec_len:.0f}m corridor=±{corridor}m. "
            f"Collar CRS={collar_epsg} Section CRS={section_crs.authid()} "
            f"xform={'yes' if needs_xform else 'no'}",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

        dem_raster = self._resolve_dem_raster() if use_raster else None

        # Colour palette — one colour per hole
        _COLOURS = [
            '#8B0000', '#00008B', '#006400', '#8B4513', '#4B0082',
            '#FF6B00', '#005F5F', '#6B0068', '#3B3B00', '#00446B',
        ]

        plotted = 0
        for col_idx, collar in enumerate(collars):
            colour = _COLOURS[col_idx % len(_COLOURS)]

            # ── CRS transform ────────────────────────────────────────────
            east_proj, north_proj = collar['east'], collar['north']
            if crs_xform is not None:
                try:
                    pt = crs_xform.transform(QgsPointXY(east_proj, north_proj))
                    east_proj, north_proj = pt.x(), pt.y()
                except Exception as e:
                    QgsMessageLog.logMessage(
                        f"DrillHoles: CRS transform failed for {collar['id']}: {e}",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                    continue

            # ── check collar is within corridor ──────────────────────────
            vec_e = east_proj - E0
            vec_n = north_proj - N0
            collar_chainage = vec_e * ux + vec_n * uy
            collar_perp     = vec_e * (-uy) + vec_n * ux

            QgsMessageLog.logMessage(
                f"DrillHoles: {collar['id']} E={east_proj:.0f} N={north_proj:.0f} "
                f"chainage={collar_chainage:.0f}m perp={collar_perp:.0f}m "
                f"(corridor ±{corridor}m) → {'IN' if abs(collar_perp) <= corridor else 'OUT'}",
                "IntegratedProfileAnalyzer", Qgis.Info
            )

            if abs(collar_perp) > corridor:
                continue

            # ── resolve collar RL ─────────────────────────────────────────
            # NOTE: _sample_raster_rl expects the *original* collar CRS coords
            # (from dh_crs_combo) because it does its own internal transform to
            # the raster CRS.  Do NOT pass the already-projected east_proj/north_proj.
            collar_rl = collar['rl']
            if use_raster or collar_rl is None:
                _raster = dem_raster if dem_raster is not None else self._resolve_dem_raster()
                sampled = self._sample_raster_rl(collar['east'], collar['north'], raster=_raster)
                if sampled is not None:
                    collar_rl = sampled
                    QgsMessageLog.logMessage(
                        f"DrillHoles: {collar['id']} RL from raster = {sampled:.1f} mRL",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                elif collar_rl is None:
                    QgsMessageLog.logMessage(
                        f"DrillHoles: {collar['id']} – no RL available; skipping.",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                    continue

            # ── build deviated trace (balanced-tangent method) ───────────
            surveys = list(collar['surveys'])   # [(depth, dip, azi), ...]
            td      = collar['td']

            # Ensure depth=0 station exists
            if not surveys or surveys[0][0] > 0:
                first_dip = surveys[0][1] if surveys else -90.0
                first_azi = surveys[0][2] if surveys else 0.0
                surveys.insert(0, (0.0, first_dip, first_azi))

            # Extend to TD if the deepest station is shallower
            if surveys[-1][0] < td:
                surveys.append((td, surveys[-1][1], surveys[-1][2]))

            trace_chainage = [collar_chainage]
            trace_rl       = [collar_rl]
            trace_depths   = [surveys[0][0]]   # depth (m) at each trace station
            cum_e = 0.0
            cum_n = 0.0
            cum_z = 0.0  # cumulative vertical drop (positive = downward)

            for seg_idx in range(len(surveys) - 1):
                d0, dip0, azi0 = surveys[seg_idx]
                d1, dip1, azi1 = surveys[seg_idx + 1]
                delta = d1 - d0
                if delta <= 0:
                    continue

                # Average dip and azimuth (handle 0/360 wrap for azimuth)
                avg_dip = (dip0 + dip1) / 2.0
                if abs(azi1 - azi0) > 180:
                    if azi1 > azi0:
                        azi0 += 360
                    else:
                        azi1 += 360
                avg_azi = (azi0 + azi1) / 2.0

                abs_dip_rad = math.radians(abs(avg_dip))
                azi_rad     = math.radians(avg_azi)

                horiz   = delta * math.cos(abs_dip_rad)
                d_east  = horiz * math.sin(azi_rad)
                d_north = horiz * math.cos(azi_rad)
                d_down  = delta * math.sin(abs_dip_rad)

                cum_e += d_east
                cum_n += d_north
                cum_z += d_down

                trace_chainage.append(collar_chainage + cum_e * ux + cum_n * uy)
                trace_rl.append(collar_rl - cum_z)
                trace_depths.append(d1)

            # ── plot geology colour bands on trace ────────────────────────
            if geo_trace_enabled and collar.get('fpath') and len(trace_depths) > 1:
                import numpy as np_interp
                td_arr = np_interp.array(trace_depths, dtype=float)
                ch_arr = np_interp.array(trace_chainage, dtype=float)
                rl_arr = np_interp.array(trace_rl, dtype=float)
                max_depth_trace = td_arr[-1]

                intervals = self._read_drill_geology(
                    collar['fpath'], geo_sheet, geo_from_col, geo_to_col, geo_code_col
                )
                for from_d, to_d, code in intervals:
                    if to_d <= td_arr[0] or from_d >= max_depth_trace:
                        continue
                    from_d = max(from_d, float(td_arr[0]))
                    to_d   = min(to_d,   float(max_depth_trace))
                    ch_f = float(np_interp.interp(from_d, td_arr, ch_arr))
                    rl_f = float(np_interp.interp(from_d, td_arr, rl_arr))
                    ch_t = float(np_interp.interp(to_d,   td_arr, ch_arr))
                    rl_t = float(np_interp.interp(to_d,   td_arr, rl_arr))
                    if code not in lith_colour_map:
                        lith_colour_map[code] = _LITH_PALETTE[
                            len(lith_colour_map) % len(_LITH_PALETTE)
                        ]
                    ax.plot([ch_f, ch_t], [rl_f, rl_t],
                            color=lith_colour_map[code],
                            linewidth=5,
                            solid_capstyle='butt',
                            zorder=4)

            # ── plot collar symbol ────────────────────────────────────────
            ax.plot(collar_chainage, collar_rl,
                    marker='o',
                    markersize=8,
                    color=colour,
                    markeredgecolor='black',
                    markeredgewidth=0.8,
                    linestyle='none',
                    zorder=6)

            if show_id:
                ax.annotate(
                    collar['id'],
                    xy=(collar_chainage, collar_rl),
                    xytext=(4, 4),
                    textcoords='offset points',
                    fontsize=7,
                    color=colour,
                    fontweight='bold',
                    zorder=7
                )

            # ── plot deviated trace ───────────────────────────────────────
            ax.plot(trace_chainage, trace_rl,
                    color=colour,
                    linewidth=1.5,
                    linestyle='-',
                    label='Drill Holes' if plotted == 0 else '_nolegend_',
                    zorder=5)

            if show_td and len(trace_chainage) > 1:
                ax.annotate(
                    f"{td:.0f} m",
                    xy=(trace_chainage[-1], trace_rl[-1]),
                    xytext=(3, -10),
                    textcoords='offset points',
                    fontsize=6,
                    color=colour,
                    zorder=7
                )

            plotted += 1

        QgsMessageLog.logMessage(
            f"DrillHoles: plotted {plotted} hole(s) on section '{sheet_name}' "
            f"(corridor ±{corridor} m)",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

        # ── add lithology legend entries ──────────────────────────────────
        if geo_trace_enabled and lith_colour_map:
            import matplotlib.patches as mpatches
            for code, col in sorted(lith_colour_map.items()):
                ax.plot([], [],
                        color=col,
                        linewidth=5,
                        label=code,
                        solid_capstyle='butt')


    def _plot_pegmatite_bands(self, ax, sheet_name: str):
        """
        Draw hatched bands on the cross-section for each interpreted pegmatite
        polygon that intersects the section line.

        Only features with a non-null/non-empty 'label' field are included when
        the 'Only labeled features' checkbox is ticked.

        Bands are drawn as semi-transparent hatched vertical spans covering the
        full axis height, using a distinct visual style (purple hatch) to
        differentiate from geology bands.
        """
        import math

        if not getattr(self, 'peg_band_chk', None) or not self.peg_band_chk.isChecked():
            return

        layer_id = (self.peg_band_layer_combo.currentData()
                    if getattr(self, 'peg_band_layer_combo', None) else None)
        if not layer_id:
            QgsMessageLog.logMessage(
                f"PegmatiteBands: No polygon layer selected for sheet '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        peg_layer = QgsProject.instance().mapLayer(layer_id)
        if not peg_layer:
            return

        labeled_only = (getattr(self, 'peg_labeled_only_chk', None) and
                        self.peg_labeled_only_chk.isChecked())
        show_labels = (getattr(self, 'peg_band_label_chk', None) and
                       self.peg_band_label_chk.isChecked())

        # ── section line (reuse same combos as geology bands) ────────────
        geo_sec_layer_id = (self.geo_section_line_combo.currentData()
                            if getattr(self, 'geo_section_line_combo', None) else None)
        dh_sec_layer_id  = (self.dh_layer_combo.currentData()
                            if getattr(self, 'dh_layer_combo', None) else None)
        active_sec_layer_id = geo_sec_layer_id or dh_sec_layer_id

        if not active_sec_layer_id:
            QgsMessageLog.logMessage(
                f"PegmatiteBands: No section line layer selected for sheet '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        section = self._get_section_line_from_layer(sheet_name, active_sec_layer_id)
        if section is None:
            QgsMessageLog.logMessage(
                f"PegmatiteBands: No section line feature found for sheet '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return
        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        sec_lyr = QgsProject.instance().mapLayer(active_sec_layer_id)
        section_crs = sec_lyr.crs() if sec_lyr else QgsProject.instance().crs()
        section_geom = QgsGeometry.fromPolylineXY([QgsPointXY(E0, N0), QgsPointXY(E1, N1)])

        # CRS transform: pegmatite layer → section CRS
        peg_crs = peg_layer.crs()
        needs_xform = (peg_crs.isValid() and section_crs.isValid() and
                       peg_crs.authid() != section_crs.authid())
        peg_xform = (QgsCoordinateTransform(peg_crs, section_crs, QgsProject.instance())
                     if needs_xform else None)

        # ── find 'label' field (case-insensitive) ──────────────────────
        label_field = None
        for f in peg_layer.fields():
            if f.name().lower() == 'label':
                label_field = f.name()
                break

        # ── intersect each polygon with the section line ───────────────
        bands = []   # [(d_start, d_end, label_text)]

        for feat in peg_layer.getFeatures():
            # Filter unlabeled features if requested
            lbl = ''
            if label_field:
                raw = feat.attribute(label_field)
                lbl = '' if raw is None else str(raw).strip()
                if lbl.upper() in ('NULL', 'NONE', 'N/A'):
                    lbl = ''

            if labeled_only and not lbl:
                continue

            geom = feat.geometry()
            if geom.isEmpty():
                continue
            if peg_xform:
                geom.transform(peg_xform)

            inter = section_geom.intersection(geom)
            if inter is None or inter.isEmpty():
                continue

            dists = []
            for v in inter.vertices():
                vec_e = v.x() - E0
                vec_n = v.y() - N0
                dists.append(vec_e * ux + vec_n * uy)

            if len(dists) < 2:
                continue

            d_start = min(dists)
            d_end   = max(dists)
            if d_end <= d_start:
                continue

            bands.append((d_start, d_end, lbl))

        if not bands:
            QgsMessageLog.logMessage(
                f"PegmatiteBands: no polygons intersect section '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            return

        # ── draw hatched vertical bands spanning full axis height ────────
        # Purple palette — visually distinct from geology bands
        _PEG_COLOURS = [
            '#9b59b6', '#8e44ad', '#a569bd', '#76448a',
            '#bb8fce', '#6c3483', '#d2b4de', '#5b2c6f',
        ]

        drawn = set()
        for idx, (d_start, d_end, lbl) in enumerate(sorted(bands, key=lambda x: x[0])):
            colour = _PEG_COLOURS[idx % len(_PEG_COLOURS)]
            legend_lbl = f'Pegmatite: {lbl}' if lbl and lbl not in drawn else '_nolegend_'
            drawn.add(lbl)

            ax.axvspan(
                d_start, d_end,
                ymin=0.0, ymax=1.0,
                facecolor=colour,
                alpha=0.18,
                edgecolor=colour,
                linewidth=1.2,
                linestyle='-',
                hatch='///',
                zorder=3,
                label=legend_lbl
            )

            if show_labels and lbl:
                x_mid = (d_start + d_end) / 2.0
                import matplotlib.transforms as _mtrans
                trans = _mtrans.blended_transform_factory(ax.transData, ax.transAxes)
                rotation = 90 if (d_end - d_start) < 120 else 0
                ax.text(
                    x_mid, 0.5, lbl,
                    transform=trans,
                    ha='center', va='center',
                    fontsize=7, fontweight='bold',
                    color=colour,
                    rotation=rotation,
                    zorder=6,
                    clip_on=True
                )

        QgsMessageLog.logMessage(
            f"PegmatiteBands: drew {len(bands)} pegmatite band(s) for section '{sheet_name}'.",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

    def _plot_geology_bands(self, ax, sheet_name: str, df=None, dist_col=None, elev_col=None):
        """
        Draw colored geology bands draped over the surface profile.

        For each polygon in the selected geology layer that intersects the section
        line, a filled colour band is drawn from the terrain surface upward by a
        fixed height (band_pct × Y-range metres).  Labels are placed inside each
        band at the midpoint of the segment.

        Parameters
        ----------
        df       : pandas DataFrame containing the surface profile (optional).
        dist_col : column name for distance values.
        elev_col : column name for elevation values.

        All geometry is processed in the section line's CRS; the geology layer is
        transformed automatically if it is in a different CRS.
        """
        import math

        if not getattr(self, 'geo_band_chk', None) or not self.geo_band_chk.isChecked():
            return

        layer_id = (self.geo_band_layer_combo.currentData()
                    if getattr(self, 'geo_band_layer_combo', None) else None)
        if not layer_id:
            return

        geology_layer = QgsProject.instance().mapLayer(layer_id)
        if not geology_layer:
            return

        field_name = (self.geo_band_field_combo.currentText()
                      if getattr(self, 'geo_band_field_combo', None) else '')
        band_pct = (self.geo_band_pct_spin.value()
                    if getattr(self, 'geo_band_pct_spin', None) else 8) / 100.0
        show_labels = (getattr(self, 'geo_band_label_chk', None) and
                       self.geo_band_label_chk.isChecked())

        # ── resolve section line layer ────────────────────────────────────
        # Prefer the geology-bands-specific section line combo; fall back to
        # the drill-hole section line combo if no dedicated one is selected.
        geo_sec_layer_id = (self.geo_section_line_combo.currentData()
                            if getattr(self, 'geo_section_line_combo', None) else None)
        dh_sec_layer_id  = (self.dh_layer_combo.currentData()
                            if getattr(self, 'dh_layer_combo', None) else None)
        active_sec_layer_id = geo_sec_layer_id or dh_sec_layer_id

        if not active_sec_layer_id:
            QgsMessageLog.logMessage(
                f"GeologyBands: No section line layer selected for sheet '{sheet_name}'. "
                "Please choose a section line layer in the Geology Bands group "
                "(or in the Drill Hole Overlay group as a fallback).",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        # ── section line in section-layer CRS ────────────────────────────
        section = self._get_section_line_from_layer(sheet_name, active_sec_layer_id)
        if section is None:
            QgsMessageLog.logMessage(
                f"GeologyBands: No section line feature matching sheet '{sheet_name}' "
                f"was found in the selected section line layer. "
                "Ensure a feature's name/text field matches the sheet name exactly.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return

        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        # Determine section layer CRS from the active section line layer
        sec_lyr = QgsProject.instance().mapLayer(active_sec_layer_id)
        section_crs = sec_lyr.crs() if sec_lyr else QgsProject.instance().crs()

        section_geom = QgsGeometry.fromPolylineXY([QgsPointXY(E0, N0), QgsPointXY(E1, N1)])

        # CRS transform: geology layer → section CRS
        geo_crs = geology_layer.crs()
        needs_xform = (geo_crs.isValid() and section_crs.isValid() and
                       geo_crs.authid() != section_crs.authid())
        geo_xform = (QgsCoordinateTransform(geo_crs, section_crs, QgsProject.instance())
                     if needs_xform else None)

        # ── intersect each polygon with the section line ──────────────────
        bands = []   # [(d_start, d_end, label)]

        for feat in geology_layer.getFeatures():
            geom = feat.geometry()
            if geom.isEmpty():
                continue
            if geo_xform:
                geom.transform(geo_xform)

            inter = section_geom.intersection(geom)
            if inter is None or inter.isEmpty():
                continue

            # Collect distances of all vertices in the intersection geometry
            dists = []
            for v in inter.vertices():
                vec_e = v.x() - E0
                vec_n = v.y() - N0
                dists.append(vec_e * ux + vec_n * uy)

            if len(dists) < 2:
                continue

            d_start = min(dists)
            d_end   = max(dists)
            if d_end <= d_start:
                continue

            label = ''
            if field_name:
                try:
                    raw = feat.attribute(field_name)
                    label = '' if raw is None else str(raw).strip()
                    if label.upper() in ('NULL', 'NONE', 'N/A'):
                        label = ''
                except Exception:
                    label = ''

            bands.append((d_start, d_end, label))

        if not bands:
            QgsMessageLog.logMessage(
                f"GeologyBands: no polygons intersect section '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            return

        # ── assign a consistent color to each unique label ────────────────
        # Colors are derived from ALL features in the geology layer (not just
        # those that intersect this section) so that each label gets the same
        # color on every cross-section plot.
        try:
            import matplotlib.cm as _mcm
            all_labels_in_layer = sorted(set(
                str(feat.attribute(field_name)).strip()
                for feat in geology_layer.getFeatures()
                if feat.attribute(field_name) is not None
                and str(feat.attribute(field_name)).strip() not in ('', 'NULL', 'NONE', 'N/A')
            )) if field_name else []
            # Merge with any labels found in this section (defensive)
            section_labels = sorted(set(b[2] for b in bands if b[2]))
            unique_labels = sorted(set(all_labels_in_layer) | set(section_labels))
            n = max(len(unique_labels), 1)
            _cmap = _mcm.get_cmap('tab20', n)
            label_color = {lbl: _cmap(i) for i, lbl in enumerate(unique_labels)}
        except Exception:
            _palette = ['#a6cee3','#1f78b4','#b2df8a','#33a02c','#fb9a99',
                        '#e31a1c','#fdbf6f','#ff7f00','#cab2d6','#6a3d9a',
                        '#ffff99','#b15928']
            all_labels_in_layer = sorted(set(
                str(feat.attribute(field_name)).strip()
                for feat in geology_layer.getFeatures()
                if feat.attribute(field_name) is not None
                and str(feat.attribute(field_name)).strip() not in ('', 'NULL', 'NONE', 'N/A')
            )) if field_name else []
            section_labels = sorted(set(b[2] for b in bands if b[2]))
            unique_labels = sorted(set(all_labels_in_layer) | set(section_labels))
            label_color = {lbl: _palette[i % len(_palette)]
                           for i, lbl in enumerate(unique_labels)}

        # ── build surface profile arrays for terrain-following bands ──────
        import numpy as _np
        has_profile = (df is not None and dist_col is not None and
                       elev_col is not None and
                       dist_col in df.columns and elev_col in df.columns)
        if has_profile:
            _valid = df[dist_col].notna() & df[elev_col].notna()
            _d_arr = _np.array(df[dist_col][_valid], dtype=float)
            _e_arr = _np.array(df[elev_col][_valid], dtype=float)
            # Sort by distance
            _sort_idx = _np.argsort(_d_arr)
            _d_arr = _d_arr[_sort_idx]
            _e_arr = _e_arr[_sort_idx]
        else:
            _d_arr = _np.array([])
            _e_arr = _np.array([])

        # Band height in data units = band_pct × current Y range
        y_lo, y_hi = ax.get_ylim()
        band_height = band_pct * (y_hi - y_lo)

        # ── build renderer-label lookup (short code → full description) ────
        # When the geology layer uses a categorized renderer (as in the QGIS
        # legend panel, e.g. "Aba – Amphibolite; fine- to medium-grained…"),
        # extract each category's display label so the matplotlib legend shows
        # the full description rather than just the short field value.
        renderer_labels = {}   # {field_value_str: renderer_category_label}
        try:
            _renderer = geology_layer.renderer()
            if _renderer and _renderer.type() == 'categorizedSymbol':
                for _cat in _renderer.categories():
                    _val = str(_cat.value()).strip()
                    _lbl = str(_cat.label()).strip()
                    if _val and _lbl and _val != _lbl:
                        renderer_labels[_val] = _lbl
        except Exception:
            pass

        # ── draw terrain-following bands ──────────────────────────────────
        drawn_labels = set()
        for d_start, d_end, label in sorted(bands, key=lambda x: x[0]):
            color = label_color.get(label, '#cccccc')
            if label not in drawn_labels:
                legend_label = renderer_labels.get(label, label)
            else:
                legend_label = '_nolegend_'
            drawn_labels.add(label)

            if has_profile and len(_d_arr) >= 2:
                # Extract surface points inside [d_start, d_end]
                mask = (_d_arr >= d_start) & (_d_arr <= d_end)
                x_pts = list(_d_arr[mask])
                y_pts = list(_e_arr[mask])

                # Interpolate surface elevation at the exact band edges
                e_at_start = float(_np.interp(d_start, _d_arr, _e_arr))
                e_at_end   = float(_np.interp(d_end,   _d_arr, _e_arr))

                # Prepend/append boundary points
                if not x_pts or x_pts[0] > d_start:
                    x_pts.insert(0, d_start)
                    y_pts.insert(0, e_at_start)
                if not x_pts or x_pts[-1] < d_end:
                    x_pts.append(d_end)
                    y_pts.append(e_at_end)

                x_pts = _np.array(x_pts)
                y_surf = _np.array(y_pts)
                y_top  = y_surf + band_height

                ax.fill_between(x_pts, y_surf, y_top,
                                alpha=0.75,
                                facecolor=color,
                                edgecolor='white',
                                linewidth=0.5,
                                zorder=4,
                                label=legend_label if label else '_nolegend_')

                if show_labels and label:
                    x_mid = (d_start + d_end) / 2.0
                    y_surf_mid = float(_np.interp(x_mid, x_pts, y_surf))
                    y_label    = y_surf_mid + band_height / 2.0
                    rotation = 90 if (d_end - d_start) < 120 else 0
                    ax.text(x_mid, y_label, label,
                            ha='center', va='center',
                            fontsize=7, fontweight='bold',
                            color='black',
                            rotation=rotation,
                            zorder=6,
                            clip_on=True)
            else:
                # Fallback: flat band at top using axes coordinates
                import matplotlib.transforms as _mtrans
                trans = _mtrans.blended_transform_factory(ax.transData, ax.transAxes)
                y_bot_ax = 1.0 - band_pct
                ax.axvspan(d_start, d_end,
                           ymin=y_bot_ax, ymax=1.0,
                           alpha=0.75,
                           facecolor=color,
                           edgecolor='white',
                           linewidth=0.5,
                           zorder=4,
                           label=legend_label if label else '_nolegend_')
                if show_labels and label:
                    x_mid = (d_start + d_end) / 2.0
                    y_mid_ax = y_bot_ax + band_pct / 2.0
                    ax.text(x_mid, y_mid_ax, label,
                            transform=trans,
                            ha='center', va='center',
                            fontsize=7, fontweight='bold',
                            color='black',
                            rotation=90 if (d_end - d_start) < 120 else 0,
                            zorder=4,
                            clip_on=True)

        # Vertical dashed boundary lines at each geology transition
        all_bounds = sorted(set(d for d_start, d_end, _ in bands
                                for d in (d_start, d_end)))
        for d in all_bounds:
            ax.axvline(x=d, color='#555555', linestyle='--',
                       linewidth=0.6, alpha=0.5, zorder=5)

        QgsMessageLog.logMessage(
            f"GeologyBands: drew {len(bands)} band(s) for section '{sheet_name}'.",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

    # =========================================================================

    def _plot_plate_locations(self, ax, sheet_name: str):
        """
        Project photo plate locations onto the cross-section and annotate them
        at the top of the plot with a downward-pointing triangle and label.

        Markers are placed in axes coordinates (y=1.0 = top edge) so they
        never obscure data inside the plot. Labels are rotated 90° for
        compactness when many plates fall on the same section.
        """
        import math, os
        import matplotlib.transforms as _mtrans

        if not getattr(self, 'plate_enable_chk', None) or not self.plate_enable_chk.isChecked():
            return

        layer_id = (self.plate_layer_combo.currentData()
                    if getattr(self, 'plate_layer_combo', None) else None)
        if not layer_id:
            return
        plate_layer = QgsProject.instance().mapLayer(layer_id)
        if not plate_layer:
            return

        show_label = (getattr(self, 'plate_label_chk', None) and
                      self.plate_label_chk.isChecked())
        corridor = (self.plate_corridor_spin.value()
                    if getattr(self, 'plate_corridor_spin', None) else 500.0)

        # ── section line ────────────────────────────────────────────────
        geo_sec_layer_id = (self.geo_section_line_combo.currentData()
                            if getattr(self, 'geo_section_line_combo', None) else None)
        dh_sec_layer_id  = (self.dh_layer_combo.currentData()
                            if getattr(self, 'dh_layer_combo', None) else None)
        active_sec_layer_id = geo_sec_layer_id or dh_sec_layer_id
        if not active_sec_layer_id:
            QgsMessageLog.logMessage(
                f"PlateLoc: No section line layer selected for sheet '{sheet_name}'.",
                "IntegratedProfileAnalyzer", Qgis.Warning)
            return

        section = self._get_section_line_from_layer(sheet_name, active_sec_layer_id)
        if section is None:
            return

        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return
        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        sec_lyr     = QgsProject.instance().mapLayer(active_sec_layer_id)
        section_crs = sec_lyr.crs() if sec_lyr else QgsProject.instance().crs()

        # ── CRS transform: plate layer → section CRS ────────────────────
        plate_crs   = plate_layer.crs()
        needs_xform = (plate_crs.isValid() and section_crs.isValid() and
                       plate_crs.authid() != section_crs.authid())
        xform = (QgsCoordinateTransform(plate_crs, section_crs, QgsProject.instance())
                 if needs_xform else None)

        # ── find relevant attribute fields (case-insensitive) ────────────
        field_map = {f.name().lower(): f.name() for f in plate_layer.fields()}
        fn_field  = field_map.get('filename') or field_map.get('name') or field_map.get('label')

        # ── blended transform: data-x, axes-y (so markers sit at top edge) ─
        trans_top = _mtrans.blended_transform_factory(ax.transData, ax.transAxes)

        colour  = '#E65100'   # deep orange — visible against all backgrounds
        plotted = 0

        for feat in plate_layer.getFeatures():
            geom = feat.geometry()
            if geom.isEmpty():
                continue

            pt = geom.asPoint()
            east_p, north_p = pt.x(), pt.y()

            if xform is not None:
                try:
                    tp = xform.transform(QgsPointXY(east_p, north_p))
                    east_p, north_p = tp.x(), tp.y()
                except Exception:
                    continue

            vec_e    = east_p - E0
            vec_n    = north_p - N0
            chainage = vec_e * ux + vec_n * uy
            perp     = vec_e * (-uy) + vec_n * ux

            if abs(perp) > corridor:
                continue

            # ── plate name for label ──────────────────────────────────
            if fn_field:
                raw = feat.attribute(fn_field)
                label_text = os.path.splitext(str(raw).strip())[0] if raw else ''
            else:
                label_text = ''

            # ── downward triangle just inside the top of the axes ────────
            # y=0.97 keeps the marker visible within the plot frame.
            # The label hangs downward from the marker (va='top') so it
            # reads into the plot area and is always fully visible.
            legend_lbl = 'Plate location' if plotted == 0 else '_nolegend_'
            ax.plot(chainage, 0.97,
                    transform=trans_top,
                    marker='v',
                    markersize=8,
                    color=colour,
                    markeredgecolor='white',
                    markeredgewidth=0.8,
                    linestyle='none',
                    zorder=10,
                    label=legend_lbl,
                    clip_on=False)

            # ── label hangs downward from the marker ──────────────────────
            if show_label and label_text:
                ax.text(chainage, 0.94,
                        label_text,
                        transform=trans_top,
                        ha='center', va='top',
                        fontsize=6,
                        color=colour,
                        fontweight='bold',
                        rotation=90,
                        zorder=10,
                        clip_on=True)

            plotted += 1

        QgsMessageLog.logMessage(
            f"PlateLoc '{sheet_name}': plotted {plotted} plate location(s).",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

    def _plot_structural_measurements(self, ax, sheet_name: str):
        """
        Project structural measurement points onto the section and draw dip ticks.

        For each point within the corridor:
          1. Project onto the section to get chainage.
          2. Sample elevation from the DEM raster.
          3. Draw a tick line of length tick_len_m at the true dip angle,
             extending from the surface downward in the section plane.
          4. Optionally label with the dip value.
        """
        import math

        if not getattr(self, 'struct_enable_chk', None) or not self.struct_enable_chk.isChecked():
            return

        layer_id = (self.struct_layer_combo.currentData()
                    if getattr(self, 'struct_layer_combo', None) else None)
        if not layer_id:
            return
        struct_layer = QgsProject.instance().mapLayer(layer_id)
        if not struct_layer:
            return

        dip_field = (self.struct_dip_combo.currentText()
                     if getattr(self, 'struct_dip_combo', None) else '')
        dir_field = (self.struct_dir_combo.currentText()
                     if getattr(self, 'struct_dir_combo', None) else '')
        tick_len  = (self.struct_tick_spin.value()
                     if getattr(self, 'struct_tick_spin', None) else 5.0)
        show_label = (getattr(self, 'struct_label_chk', None) and
                      self.struct_label_chk.isChecked())
        corridor  = (self.struct_corridor_spin.value()
                     if getattr(self, 'struct_corridor_spin', None) else 500.0)

        # ── section geometry ─────────────────────────────────────────────
        # Prefer the geology-bands section line combo; fall back to drill-hole combo.
        geo_sec_layer_id = (self.geo_section_line_combo.currentData()
                            if getattr(self, 'geo_section_line_combo', None) else None)
        dh_sec_layer_id  = (self.dh_layer_combo.currentData()
                            if getattr(self, 'dh_layer_combo', None) else None)
        active_sec_layer_id = geo_sec_layer_id or dh_sec_layer_id

        if not active_sec_layer_id:
            QgsMessageLog.logMessage(
                f"StructDip: No section line layer selected for sheet '{sheet_name}'. "
                "Select a layer in the Geology Bands or Drill Hole Overlay group.",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return

        section = self._get_section_line_from_layer(sheet_name, active_sec_layer_id)
        if section is None:
            return
        (E0, N0), (E1, N1) = section
        sec_dE, sec_dN = E1 - E0, N1 - N0
        sec_len = math.hypot(sec_dE, sec_dN)
        if sec_len == 0:
            return
        ux = sec_dE / sec_len
        uy = sec_dN / sec_len

        # ── CRS transform: struct layer → section layer CRS ──────────────
        if active_sec_layer_id:
            sec_lyr     = QgsProject.instance().mapLayer(active_sec_layer_id)
            section_crs = sec_lyr.crs() if sec_lyr else QgsProject.instance().crs()
        else:
            section_crs = QgsProject.instance().crs()

        struct_crs  = struct_layer.crs()
        needs_xform = (struct_crs.isValid() and section_crs.isValid() and
                       struct_crs.authid() != section_crs.authid())
        xform = (QgsCoordinateTransform(struct_crs, section_crs, QgsProject.instance())
                 if needs_xform else None)

        # ── DEM raster for surface elevation ─────────────────────────────
        dem_raster = self._resolve_dem_raster()

        colour   = '#7B1FA2'   # purple
        plotted  = 0
        skipped_corridor = 0
        skipped_elev     = 0
        skipped_dip      = 0

        for feat in struct_layer.getFeatures():
            geom = feat.geometry()
            if geom.isEmpty():
                continue

            pt = geom.asPoint()
            east_orig, north_orig = pt.x(), pt.y()   # in struct layer CRS

            # Transform to section CRS for corridor / chainage check
            if xform is not None:
                try:
                    tp = xform.transform(QgsPointXY(east_orig, north_orig))
                    east_s, north_s = tp.x(), tp.y()
                except Exception:
                    skipped_corridor += 1
                    continue
            else:
                east_s, north_s = east_orig, north_orig

            vec_e = east_s - E0
            vec_n = north_s - N0
            chainage = vec_e * ux + vec_n * uy
            perp     = vec_e * (-uy) + vec_n * ux

            if abs(perp) > corridor:
                skipped_corridor += 1
                QgsMessageLog.logMessage(
                    f"StructDip SKIP (corridor): perp={abs(perp):.1f}m > {corridor}m  "
                    f"chainage={chainage:.1f}m  feat_id={feat.id()}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                continue

            # Sample elevation from DEM using section CRS coords
            rl = None
            if dem_raster is not None and dem_raster.isValid():
                from qgis.core import QgsRaster
                raster_crs = dem_raster.crs()
                if section_crs.authid() != raster_crs.authid():
                    r_xform = QgsCoordinateTransform(
                        section_crs, raster_crs, QgsProject.instance())
                    try:
                        rp = r_xform.transform(QgsPointXY(east_s, north_s))
                    except Exception:
                        rp = None
                else:
                    rp = QgsPointXY(east_s, north_s)

                if rp is not None:
                    result = dem_raster.dataProvider().identify(
                        rp, QgsRaster.IdentifyFormatValue)
                    if result.isValid():
                        val = result.results().get(1)
                        if val is not None:
                            try:
                                rl = float(val)
                            except (TypeError, ValueError):
                                pass

            # Fallback: try _sample_raster_rl with original coords
            if rl is None:
                try:
                    rl = self._sample_raster_rl(east_orig, north_orig,
                                                raster=dem_raster)
                except Exception:
                    pass

            if rl is None:
                skipped_elev += 1
                QgsMessageLog.logMessage(
                    f"StructDip SKIP (no elevation): feat_id={feat.id()}  "
                    f"chainage={chainage:.1f}m  perp={perp:.1f}m  "
                    f"dem={'set' if dem_raster else 'None'}",
                    "IntegratedProfileAnalyzer", Qgis.Warning
                )
                continue

            # ── get dip value ────────────────────────────────────────────
            try:
                dip = float(feat.attribute(dip_field))
            except (TypeError, ValueError, KeyError):
                skipped_dip += 1
                continue

            # ── draw tick: from surface point downward at true dip ───────
            # Horizontal projection along section = tick_len * cos(dip)
            # Vertical drop                        = tick_len * sin(dip)
            dip_rad = math.radians(abs(dip))

            # Dip direction relative to section — determines left/right lean
            try:
                dip_dir = float(feat.attribute(dir_field))
            except (TypeError, ValueError, KeyError):
                dip_dir = None

            if dip_dir is not None:
                # Section azimuth (angle from North of section direction)
                sec_az = math.degrees(math.atan2(sec_dE, sec_dN)) % 360
                # Component of dip direction along section (+ve = forward)
                along = math.cos(math.radians(dip_dir - sec_az))
                h_sign = 1.0 if along >= 0 else -1.0
            else:
                h_sign = 1.0

            dx = h_sign * tick_len * math.cos(dip_rad)
            dz = tick_len * math.sin(dip_rad)          # always drops down

            x0, y0 = chainage, rl
            x1, y1 = chainage + dx, rl - dz

            label_str = 'Structural dip' if plotted == 0 else '_nolegend_'
            ax.plot([x0, x1], [y0, y1],
                    color=colour, linewidth=2.5,
                    solid_capstyle='butt',
                    label=label_str, zorder=7)

            if show_label:
                ax.annotate(
                    f"{abs(dip):.0f}°",
                    xy=(x0, y0),
                    xytext=(3, 4),
                    textcoords='offset points',
                    fontsize=7,
                    color=colour,
                    fontweight='bold',
                    zorder=9
                )
            plotted += 1

        QgsMessageLog.logMessage(
            f"StructDip '{sheet_name}': plotted={plotted}  "
            f"skipped_corridor={skipped_corridor}  skipped_elev={skipped_elev}  "
            f"skipped_dip={skipped_dip}  corridor=±{corridor}m",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

    # =========================================================================

    def _plot_intersection_points(self, ax, intersection_df, sheet_name: str):
        """Enhanced intersection point plotting with better styling and logging"""
        if intersection_df is None:
            return

        # Respect Plot Options marker size (0 = hidden)
        _mk_sz = (self.plot_marker_size_spin.value()
                  if getattr(self, 'plot_marker_size_spin', None) else 6)
        if _mk_sz == 0:
            return
        _scatter_s = _mk_sz ** 2

        if PANDAS_AVAILABLE and hasattr(intersection_df, 'empty') and not intersection_df.empty:
            # Handle pandas DataFrame
            ax.scatter(intersection_df['distance'], intersection_df['elevation'],
                      color='#FF0000',
                      marker='^',
                      s=_scatter_s,
                      label='Intersection Points',
                      zorder=5,
                      edgecolors='black',
                      linewidth=0.5,
                      alpha=0.9)
                      
            QgsMessageLog.logMessage(
                f"Plotted {len(intersection_df)} intersection points for {sheet_name}", 
                "IntegratedProfileAnalyzer", Qgis.Info
            )
        elif isinstance(intersection_df, list) and len(intersection_df) > 0:
            # Handle list of dicts (when pandas not available)
            distances = [pt['distance'] for pt in intersection_df if 'distance' in pt and 'elevation' in pt]
            elevations = [pt['elevation'] for pt in intersection_df if 'distance' in pt and 'elevation' in pt]
            
            if distances and elevations:
                ax.scatter(distances, elevations,
                          color='#FF0000',
                          marker='^',
                          s=_scatter_s,
                          label='Intersection Points',
                          zorder=5,
                          edgecolors='black',
                          linewidth=0.5,
                          alpha=0.9)
                          
                QgsMessageLog.logMessage(
                    f"Plotted {len(distances)} intersection points for {sheet_name}", 
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
            
    def plot_single_profile(self, df: pd.DataFrame, sheet_name: str, params: Dict[str, Any], intersection_df: Optional[Any] = None) -> Optional[plt.Figure]:
        """Plot a single profile with all available data columns using dynamic column detection"""
        try:
            # Detect all available distance/elevation column pairs dynamically
            column_pairs = self._detect_profile_columns(df)
            
            if not column_pairs:
                raise ValueError(f"No valid distance/elevation column pairs found in sheet '{sheet_name}'. Available columns: {list(df.columns)}")
            
            QgsMessageLog.logMessage(
                f"Detected {len(column_pairs)} profile pairs for sheet '{sheet_name}': {column_pairs}",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            # Calculate overall data bounds for all columns
            all_distances = []
            all_elevations = []
            
            for dist_col, elev_col in column_pairs:
                valid_mask = df[dist_col].notna() & df[elev_col].notna()
                if valid_mask.any():
                    all_distances.extend(df[dist_col][valid_mask].tolist())
                    all_elevations.extend(df[elev_col][valid_mask].tolist())
            
            if not all_distances or not all_elevations:
                raise ValueError(f"No valid data points found in any column pairs for sheet '{sheet_name}'")
            
            # Enforce fixed figure size for all plots (1:1 ratio for the plot
            # area). Extra height is added below to hold the legend.
            fixed_size = 8  # inches — plot area width and height
            legend_height = 2.5  # inches reserved below the plot for the legend
            fig = plt.figure(figsize=(fixed_size, fixed_size + legend_height))
            # Place the axes in the top portion; leave the bottom for the legend
            ax = fig.add_axes([0.10, legend_height / (fixed_size + legend_height),
                               0.85, fixed_size / (fixed_size + legend_height) * 0.90])
            # 1:1 data aspect ratio — 1 metre horizontal == 1 metre vertical.
            # This is set here unconditionally and again after axis limits are
            # finalised later in this function.
            ax.set_aspect('equal', adjustable='datalim')
            
            # Set plot bounds with padding
            x_min, x_max = min(all_distances), max(all_distances)
            y_min, y_max = min(all_elevations), max(all_elevations)
            
            x_padding = (x_max - x_min) * 0.1 if x_max > x_min else 1
            y_padding = (y_max - y_min) * 0.1 if y_max > y_min else 1
            
            ax.set_xlim(x_min - x_padding, x_max + x_padding)
            ax.set_ylim(y_min - y_padding, y_max + y_padding)
            
            # Set font sizes
            plt.rcParams['font.size'] = int(self.plot_config.get('text_font_size', '10'))
            
            # Set background
            ax.set_facecolor('white')
            fig.patch.set_facecolor('white')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            
            # Define dynamic styling for multiple profiles
            profile_colors = [
                self.plot_config.get('main_line_color', '#000000'),    # Black for main
                self.plot_config.get('feature_line_color', '#0000FF'), # Blue for feature/secondary
                '#008000',  # Green for additional profiles
                '#FFA500',  # Orange for additional profiles
                '#9C27B0',  # Purple
                '#E91E63',  # Pink
                '#FF5722',  # Deep Orange
                '#795548',  # Brown
                '#607D8B',  # Blue Grey
                '#FF9800',  # Orange
            ]
            
            profile_line_styles = ['-', '-', '-', '-', '--', '-.', ':', '-', '--', '-.']
            # Get configurable line widths
            profile_line_widths = [
                float(self.plot_config.get('main_line_thickness', '2.0')),     # Main profile
                float(self.plot_config.get('feature_line_thickness', '1.5')),  # Feature line
                float(self.plot_config.get('profile_3_thickness', '1.5')),     # Profile 3
                float(self.plot_config.get('feature_4_thickness', '1.5')),     # Feature 4
                1.5, 1.5, 1.5, 1.5, 1.5, 1.5  # Additional profiles default thickness
            ]
            
            # Plot all profile pairs with dynamic styling
            profiles_plotted = 0
            for idx, (dist_col, elev_col) in enumerate(column_pairs):
                QgsMessageLog.logMessage(
                    f"Processing profile pair {idx+1}/{len(column_pairs)}: {dist_col} / {elev_col}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                
                # Check if columns exist in dataframe
                if dist_col not in df.columns or elev_col not in df.columns:
                    QgsMessageLog.logMessage(
                        f"ERROR: Column {dist_col} or {elev_col} not found in dataframe",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                    continue
                
                valid_mask = df[dist_col].notna() & df[elev_col].notna()
                valid_count = valid_mask.sum()
                
                QgsMessageLog.logMessage(
                    f"Column {dist_col}: {df[dist_col].notna().sum()} non-null values",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                QgsMessageLog.logMessage(
                    f"Column {elev_col}: {df[elev_col].notna().sum()} non-null values",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                QgsMessageLog.logMessage(
                    f"Combined valid pairs: {valid_count}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                
                if not valid_mask.any():
                    QgsMessageLog.logMessage(
                        f"No valid data for profile pair {dist_col}/{elev_col}",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                    continue
                
                # Get data for this profile
                x_data = df[dist_col][valid_mask]
                y_data = df[elev_col][valid_mask]
                
                QgsMessageLog.logMessage(
                    f"X data range: {x_data.min():.2f} to {x_data.max():.2f}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                QgsMessageLog.logMessage(
                    f"Y data range: {y_data.min():.2f} to {y_data.max():.2f}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                
                # Determine styling
                color = profile_colors[idx % len(profile_colors)]
                line_style = profile_line_styles[idx % len(profile_line_styles)]
                line_width = profile_line_widths[idx % len(profile_line_widths)]
                
                # Create legend label based on column name
                legend_label = self._create_legend_label(dist_col, elev_col, idx)
                
                QgsMessageLog.logMessage(
                    f"Plotting profile '{legend_label}' with color {color}, style {line_style}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                
                # Plot the profile
                ax.plot(x_data, y_data,
                       color=color,
                       linestyle=line_style,
                       linewidth=line_width,
                       label=legend_label,
                       alpha=0.9,
                       zorder=3)
                
                profiles_plotted += 1
                
                QgsMessageLog.logMessage(
                    f"✓ Successfully plotted profile '{legend_label}' with {valid_count} points from {dist_col}/{elev_col}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
            
            QgsMessageLog.logMessage(
                f"Total profiles plotted: {profiles_plotted} out of {len(column_pairs)} detected pairs",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            # Plot marker columns if they exist
            self._plot_marker_columns(ax, df)

            # Plot intersection points if available
            self._plot_intersection_points(ax, intersection_df, sheet_name)

            # Plot geology bands from polygon layer if enabled
            if column_pairs:
                _gc_dist, _gc_elev = column_pairs[0]
                self._plot_geology_bands(ax, sheet_name, df=df,
                                         dist_col=_gc_dist, elev_col=_gc_elev)
            else:
                self._plot_geology_bands(ax, sheet_name)

            # Plot interpreted pegmatite bands if enabled
            self._plot_pegmatite_bands(ax, sheet_name)

            # Plot structural dip ticks if enabled
            self._plot_structural_measurements(ax, sheet_name)

            # Plot drill hole collars and traces if enabled
            self._plot_drillholes(ax, sheet_name)

            # Plot planned (proposed) drill holes if enabled
            self._plot_planned_drillholes(ax, sheet_name)

            # Plot plate locations at top of plot if enabled
            self._plot_plate_locations(ax, sheet_name)

            # Calculate and plot slopes using the main profile (first column pair)
            if column_pairs:
                self._plot_slopes(ax, df, params, column_pairs)

            # ── expand axis limits to include ALL plotted data (drill traces
            # may extend well below the surface profile elevation) ──────────
            import numpy as _np
            _ax_x, _ax_y = [], []
            for _ln in ax.get_lines():
                _xd = _ln.get_xdata()
                _yd = _ln.get_ydata()
                if len(_xd):
                    _ax_x.extend(_xd)
                    _ax_y.extend(_yd)
            for _coll in ax.collections:
                try:
                    _offs = _coll.get_offsets()
                    if len(_offs):
                        _ax_x.extend(_offs[:, 0])
                        _ax_y.extend(_offs[:, 1])
                except Exception:
                    pass
            if _ax_x and _ax_y:
                _xlo = float(_np.nanmin(_ax_x))
                _xhi = float(_np.nanmax(_ax_x))
                _ylo = float(_np.nanmin(_ax_y))
                _yhi = float(_np.nanmax(_ax_y))
                _xpad = max((_xhi - _xlo) * 0.05, 1.0)
                _ypad = max((_yhi - _ylo) * 0.05, 1.0)
                _y_extra = (self.plot_y_extra_spin.value()
                            if getattr(self, 'plot_y_extra_spin', None) else 0)
                ax.set_xlim(_xlo - _xpad, _xhi + _xpad)
                ax.set_ylim(_ylo - _ypad - _y_extra, _yhi + _ypad + _y_extra)
                ax.set_aspect('equal')

            # Set labels and title
            ax.set_xlabel('Distance (m)', fontsize=int(self.plot_config.get('label_font_size', '12')))
            ax.set_ylabel('Elevation (mPD)', fontsize=int(self.plot_config.get('label_font_size', '12')))
            ax.set_title(sheet_name, fontsize=int(self.plot_config.get('title_font_size', '14')), y=1.02)
            ax.grid(True, alpha=0.3, linestyle='-', linewidth=float(self.plot_config.get('grid_lines_thickness', '0.5')))
            
            # Create legend below the plot area so it doesn't obscure the
            # cross-section. ncols wraps entries into multiple columns so the
            # legend stays compact even with many geology entries.
            handles, labels = ax.get_legend_handles_labels()
            n_items = max(len(handles), 1)
            ncols = min(n_items, 3)   # up to 3 columns; adjust if needed
            legend = ax.legend(handles, labels,
                             fontsize=int(self.plot_config.get('legend_font_size', '10')),
                             loc='upper center',
                             bbox_to_anchor=(0.5, -0.12),
                             ncol=ncols,
                             borderaxespad=0,
                             frameon=True,
                             fancybox=True,
                             shadow=False,
                             framealpha=0.95)
            legend.get_frame().set_facecolor('white')

            # tight_layout is not used here because the axes are positioned
            # manually via add_axes to preserve the plot area size while the
            # legend occupies the extra figure height below.
            
            QgsMessageLog.logMessage(
                f"Successfully created plot for '{sheet_name}' with {len(column_pairs)} profiles",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
            return fig
            
        except Exception as e:
            QgsMessageLog.logMessage(f"Error plotting profile for sheet {sheet_name}: {str(e)}", 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
            return None
            
    def _plot_slopes(self, ax, df, params, column_pairs):
        """Plot slope lines and calculations using the main profile (first column pair)"""
        if not column_pairs:
            QgsMessageLog.logMessage(
                "No column pairs available for slope calculations",
                "IntegratedProfileAnalyzer", Qgis.Warning
            )
            return
            
        # Use the first column pair for slope calculations (main profile)
        main_dist_col, main_elev_col = column_pairs[0]
        
        try:
            # Import numpy for calculations
            import numpy as np
            import math
            
            # Main slope
            mask = (df[main_dist_col] >= params['start_slope']) & (df[main_dist_col] <= params['end_slope'])
            if sum(mask) > 1:
                filtered_distances = df[main_dist_col][mask]
                filtered_heights = df[main_elev_col][mask]
                
                slope = (filtered_heights.iloc[-1] - filtered_heights.iloc[0]) / \
                       (filtered_distances.iloc[-1] - filtered_distances.iloc[0])
                angle = np.arctan(slope) * (180 / np.pi)
                
                y_intercept = filtered_heights.iloc[0] - slope * filtered_distances.iloc[0]
                x_values = np.array([filtered_distances.iloc[0], filtered_distances.iloc[-1]])
                y_values = slope * x_values + y_intercept
                
                ax.plot(x_values, y_values, 
                       color=self.plot_config.get('slope_line_color', '#FF00FF'),
                       linestyle='--', 
                       linewidth=float(self.plot_config.get('slope_line_thickness', '2.0')),
                       label=f'Main slope angle: {angle:.2f}°',
                       zorder=2)
                       
                QgsMessageLog.logMessage(
                    f"Plotted main slope: {angle:.2f}° using {main_dist_col}/{main_elev_col}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                       
            # Runout slope
            mask_runout = (df[main_dist_col] >= params['runout_start']) & (df[main_dist_col] <= params['runout_end'])
            if sum(mask_runout) > 1:
                filtered_distances = df[main_dist_col][mask_runout]
                filtered_heights = df[main_elev_col][mask_runout]
                
                slope = (filtered_heights.iloc[-1] - filtered_heights.iloc[0]) / \
                       (filtered_distances.iloc[-1] - filtered_distances.iloc[0])
                angle = np.arctan(slope) * (180 / np.pi)
                
                y_intercept = filtered_heights.iloc[0] - slope * filtered_distances.iloc[0]
                x_values = np.array([filtered_distances.iloc[0], filtered_distances.iloc[-1]])
                y_values = slope * x_values + y_intercept
                
                ax.plot(x_values, y_values, 
                       color=self.plot_config.get('runout_line_color', '#00FFFF'),
                       linestyle='--', 
                       linewidth=float(self.plot_config.get('runout_line_thickness', '2.0')),
                       label=f'Runout slope angle: {angle:.2f}°',
                       zorder=2)
                       
                QgsMessageLog.logMessage(
                    f"Plotted runout slope: {angle:.2f}° using {main_dist_col}/{main_elev_col}",
                    "IntegratedProfileAnalyzer", Qgis.Info
                )
                       
            # Reference angle line (measured from horizontal)
            if params.get('ref_angle', 0) != 0:
                try:
                    max_height_idx = df[main_elev_col].idxmax()
                    max_height = df[main_elev_col][max_height_idx]
                    max_height_distance = df[main_dist_col][max_height_idx]
                    
                    # Convert angle to radians (now measured from horizontal)
                    angle_rad = math.radians(params['ref_angle'])
                    
                    # Calculate slope from the angle (tan of angle from horizontal)
                    slope_ref = -math.tan(angle_rad)  # Negative because we're going down
                    
                    # Find where the line intersects the minimum height
                    min_height = df[main_elev_col].min()
                    
                    # Calculate the horizontal distance needed
                    vertical_drop = max_height - min_height
                    horizontal_distance = vertical_drop / abs(slope_ref) if slope_ref != 0 else 100
                    
                    # Create the reference line
                    x_ref = np.array([max_height_distance, max_height_distance + horizontal_distance])
                    y_ref = np.array([max_height, min_height])
                    
                    ax.plot(x_ref, y_ref, 
                           color=self.plot_config.get('reference_line_color', '#FFFF00'),
                           linestyle='--', 
                           linewidth=float(self.plot_config.get('reference_line_thickness', '1.5')),
                           label=f"{self.plot_config.get('legend_angular_elevation', 'Reference Angle')} ({params['ref_angle']}°)",
                           zorder=2)
                           
                    QgsMessageLog.logMessage(
                        f"Plotted reference angle: {params['ref_angle']}° using {main_dist_col}/{main_elev_col}",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                except Exception as ref_error:
                    QgsMessageLog.logMessage(
                        f"Error plotting reference angle: {str(ref_error)}",
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
                   
        except Exception as e:
            QgsMessageLog.logMessage(f"Error plotting slopes: {str(e)}", 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
    
    # Note: _plot_additional_profiles method has been removed as all profiles 
    # are now automatically detected and plotted in plot_single_profile method
        
    def save_project(self):
        """Save the current project"""
        try:
            project = QgsProject.instance()
            if project.fileName():
                if project.write():
                    QMessageBox.information(self, "Success", "Project saved successfully!")
                    QgsMessageLog.logMessage("Project saved successfully", "IntegratedProfileAnalyzer", Qgis.Info)
                else:
                    QMessageBox.warning(self, "Error", "Failed to save project!")
            else:
                # Project hasn't been saved before
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Save Project",
                    self.last_directory,
                    "QGIS Project Files (*.qgz *.qgs)"
                )
                if file_path:
                    if project.write(file_path):
                        QMessageBox.information(self, "Success", f"Project saved to:\n{file_path}")
                        self.last_directory = os.path.dirname(file_path)
                        self.settings.setValue("lastDirectory", self.last_directory)
                    else:
                        QMessageBox.warning(self, "Error", "Failed to save project!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving project: {str(e)}")
    
    def setup_polyline_tab(self):
        """Setup the polyline creation tab"""
        layout = QVBoxLayout()
        
        # Layer settings
        layer_group = QGroupBox("Layer Settings")
        layer_layout = QVBoxLayout()
        
        # Layer name
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Layer Name:"))
        self.layer_name = QLineEdit("Polylines")
        name_layout.addWidget(self.layer_name)
        layer_layout.addLayout(name_layout)
        
        # Save format
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Save As:"))
        self.format_combo = QComboBox()
        self.format_combo.addItems(["Memory Layer", "GeoPackage", "Shapefile"])
        format_layout.addWidget(self.format_combo)
        layer_layout.addLayout(format_layout)
        
        # 3D option
        self.enable_3d = QCheckBox("Create 3D polylines (with Z values)")
        layer_layout.addWidget(self.enable_3d)
        
        layer_group.setLayout(layer_layout)
        layout.addWidget(layer_group)
        
        # Draping settings
        drape_group = QGroupBox("Raster Draping Settings")
        drape_layout = QVBoxLayout()
        
        # Enable draping
        self.enable_draping = QCheckBox("Enable raster draping")
        self.enable_draping.stateChanged.connect(self.toggle_draping_options)
        drape_layout.addWidget(self.enable_draping)
        
        # Raster selection
        raster_layout = QHBoxLayout()
        raster_layout.addWidget(QLabel("DEM/Raster:"))
        self.raster_combo = QComboBox()
        self.refresh_raster_list()
        raster_layout.addWidget(self.raster_combo)
        drape_layout.addLayout(raster_layout)
        
        # Refresh button
        refresh_btn = QPushButton("Refresh Raster List")
        refresh_btn.clicked.connect(self.refresh_raster_list)
        drape_layout.addWidget(refresh_btn)
        
        # Sampling options
        self.sample_group = QGroupBox("Sampling Options")
        sample_layout = QVBoxLayout()
        
        # Sampling distance
        dist_layout = QHBoxLayout()
        dist_layout.addWidget(QLabel("Sample every:"))
        self.sample_distance = QDoubleSpinBox()
        self.sample_distance.setRange(0.1, 1000)
        self.sample_distance.setValue(10)
        self.sample_distance.setSuffix(" m")
        dist_layout.addWidget(self.sample_distance)
        sample_layout.addLayout(dist_layout)
        
        # Offset
        offset_layout = QHBoxLayout()
        offset_layout.addWidget(QLabel("Vertical offset:"))
        self.vertical_offset = QDoubleSpinBox()
        self.vertical_offset.setRange(-1000, 1000)
        self.vertical_offset.setValue(0)
        self.vertical_offset.setSuffix(" m")
        offset_layout.addWidget(self.vertical_offset)
        sample_layout.addLayout(offset_layout)
        
        self.sample_group.setLayout(sample_layout)
        drape_layout.addWidget(self.sample_group)
        
        drape_group.setLayout(drape_layout)
        layout.addWidget(drape_group)
        
        # Drawing tools
        draw_group = QGroupBox("Drawing Tools")
        draw_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel("Left-click: Add point\nRight-click: Finish polyline")
        instructions.setStyleSheet("background-color: #f0f0f0; padding: 5px; border-radius: 3px;")
        draw_layout.addWidget(instructions)
        
        # Buttons with tooltips
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("Start Drawing")
        self.start_btn.clicked.connect(self.start_drawing)
        self.start_btn.setToolTip("Click to activate drawing mode, then click on the map to create polylines")
        btn_layout.addWidget(self.start_btn)
        
        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.clicked.connect(self.clear_polylines)
        self.clear_btn.setToolTip("Remove all drawn polylines from the list")
        btn_layout.addWidget(self.clear_btn)
        
        draw_layout.addLayout(btn_layout)
        
        # Polyline list with compact sizing
        draw_layout.addWidget(QLabel("Created Polylines:"))
        self.polyline_list = QListWidget()
        self.polyline_list.setMinimumHeight(60)  # Reduced from 80
        self.polyline_list.setMaximumHeight(120)  # Reduced from 150
        self.polyline_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.polyline_list.setAlternatingRowColors(True)
        self.polyline_list.setToolTip("List of drawn polylines. Each polyline will be numbered as you create them.")
        # Make list items more compact
        self.polyline_list.setStyleSheet("QListWidget::item { padding: 2px; }")
        draw_layout.addWidget(self.polyline_list)
        
        draw_group.setLayout(draw_layout)
        layout.addWidget(draw_group)
        
        # Save button with tooltip
        self.save_btn = QPushButton("Save Polylines to Layer")
        self.save_btn.clicked.connect(self.save_polylines)
        self.save_btn.setEnabled(False)
        self.save_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 8px; }")
        self.save_btn.setToolTip("Save all drawn polylines to a new layer in QGIS. Make sure you have at least one polyline drawn.")
        layout.addWidget(self.save_btn)
        
        # Wrap in scroll area
        _scroll = QScrollArea()
        _scroll.setWidgetResizable(True)
        _scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _content = QWidget()
        _content.setLayout(layout)
        _scroll.setWidget(_content)
        _outer = QVBoxLayout()
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.addWidget(_scroll)
        self.polyline_tab.setLayout(_outer)
        self.toggle_draping_options()
        
    def setup_profile_tab(self):
        """Setup the profile extraction tab"""
        layout = QVBoxLayout()
        
        # Profile settings
        profile_group = QGroupBox("Profile Extraction Settings")
        profile_layout = QVBoxLayout()
        
        # Sample interval
        interval_layout = QHBoxLayout()
        interval_layout.addWidget(QLabel("Sample Interval:"))
        self.profile_interval = QDoubleSpinBox()
        self.profile_interval.setRange(0.1, 1000.0)
        self.profile_interval.setValue(0.5)
        self.profile_interval.setSingleStep(0.1)
        self.profile_interval.setDecimals(2)
        self.profile_interval.setSuffix(" map units")
        interval_layout.addWidget(self.profile_interval)
        profile_layout.addLayout(interval_layout)
        
        # Column configuration button
        self.config_columns_btn = QPushButton("⚙️ Configure Output Columns")
        self.config_columns_btn.clicked.connect(self.configure_output_columns)
        self.config_columns_btn.setStyleSheet("""
            QPushButton {
                background-color: #673AB7;
                color: white;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #5E35B1;
            }
        """)
        profile_layout.addWidget(self.config_columns_btn)
        
        # Display current column configuration
        self.column_config_display = QLabel()
        self.column_config_display.setWordWrap(True)
        self.column_config_display.setStyleSheet("""
            QLabel {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-family: monospace;
                font-size: 10px;
            }
        """)
        self.update_column_config_display()
        profile_layout.addWidget(self.column_config_display)
        
        # Output format
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Output Format:"))
        self.output_format = QComboBox()
        if EXCEL_ENGINE == 'xlsxwriter':
            self.output_format.addItems(["Excel (XLSX)", "CSV Files"])
        else:
            self.output_format.addItems(["CSV Files"])
        format_layout.addWidget(self.output_format)
        profile_layout.addLayout(format_layout)
        
        # Output settings
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("Output Location:"))
        self.output_path_edit = QLineEdit()
        self.output_path_edit.setReadOnly(True)
        output_layout.addWidget(self.output_path_edit)
        
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self.browse_output)
        output_layout.addWidget(browse_btn)
        profile_layout.addLayout(output_layout)
        
        profile_group.setLayout(profile_layout)
        layout.addWidget(profile_group)
        
        # Layer selection
        layer_group = QGroupBox("Layer Selection")
        layer_layout = QVBoxLayout()
        
        # Info label
        info_label = QLabel("The tool will process all visible line layers with the first visible raster layer as DEM")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e3f2fd; padding: 5px; border-radius: 3px;")
        layer_layout.addWidget(info_label)
        
        # Refresh layers button
        refresh_layers_btn = QPushButton("Check Visible Layers")
        refresh_layers_btn.clicked.connect(self.check_visible_layers)
        layer_layout.addWidget(refresh_layers_btn)
        
        # Status display with compact sizing
        self.layer_status = QTextEdit()
        self.layer_status.setReadOnly(True)
        self.layer_status.setMinimumHeight(80)  # Reduced from 100
        self.layer_status.setMaximumHeight(150)  # Reduced from 200
        self.layer_status.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.layer_status.setStyleSheet("QTextEdit { padding: 4px; }")  # Reduce internal padding
        layer_layout.addWidget(self.layer_status)
        
        layer_group.setLayout(layer_layout)
        layout.addWidget(layer_group)
        
        # Extract profiles button
        self.extract_btn = QPushButton("Extract Profiles")
        self.extract_btn.clicked.connect(self.extract_profiles)
        self.extract_btn.setEnabled(False)
        self.extract_btn.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 8px; }")
        layout.addWidget(self.extract_btn)
        
        layout.addStretch()
        # Wrap in scroll area
        _scroll = QScrollArea()
        _scroll.setWidgetResizable(True)
        _scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _content = QWidget()
        _content.setLayout(layout)
        _scroll.setWidget(_content)
        _outer = QVBoxLayout()
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.addWidget(_scroll)
        self.profile_tab.setLayout(_outer)
        
    def setup_intersection_tab(self):
        """Setup the intersection points tab"""
        layout = QVBoxLayout()
        
        # Line layer selection
        line_group = QGroupBox("Select Line Layers")
        line_layout = QVBoxLayout()
        
        line_layout.addWidget(QLabel("Select line layers for intersection:"))
        
        # Add refresh button for line layers
        refresh_line_btn = QPushButton("🔄 Refresh Visible Layers")
        refresh_line_btn.clicked.connect(self.refresh_intersection_layers)
        refresh_line_btn.setToolTip("Refresh layer lists to show only currently visible layers in QGIS.\n" +
                                   "Hidden layers will be excluded from the lists.")
        line_layout.addWidget(refresh_line_btn)
        
        self.line_list = QListWidget()
        self.line_list.setSelectionMode(QAbstractItemView.MultiSelection)
        self.line_list.setMinimumHeight(100)  # Reduced from 120
        self.line_list.setMaximumHeight(150)  # Set maximum to prevent excessive growth
        self.line_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.line_list.setStyleSheet("QListWidget::item { padding: 2px; }")  # Compact items
        self.populate_line_layers()
        line_layout.addWidget(self.line_list)
        
        line_group.setLayout(line_layout)
        layout.addWidget(line_group)
        
        # Intersection settings
        intersect_group = QGroupBox("Intersection Settings")
        intersect_layout = QVBoxLayout()
        
        # Layer to intersect
        intersect_layer_layout = QHBoxLayout()
        intersect_layer_layout.addWidget(QLabel("Layer to Intersect:"))
        self.intersect_combo = QComboBox()
        self.populate_vector_layers()
        intersect_layer_layout.addWidget(self.intersect_combo)
        intersect_layout.addLayout(intersect_layer_layout)
        
        # DEM selection
        dem_layout = QHBoxLayout()
        dem_layout.addWidget(QLabel("DEM Raster:"))
        self.dem_combo = QComboBox()
        self.populate_raster_layers()
        dem_layout.addWidget(self.dem_combo)
        intersect_layout.addLayout(dem_layout)
        
        # Point interval
        interval_layout = QHBoxLayout()
        interval_layout.addWidget(QLabel("Point Interval (meters):"))
        self.intersect_interval = QSpinBox()
        self.intersect_interval.setRange(1, 10000)
        self.intersect_interval.setValue(10)
        interval_layout.addWidget(self.intersect_interval)
        intersect_layout.addLayout(interval_layout)
        
        # Include interval points
        self.include_interval_points = QCheckBox("Include regular interval points")
        self.include_interval_points.setChecked(True)
        intersect_layout.addWidget(self.include_interval_points)

        # Buffer / corridor width
        buffer_layout = QHBoxLayout()
        buffer_layout.addWidget(QLabel("Section corridor (m):"))
        self.intersect_buffer_spin = QSpinBox()
        self.intersect_buffer_spin.setRange(0, 10000)
        self.intersect_buffer_spin.setValue(0)
        self.intersect_buffer_spin.setSuffix(" m")
        self.intersect_buffer_spin.setToolTip(
            "Buffer the section polyline by this distance before intersecting.\n"
            "Strike lines within this corridor are projected perpendicularly\n"
            "onto the section centerline so they generate an intersection point.\n"
            "Set to 0 for exact geometric intersection only."
        )
        buffer_layout.addWidget(self.intersect_buffer_spin)
        buffer_layout.addStretch()
        intersect_layout.addLayout(buffer_layout)

        # Output prefix
        prefix_layout = QHBoxLayout()
        prefix_layout.addWidget(QLabel("Output Layer Prefix:"))
        self.output_prefix = QLineEdit("Profile_Points")
        prefix_layout.addWidget(self.output_prefix)
        intersect_layout.addLayout(prefix_layout)
        
        intersect_group.setLayout(intersect_layout)
        layout.addWidget(intersect_group)
        
        # Output settings
        output_group = QGroupBox("Output Settings")
        output_layout = QVBoxLayout()
        
        # Output format selection
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Output Format:"))
        self.intersect_output_format = QComboBox()
        self.intersect_output_format.addItems(["Excel (.xlsx)", "CSV Files"])
        self.intersect_output_format.currentTextChanged.connect(self.on_intersect_format_changed)
        format_layout.addWidget(self.intersect_output_format)
        output_layout.addLayout(format_layout)
        
        # Output path
        output_path_layout = QHBoxLayout()
        output_path_layout.addWidget(QLabel("Output Location:"))
        self.intersect_output_edit = QLineEdit()
        self.intersect_output_edit.setReadOnly(True)
        output_path_layout.addWidget(self.intersect_output_edit)
        
        self.intersect_browse_btn = QPushButton("Browse...")
        self.intersect_browse_btn.clicked.connect(self.browse_intersect_output)
        output_path_layout.addWidget(self.intersect_browse_btn)
        output_layout.addLayout(output_path_layout)
        
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)
        
        # Add geometry check button
        fix_settings_btn = QPushButton("⚙️ Configure Processing to Skip Invalid Geometries")
        fix_settings_btn.clicked.connect(self.configure_processing_settings)
        fix_settings_btn.setStyleSheet("QPushButton { background-color: #FFC107; color: black; padding: 6px; }")
        layout.addWidget(fix_settings_btn)
        
        # Process button
        self.process_intersection_btn = QPushButton("Create Intersection Points")
        self.process_intersection_btn.clicked.connect(self.process_intersections)
        self.process_intersection_btn.setEnabled(False)
        self.process_intersection_btn.setStyleSheet("QPushButton { background-color: #FF9800; color: white; font-weight: bold; padding: 8px; }")
        layout.addWidget(self.process_intersection_btn)
        
        layout.addStretch()
        # Wrap in scroll area
        _scroll = QScrollArea()
        _scroll.setWidgetResizable(True)
        _scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        _content = QWidget()
        _content.setLayout(layout)
        _scroll.setWidget(_content)
        _outer = QVBoxLayout()
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.addWidget(_scroll)
        self.intersection_tab.setLayout(_outer)
        
    def refresh_intersection_layers(self):
        """Refresh all layer lists in the intersection tab - only show visible layers"""
        # Save current selections
        selected_line_items = [item.text() for item in self.line_list.selectedItems()]
        current_intersect = self.intersect_combo.currentText()
        current_dem = self.dem_combo.currentText()
        
        # Refresh lists (now only showing visible layers)
        self.populate_line_layers()
        self.populate_vector_layers()
        self.populate_raster_layers()
        
        # Track restoration success
        restored_lines = 0
        intersect_restored = False
        dem_restored = False
        
        # Restore line layer selections where possible (only if layers are still visible)
        for i in range(self.line_list.count()):
            item = self.line_list.item(i)
            if item.text() in selected_line_items:
                item.setSelected(True)
                restored_lines += 1
                
        # Restore intersection combo selection (only if layer is still visible)
        index = self.intersect_combo.findText(current_intersect)
        if index >= 0:
            self.intersect_combo.setCurrentIndex(index)
            intersect_restored = True
        else:
            # Layer is no longer visible, show warning
            if current_intersect and self.intersect_combo.count() > 0:
                QMessageBox.information(self, "Layer Hidden", 
                    f"Previously selected intersection layer '{current_intersect}' is now hidden.")
                
        # Restore DEM combo selection (only if layer is still visible)
        index = self.dem_combo.findText(current_dem)
        if index >= 0:
            self.dem_combo.setCurrentIndex(index)
            dem_restored = True
        else:
            # Layer is no longer visible, show warning
            if current_dem and self.dem_combo.count() > 0:
                QMessageBox.information(self, "Layer Hidden", 
                    f"Previously selected DEM layer '{current_dem}' is now hidden.")
        
        # Count visible layers by type
        visible_line_count = self.line_list.count()
        visible_vector_count = self.intersect_combo.count()
        visible_raster_count = self.dem_combo.count()
        
        # Show comprehensive feedback dialog
        feedback_message = f"Layer lists refreshed!\n\n"
        feedback_message += f"📊 Visible Layers Summary:\n"
        feedback_message += f"• Line layers: {visible_line_count}\n"
        feedback_message += f"• Vector layers: {visible_vector_count}\n"
        feedback_message += f"• Raster layers: {visible_raster_count}\n\n"
        
        if selected_line_items or current_intersect or current_dem:
            feedback_message += f"🔄 Selection Restoration:\n"
            if selected_line_items:
                feedback_message += f"• Line layers: {restored_lines}/{len(selected_line_items)} restored\n"
            if current_intersect:
                status = "✓ restored" if intersect_restored else "✗ hidden"
                feedback_message += f"• Intersection layer: {status}\n"
            if current_dem:
                status = "✓ restored" if dem_restored else "✗ hidden"
                feedback_message += f"• DEM layer: {status}\n\n"
        
        feedback_message += f"💡 Note: Only visible layers are shown.\n"
        feedback_message += f"Toggle layer visibility in QGIS Layers Panel to add/remove layers from these lists."
        
        # Show warnings if no layers of any type are visible
        warnings = []
        if visible_line_count == 0:
            warnings.append("⚠️ No visible line layers found!")
        if visible_vector_count == 0:
            warnings.append("⚠️ No visible vector layers found!")
        if visible_raster_count == 0:
            warnings.append("⚠️ No visible raster layers found!")
        
        if warnings:
            feedback_message += f"\n\n" + "\n".join(warnings)
            feedback_message += f"\n\nMake some layers visible in QGIS and refresh again."
        
        QMessageBox.information(self, "Refresh Complete", feedback_message)
        
    def on_intersect_format_changed(self, text):
        """Handle intersection output format change"""
        self.intersect_output_edit.clear()
        self.process_intersection_btn.setEnabled(False)
        
    def browse_intersect_output(self):
        """Browse for intersection output location"""
        if self.intersect_output_format.currentText() == "Excel (.xlsx)":
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                os.path.join(self.last_directory, "intersection_points.xlsx"),
                "Excel Files (*.xlsx);;All Files (*.*)"
            )
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                self.intersect_output_edit.setText(file_path)
                self.last_directory = os.path.dirname(file_path)
                self.settings.setValue("lastDirectory", self.last_directory)
                self.process_intersection_btn.setEnabled(True)
        else:  # CSV Files
            dir_path = QFileDialog.getExistingDirectory(
                self,
                "Select Output Directory for CSV Files",
                self.last_directory
            )
            if dir_path:
                self.intersect_output_edit.setText(dir_path)
                self.last_directory = dir_path
                self.settings.setValue("lastDirectory", self.last_directory)
            
    def populate_line_layers(self):
        """Populate the list widget with visible line layers only"""
        self.line_list.clear()
        
        # Get the layer tree root to check visibility
        root = QgsProject.instance().layerTreeRoot()
        
        layers = QgsProject.instance().mapLayers().values()
        for layer in layers:
            # Check if layer is line geometry and visible
            if (layer.type() == QgsMapLayer.VectorLayer and 
                layer.geometryType() == QgsWkbTypes.LineGeometry):
                
                # Find the layer node in the tree and check if it's visible
                node = root.findLayer(layer.id())
                if node and node.isVisible():
                    self.line_list.addItem(layer.name())
                
    def populate_vector_layers(self):
        """Populate combo box with visible vector layers only"""
        self.intersect_combo.clear()
        
        # Get the layer tree root to check visibility
        root = QgsProject.instance().layerTreeRoot()
        
        layers = QgsProject.instance().mapLayers().values()
        for layer in layers:
            if layer.type() == QgsMapLayer.VectorLayer:
                # Find the layer node in the tree and check if it's visible
                node = root.findLayer(layer.id())
                if node and node.isVisible():
                    self.intersect_combo.addItem(layer.name(), layer)
                
    def populate_raster_layers(self):
        """Populate combo box with visible raster layers only"""
        self.dem_combo.clear()
        
        # Get the layer tree root to check visibility
        root = QgsProject.instance().layerTreeRoot()
        
        layers = QgsProject.instance().mapLayers().values()
        for layer in layers:
            if layer.type() == QgsMapLayer.RasterLayer:
                # Find the layer node in the tree and check if it's visible
                node = root.findLayer(layer.id())
                if node and node.isVisible():
                    self.dem_combo.addItem(layer.name(), layer)
                
    def get_selected_line_layers(self):
        """Get selected line layers from the list widget"""
        selected_layers = []
        for item in self.line_list.selectedItems():
            layer_name = item.text()
            for layer in QgsProject.instance().mapLayers().values():
                if layer.name() == layer_name:
                    selected_layers.append(layer)
        return selected_layers
        
    def create_intersection_points(self, line_layer, intersect_layer):
        """Create intersection points between line layer and another layer."""
        try:
            buffer_dist = (self.intersect_buffer_spin.value()
                           if getattr(self, 'intersect_buffer_spin', None) else 0)

            # Fix geometries in the intersect layer
            QgsMessageLog.logMessage("Checking for invalid geometries...", "IntegratedProfileAnalyzer", Qgis.Info)
            params = {'INPUT': intersect_layer, 'OUTPUT': 'memory:fixed_geometries'}
            try:
                fixed_intersect_layer = processing.run("native:fixgeometries", params)['OUTPUT']
                QgsMessageLog.logMessage("Fixed invalid geometries", "IntegratedProfileAnalyzer", Qgis.Info)
            except Exception as fix_error:
                QgsMessageLog.logMessage(f"Could not fix geometries, using original: {fix_error}",
                                         "IntegratedProfileAnalyzer", Qgis.Warning)
                fixed_intersect_layer = intersect_layer

            geom_type = fixed_intersect_layer.geometryType()

            # ── Buffered corridor mode ─────────────────────────────────────
            if buffer_dist > 0 and geom_type == QgsWkbTypes.LineGeometry:
                return self._create_buffered_intersection_points(
                    line_layer, fixed_intersect_layer, buffer_dist)

            # ── Exact intersection mode (original behaviour) ───────────────
            if geom_type == QgsWkbTypes.PolygonGeometry:
                params = {'INPUT': fixed_intersect_layer, 'OUTPUT': 'memory:boundaries'}
                boundaries = processing.run("native:polygonstolines", params)
                params = {
                    'INPUT': line_layer,
                    'INTERSECT': boundaries['OUTPUT'],
                    'OUTPUT': 'memory:raw_intersections',
                    'INPUT_FIELDS': [], 'INTERSECT_FIELDS': [],
                    'INTERSECT_FIELDS_PREFIX': ''
                }
                intersections = processing.run("native:lineintersections", params)
                params = {'INPUT': intersections['OUTPUT'], 'OUTPUT': 'memory:intersection_points'}
                result = processing.run("native:deleteduplicategeometries", params)
                return result['OUTPUT']

            elif geom_type == QgsWkbTypes.LineGeometry:
                params = {
                    'INPUT': line_layer,
                    'INTERSECT': fixed_intersect_layer,
                    'OUTPUT': 'memory:intersection_points',
                    'INPUT_FIELDS': [], 'INTERSECT_FIELDS': [],
                    'INTERSECT_FIELDS_PREFIX': ''
                }
                result = processing.run("native:lineintersections", params)
                return result['OUTPUT']

            elif geom_type == QgsWkbTypes.PointGeometry:
                QMessageBox.warning(self, 'Warning',
                    'Point layers are not supported for intersection. Please select a line or polygon layer.')
                return None
                
        except Exception as e:
            error_msg = str(e)
            if "invalid geometry" in error_msg.lower():
                QMessageBox.critical(self, 'Geometry Error',
                    f'Invalid geometry detected in layer "{intersect_layer.name()}".\n\n'
                    f'Feature with invalid geometry cannot be processed.\n'
                    f'The layer appears to have self-intersecting polygons.\n\n'
                    f'Solutions:\n'
                    f'1. Use QGIS "Fix Geometries" tool on the layer first\n'
                    f'2. Contact the data provider about the geometry issue\n'
                    f'3. Try using a different intersection layer\n\n'
                    f'Error details: {error_msg}')
            else:
                QMessageBox.critical(self, 'Error', f'Error creating intersection points: {error_msg}')
            return None

    def _create_buffered_intersection_points(self, line_layer, intersect_layer, buffer_dist):
        """
        Buffered corridor intersection: for every strike-line feature in
        intersect_layer whose geometry falls within buffer_dist metres of the
        section polyline(s), project the nearest point on the section centerline
        and emit it as an intersection point with all original attributes preserved.

        All geometry operations are performed in the section (line_layer) CRS so
        that the buffer distance is in metres regardless of the intersect_layer CRS.
        Output points are written in the section layer CRS (matching the downstream
        add_elevation_field / distance calculation which uses the same layer).
        """
        line_crs       = line_layer.crs()
        intersect_crs  = intersect_layer.crs()
        project        = QgsProject.instance()

        needs_xform = (intersect_crs.isValid() and line_crs.isValid() and
                       intersect_crs.authid() != line_crs.authid())
        xform = (QgsCoordinateTransform(intersect_crs, line_crs, project)
                 if needs_xform else None)

        QgsMessageLog.logMessage(
            f"BufferedIntersect: line CRS={line_crs.authid()}  "
            f"intersect CRS={intersect_crs.authid()}  "
            f"needs_xform={needs_xform}  buffer={buffer_dist}m",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

        # ── Build combined section geometry in line_layer CRS ─────────────
        line_geoms = []
        for f in line_layer.getFeatures():
            g = f.geometry()
            if not g.isEmpty():
                line_geoms.append(g)
        if not line_geoms:
            QgsMessageLog.logMessage("BufferedIntersect: no section geometries found.",
                                     "IntegratedProfileAnalyzer", Qgis.Warning)
            return None

        combined_line = line_geoms[0]
        for g in line_geoms[1:]:
            combined_line = combined_line.combine(g)

        corridor_poly = combined_line.buffer(buffer_dist, 16)

        QgsMessageLog.logMessage(
            f"BufferedIntersect: corridor bbox = {corridor_poly.boundingBox().toString()}",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

        # ── Build spatial index on intersect_layer IN line_layer CRS ──────
        si = QgsSpatialIndex()
        feat_map   = {}      # fid -> original QgsFeature
        geom_map   = {}      # fid -> transformed QgsGeometry (in line CRS)

        for feat in intersect_layer.getFeatures():
            g = feat.geometry()
            if g.isEmpty():
                continue
            if xform is not None:
                g_t = QgsGeometry(g)          # copy
                try:
                    g_t.transform(xform)
                except Exception as e:
                    QgsMessageLog.logMessage(
                        f"BufferedIntersect: CRS transform failed for feat {feat.id()}: {e}",
                        "IntegratedProfileAnalyzer", Qgis.Warning)
                    continue
            else:
                g_t = g
            feat_map[feat.id()] = feat
            geom_map[feat.id()] = g_t
            # Add a temporary feature with the transformed geometry for indexing
            tmp = QgsFeature(feat.id())
            tmp.setGeometry(g_t)
            si.addFeature(tmp)

        candidate_ids = si.intersects(corridor_poly.boundingBox())
        QgsMessageLog.logMessage(
            f"BufferedIntersect: {len(feat_map)} features indexed, "
            f"{len(candidate_ids)} candidates in corridor bbox.",
            "IntegratedProfileAnalyzer", Qgis.Info
        )

        # ── Output layer: point layer in line_layer CRS, same fields ──────
        out_layer = QgsVectorLayer(
            f'Point?crs={line_crs.authid()}', 'buffered_intersections', 'memory')
        pr = out_layer.dataProvider()
        pr.addAttributes(intersect_layer.fields().toList())
        out_layer.updateFields()

        new_feats = []
        skipped   = 0

        for fid in candidate_ids:
            if fid not in geom_map:
                continue
            g_t = geom_map[fid]

            # Precise check: feature must actually intersect the corridor polygon
            if not corridor_poly.intersects(g_t):
                skipped += 1
                continue

            # Project: find the nearest point on the section line to this feature
            closest = combined_line.nearestPoint(g_t)
            if closest is None or closest.isEmpty():
                skipped += 1
                continue

            orig_feat = feat_map[fid]
            new_feat  = QgsFeature(out_layer.fields())
            new_feat.setGeometry(closest)
            for field in intersect_layer.fields():
                try:
                    new_feat.setAttribute(field.name(), orig_feat.attribute(field.name()))
                except Exception:
                    pass
            new_feats.append(new_feat)

        pr.addFeatures(new_feats)
        out_layer.updateExtents()

        QgsMessageLog.logMessage(
            f"BufferedIntersect: projected {len(new_feats)} point(s), "
            f"skipped {skipped} (outside corridor after precise check).  "
            f"corridor=±{buffer_dist}m",
            "IntegratedProfileAnalyzer", Qgis.Info
        )
        return out_layer

    def validate_layer_geometries(self, layer):
        """Check for invalid geometries in a layer"""
        invalid_count = 0
        total_count = layer.featureCount()
        invalid_features = []
        
        for feature in layer.getFeatures():
            if not feature.geometry().isGeosValid():
                invalid_count += 1
                invalid_features.append({
                    'id': feature.id(),
                    'error': feature.geometry().lastError()
                })
                if len(invalid_features) <= 5:  # Limit details to first 5 errors
                    QgsMessageLog.logMessage(
                        f"Invalid geometry in feature {feature.id()}: {feature.geometry().lastError()}", 
                        "IntegratedProfileAnalyzer", Qgis.Warning
                    )
        
        return invalid_count, total_count, invalid_features
        
    def configure_processing_settings(self):
        """Configure QGIS processing settings to handle invalid geometries"""
        from qgis.core import QgsSettings
        settings = QgsSettings()
        
        # Set to skip invalid features
        settings.setValue("Processing/Configuration/INVALID_FEATURES_FILTERING", 1)
        
        QMessageBox.information(self, "Settings Updated",
            "Processing settings updated to skip invalid geometries.\n\n"
            "This will help avoid errors when processing layers with geometry issues.\n\n"
            "Note: This setting applies globally to all QGIS processing tools.")
            
    def create_points_along_line(self, line_layer, interval):
        """Create points at regular intervals along line"""
        params = {
            'INPUT': line_layer,
            'DISTANCE': interval,
            'START_OFFSET': 0,
            'END_OFFSET': 0,
            'OUTPUT': 'memory:temp_points'
        }
        result = processing.run("native:pointsalonglines", params)
        return result['OUTPUT']
        
    def merge_point_layers(self, layers, output_name):
        """Merge multiple point layers"""
        valid_layers = [layer for layer in layers if layer is not None]
        if not valid_layers:
            return None
        params = {
            'LAYERS': valid_layers,
            'CRS': valid_layers[0].crs(),
            'OUTPUT': 'memory:' + output_name
        }
        result = processing.run("native:mergevectorlayers", params)
        return result['OUTPUT']
        
    def add_z_values(self, point_layer, dem_layer):
        """Add Z values from DEM to points"""
        params = {
            'INPUT': point_layer,
            'RASTER': dem_layer,
            'BAND': 1,
            'SCALE': 1.0,
            'OUTPUT': 'memory:points_with_z'
        }
        result = processing.run("qgis:setzfromraster", params)
        return result['OUTPUT']
        
    def setup_layer_style(self, layer):
        """Setup visual style for point layer"""
        symbol = QgsMarkerSymbol.createSimple({
            'name': 'circle',
            'size': '2',
            'color': '255,0,0'
        })
        layer.renderer().setSymbol(symbol)
        layer.triggerRepaint()
        
    def add_elevation_field(self, layer, line_layer):
        """Add elevation and distance fields to layer - FIXED VERSION"""
        try:
            layer.startEditing()
            
            # Add fields if they don't exist
            field_names = [field.name() for field in layer.fields()]
            fields_to_add = []
            
            if 'elevation' not in field_names:
                fields_to_add.append(QgsField('elevation', QVariant.Double))
            if 'distance' not in field_names:
                fields_to_add.append(QgsField('distance', QVariant.Double))
            if 'line_id' not in field_names:
                fields_to_add.append(QgsField('line_id', QVariant.Int))
            if 'line_name' not in field_names:
                fields_to_add.append(QgsField('line_name', QVariant.String))
                
            if fields_to_add:
                layer.dataProvider().addAttributes(fields_to_add)
                layer.updateFields()
            
            # Get field indices
            elevation_idx = layer.fields().indexOf('elevation')
            distance_idx = layer.fields().indexOf('distance')
            line_id_idx = layer.fields().indexOf('line_id')
            line_name_idx = layer.fields().indexOf('line_name')
            
            # Build spatial index for line features
            line_index = QgsSpatialIndex()
            line_features = {}
            
            for line_feature in line_layer.getFeatures():
                line_index.addFeature(line_feature)
                line_features[line_feature.id()] = line_feature
            
            # Process each point
            point_count = layer.featureCount()
            progress_step = max(1, point_count // 100)
            
            for idx, feature in enumerate(layer.getFeatures()):
                # Show progress
                if idx % progress_step == 0:
                    QCoreApplication.processEvents()
                    
                point_geom = feature.geometry()
                if not point_geom or point_geom.isEmpty():
                    continue
                    
                point = point_geom.asPoint()
                
                # Get elevation from Z value if 3D
                if point_geom.constGet() and point_geom.constGet().is3D():
                    elevation = point_geom.constGet().z()
                else:
                    elevation = 0
                
                # Find candidate lines using spatial index
                buffer_distance = 10.0  # Search within 10 units
                search_rect = point_geom.buffer(buffer_distance, 5).boundingBox()
                candidate_ids = line_index.intersects(search_rect)
                
                if not candidate_ids:
                    # No lines nearby - set NULL values
                    layer.changeAttributeValue(feature.id(), elevation_idx, elevation)
                    layer.changeAttributeValue(feature.id(), distance_idx, NULL)
                    layer.changeAttributeValue(feature.id(), line_id_idx, NULL)
                    layer.changeAttributeValue(feature.id(), line_name_idx, NULL)
                    continue
                
                # Find the closest line among candidates
                min_distance = float('inf')
                closest_line_distance = 0
                closest_line_id = -1
                closest_line_name = ''
                
                for line_id in candidate_ids:
                    if line_id not in line_features:
                        continue
                        
                    line_feature = line_features[line_id]
                    line_geom = line_feature.geometry()
                    
                    # Calculate distance from point to line
                    distance_to_line = point_geom.distance(line_geom)
                    
                    # If this is the closest line so far
                    if distance_to_line < min_distance:
                        min_distance = distance_to_line
                        closest_line_id = line_feature.id()
                        
                        # Get line name
                        name_field = None
                        for field_name in ['name', 'Name', 'NAME', 'id', 'ID']:
                            if field_name in [f.name() for f in line_layer.fields()]:
                                name_field = field_name
                                break
                        
                        if name_field:
                            closest_line_name = str(line_feature[name_field])
                        else:
                            closest_line_name = f"Line_{closest_line_id}"
                        
                        # Calculate distance along the line
                        closest_line_distance = line_geom.lineLocatePoint(point_geom)
                
                # Set values - only if point is reasonably close to a line
                layer.changeAttributeValue(feature.id(), elevation_idx, elevation)
                
                if min_distance < 5.0:  # Within 5 units of a line
                    layer.changeAttributeValue(feature.id(), distance_idx, closest_line_distance)
                    layer.changeAttributeValue(feature.id(), line_id_idx, closest_line_id)
                    layer.changeAttributeValue(feature.id(), line_name_idx, closest_line_name)
                else:
                    # Too far from any line
                    layer.changeAttributeValue(feature.id(), distance_idx, NULL)
                    layer.changeAttributeValue(feature.id(), line_id_idx, NULL)
                    layer.changeAttributeValue(feature.id(), line_name_idx, NULL)
            
            layer.commitChanges()
            QgsMessageLog.logMessage(
                f"Successfully added elevation and distance to {point_count} points",
                "IntegratedProfileAnalyzer", Qgis.Info
            )
            
        except Exception as e:
            if layer.isEditable():
                layer.rollBack()
            QgsMessageLog.logMessage(f"Error adding elevation field: {str(e)}", 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
            import traceback
            QgsMessageLog.logMessage(traceback.format_exc(), 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
        
    def export_intersection_to_excel(self, layers_data, excel_path):
        """Export intersection points to Excel - Fixed version with unique sheet names"""
        try:
            if EXCEL_ENGINE == 'xlsxwriter':
                # Use xlsxwriter directly to avoid openpyxl conflicts
                import xlsxwriter
                
                workbook = xlsxwriter.Workbook(excel_path)
                used_sheet_names = set()  # Track used sheet names
                
                for layer_name, layer in layers_data:
                    # Create worksheet with safe name - remove duplicate prefixes
                    base_sheet_name = layer_name.replace("Profile_Points_", "")  # Remove prefix duplication
                    safe_name = "".join(c for c in base_sheet_name if c.isalnum() or c in (' ', '-', '_'))
                    
                    # Ensure unique sheet name
                    sheet_name = safe_name[:31]  # Excel limit is 31 characters
                    counter = 1
                    original_name = sheet_name
                    
                    while sheet_name.lower() in used_sheet_names:
                        suffix = f"_{counter}"
                        max_base_len = 31 - len(suffix)
                        sheet_name = original_name[:max_base_len] + suffix
                        counter += 1
                    
                    used_sheet_names.add(sheet_name.lower())
                    worksheet = workbook.add_worksheet(sheet_name)
                    
                    # Get field names
                    field_names = [field.name() for field in layer.fields()]
                    
                    # Template columns to ensure consistent output
                    template_columns = [
                        'id', 'name', 'length_m', 'min_z', 'max_z', 'id_2', 'name_2', 
                        'width_m', 'height_m', 'area_m2', 'measured_w', 'measured_h', 
                        'distance', 'angle', 'layer', 'path', 'elevation', 
                        'line_id', 'line_name'
                    ]
                    
                    # Create column mapping
                    column_mapping = {}
                    for col in template_columns:
                        if col in field_names:
                            column_mapping[col] = field_names.index(col)
                        else:
                            column_mapping[col] = -1
                    
                    # Write headers
                    for col_idx, col_name in enumerate(template_columns):
                        worksheet.write(0, col_idx, col_name)
                    
                    # Collect and sort features by distance
                    features_data = []
                    for feature in layer.getFeatures():
                        attrs = feature.attributes()
                        
                        # Extract distance value for sorting
                        distance_idx = field_names.index('distance') if 'distance' in field_names else -1
                        distance_val = attrs[distance_idx] if distance_idx >= 0 else 0
                        
                        # Handle NULL values
                        if distance_val is None or distance_val == NULL:
                            distance_val = 0
                            
                        features_data.append((distance_val, attrs))
                    
                    # Sort by distance
                    features_data.sort(key=lambda x: x[0])
                    
                    # Write data
                    for row_idx, (_, attrs) in enumerate(features_data):
                        for col_idx, col_name in enumerate(template_columns):
                            field_idx = column_mapping[col_name]
                            
                            if field_idx >= 0 and field_idx < len(attrs):
                                value = attrs[field_idx]
                                # Handle NULL values
                                if value is None or value == NULL:
                                    value = ""
                                worksheet.write(row_idx + 1, col_idx, value)
                            else:
                                worksheet.write(row_idx + 1, col_idx, "")
                    
                    QgsMessageLog.logMessage(
                        f"Exported {len(features_data)} features to sheet '{sheet_name}'",
                        "IntegratedProfileAnalyzer", Qgis.Info
                    )
                
                workbook.close()
                
            else:
                # Fallback to CSV
                self._export_intersection_to_csv(layers_data, excel_path)
                
        except Exception as e:
            QgsMessageLog.logMessage(f"Error exporting to Excel: {str(e)}", 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
            # Fallback to CSV
            self._export_intersection_to_csv(layers_data, excel_path)
            QMessageBox.warning(None, "Export Warning", 
                f"Excel export failed, saved as CSV instead:\n{str(e)}")
            
    def _export_intersection_to_csv(self, layers_data, output_path):
        """Export intersection points to CSV"""
        if output_path.endswith('.xlsx'):
            # If Excel path provided, create CSV directory with similar name
            csv_dir = os.path.splitext(output_path)[0] + "_csv"
        else:
            # Use the directory provided
            csv_dir = os.path.join(output_path, "intersection_points")
            
        os.makedirs(csv_dir, exist_ok=True)
        
        for layer_name, layer in layers_data:
            safe_name = "".join(c for c in layer_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            csv_path = os.path.join(csv_dir, f"{safe_name}.csv")
            
            try:
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    # Get field names
                    field_names = [field.name() for field in layer.fields()]
                    writer = csv.DictWriter(f, fieldnames=field_names)
                    writer.writeheader()
                    
                    # Write features
                    for feature in layer.getFeatures():
                        attrs = feature.attributes()
                        row_dict = dict(zip(field_names, attrs))
                        writer.writerow(row_dict)
            except Exception as e:
                QgsMessageLog.logMessage(f"Error writing CSV {csv_path}: {str(e)}", 
                                       "IntegratedProfileAnalyzer", Qgis.Warning)
                    
        return csv_dir
                    
    def process_intersections(self):
        """Process intersection points"""
        # Check if project is saved
        if not QgsProject.instance().fileName():
            reply = QMessageBox.question(
                self,
                "Project Not Saved",
                "Your project hasn't been saved. It's strongly recommended to save before processing.\n\nDo you want to save now?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            if reply == QMessageBox.Yes:
                self.save_project()
            elif reply == QMessageBox.Cancel:
                return
                
        if not self.intersect_output_edit.text():

            QMessageBox.warning(self, 'Warning', 'Please select an output location')
            return            
        selected_line_layers = self.get_selected_line_layers()
        if len(selected_line_layers) == 0:
            QMessageBox.warning(self, 'Warning', 'Please select at least one line layer')
            return
            
        intersect_layer = self.intersect_combo.currentData()
        dem_layer = self.dem_combo.currentData()
        interval = self.intersect_interval.value()
        output_prefix = self.output_prefix.text().strip()
        
        if not all([intersect_layer, dem_layer, interval]):
            QMessageBox.warning(self, 'Warning', 
                'Please select intersection layer, DEM layer, and interval')
            return
            
        # Validate the intersection layer for invalid geometries
        invalid_count, total_count, invalid_features = self.validate_layer_geometries(intersect_layer)
        
        if invalid_count > 0:
            reply = QMessageBox.question(
                self,
                "Invalid Geometries Detected",
                f"The intersection layer '{intersect_layer.name()}' has {invalid_count} invalid geometries out of {total_count} features.\n\n"
                f"These invalid geometries may cause processing errors or be skipped during intersection.\n\n"
                f"Common issues include self-intersecting polygons (like the one you encountered).\n\n"
                f"Options:\n"
                f"1. Continue (invalid geometries will be automatically fixed/skipped)\n"
                f"2. Cancel and manually fix geometries first using QGIS 'Fix Geometries' tool\n\n"
                f"Continue processing?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
        try:
            processed_layers = []
            
            progress = QProgressDialog("Processing intersection points...", "Cancel", 
                0, len(selected_line_layers), self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            for idx, line_layer in enumerate(selected_line_layers):
                progress.setValue(idx)
                progress.setLabelText(f"Processing {line_layer.name()}")
                QCoreApplication.processEvents()
                
                if progress.wasCanceled():
                    break
                    
                point_layers = []
                
                # Create intersection points
                intersection_points = self.create_intersection_points(line_layer, intersect_layer)
                if intersection_points is not None:
                    point_layers.append(intersection_points)
                    
                # Create interval points if requested
                if self.include_interval_points.isChecked():
                    interval_points = self.create_points_along_line(line_layer, interval)
                    point_layers.append(interval_points)
                    
                if not point_layers:
                    QMessageBox.warning(self, 'Warning',
                        f'No points created for layer {line_layer.name()}')
                    continue
                    
                # Set output name
                if output_prefix:
                    current_output_name = f"{output_prefix}_{line_layer.name()}"
                else:
                    current_output_name = f"Points_{line_layer.name()}"
                    
                # Merge point layers
                merged_points = self.merge_point_layers(point_layers, current_output_name)
                if merged_points is None:
                    continue
                
                # Add Z values from DEM
                points_with_z = self.add_z_values(merged_points, dem_layer)
                points_with_z.setName(current_output_name)
                
                # Add elevation and distance fields
                self.add_elevation_field(points_with_z, line_layer)
                
                # Setup style
                self.setup_layer_style(points_with_z)
                
                # Add to map
                QgsProject.instance().addMapLayer(points_with_z)
                processed_layers.append((current_output_name, points_with_z))
                
            progress.close()
            
            if processed_layers:
                # Export based on selected format
                output_path = self.intersect_output_edit.text()
                
                if self.intersect_output_format.currentText() == "Excel (.xlsx)":
                    self.export_intersection_to_excel(processed_layers, output_path)
                    message = f'Excel file exported to: {output_path}'
                else:  # CSV Files
                    csv_dir = self._export_intersection_to_csv(processed_layers, output_path)
                    message = f'CSV files exported to: {csv_dir}'
                
                QMessageBox.information(self, 'Success',
                    'Intersection points created successfully!\n\n' +
                    message + '\n\n' +
                    'To view in profile:\n' +
                    '1. Open Profile Tool\n' +
                    '2. Click "Add Layer"\n' +
                    '3. Select the desired layer\n' +
                    '4. The points will show in profile along your line')
            else:
                QMessageBox.warning(self, 'Warning',
                    'No intersection points were created. Check your input layers.')
                    
        except Exception as e:
            QMessageBox.critical(self, 'Error', 
                f'An error occurred: {str(e)}\n\n{traceback.format_exc()}')
    
    def browse_output(self):
        """Browse for output location"""
        if self.output_format.currentText().startswith("Excel"):
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Save Excel File", 
                os.path.join(self.last_directory, "elevation_profiles.xlsx"),
                "Excel files (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                self.output_path_edit.setText(file_path)
                self.last_directory = os.path.dirname(file_path)
                self.settings.setValue("lastDirectory", self.last_directory)
                self.extract_btn.setEnabled(True)
        else:
            # For CSV, select directory
            dir_path = QFileDialog.getExistingDirectory(
                self,
                "Select Output Directory",
                self.last_directory
            )
            
            if dir_path:
                self.output_path_edit.setText(dir_path)
                self.last_directory = dir_path
                self.settings.setValue("lastDirectory", self.last_directory)
                self.extract_btn.setEnabled(True)
            
    def check_visible_layers(self):
        """Check and display visible layers"""
        root = QgsProject.instance().layerTreeRoot()
        
        # Check for line layers
        line_layers = []
        for layer in QgsProject.instance().mapLayers().values():
            node = root.findLayer(layer.id())
            if (node and node.isVisible() and 
                layer.isValid() and 
                layer.type() == QgsMapLayer.VectorLayer and 
                layer.geometryType() == QgsWkbTypes.LineGeometry):
                line_layers.append(layer.name())
        
        # Check for raster layers
        raster_layer = None
        for layer in QgsProject.instance().mapLayers().values():
            node = root.findLayer(layer.id())
            if (node and node.isVisible() and 
                layer.type() == QgsMapLayer.RasterLayer):
                raster_layer = layer.name()
                break
        
        # Update status
        status_text = f"<b>Visible Line Layers ({len(line_layers)}):</b><br>"
        if line_layers:
            for layer in line_layers:
                status_text += f"• {layer}<br>"
        else:
            status_text += "<font color='red'>No visible line layers found!</font><br>"
        
        status_text += f"<br><b>DEM/Raster Layer:</b><br>"
        if raster_layer:
            status_text += f"• {raster_layer}"
        else:
            status_text += "<font color='red'>No visible raster layer found!</font>"
        
        self.layer_status.setHtml(status_text)
        
    def toggle_draping_options(self):
        """Enable/disable draping options based on checkbox"""
        enabled = self.enable_draping.isChecked()
        self.raster_combo.setEnabled(enabled)
        self.sample_group.setEnabled(enabled)
        self.enable_3d.setChecked(enabled)
        
    def refresh_raster_list(self):
        """Refresh the list of available raster layers"""
        self.raster_combo.clear()
        
        for layer in QgsProject.instance().mapLayers().values():
            if isinstance(layer, QgsRasterLayer):
                self.raster_combo.addItem(layer.name(), layer)
                
        if self.raster_combo.count() == 0:
            self.raster_combo.addItem("No raster layers available")
            
    def start_drawing(self):
        """Start the drawing mode"""
        self.map_tool = PolylineMapTool(self.iface.mapCanvas())
        self.map_tool.finished.connect(self.on_polyline_finished)
        self.iface.mapCanvas().setMapTool(self.map_tool)
        
        self.start_btn.setText("Drawing... (right-click to finish)")
        self.start_btn.setEnabled(False)
        
    def on_polyline_finished(self, points):
        """Handle completed polyline"""
        if len(points) < 2:
            return
            
        polyline_data = {
            'points': points,
            'name': f"Polyline {len(self.temp_polylines) + 1}"
        }
        self.temp_polylines.append(polyline_data)
        
        self.polyline_list.addItem(polyline_data['name'])
        self.save_btn.setEnabled(True)
        
        self.start_btn.setText("Start Drawing")
        self.start_btn.setEnabled(True)
        
        if self.map_tool:
            self.map_tool.reset()
            
    def drape_polyline_on_raster(self, points, raster_layer):
        """Drape polyline points on raster surface"""
        draped_points = []
        
        provider = raster_layer.dataProvider()
        extent = raster_layer.extent()
        no_data_value = provider.sourceNoDataValue(1)
        
        sample_dist = self.sample_distance.value()
        
        for i in range(len(points) - 1):
            start_point = points[i]
            end_point = points[i + 1]
            
            dx = end_point.x() - start_point.x()
            dy = end_point.y() - start_point.y()
            segment_length = math.sqrt(dx**2 + dy**2)
            
            num_samples = max(2, int(segment_length / sample_dist))
            
            for j in range(num_samples):
                t = j / (num_samples - 1) if num_samples > 1 else 0
                x = start_point.x() + t * dx
                y = start_point.y() + t * dy
                
                sample_point = QgsPointXY(x, y)
                z = self.vertical_offset.value()
                
                if extent.contains(sample_point):
                    value, success = provider.sample(sample_point, 1)
                    
                    if success and not math.isnan(value):
                        if no_data_value is None or value != no_data_value:
                            z = float(value) + self.vertical_offset.value()
                
                draped_points.append(QgsPoint(x, y, z))
        
        return draped_points
        
    def save_polylines(self):
        """Save all polylines to a layer"""
        if not self.temp_polylines:
            QMessageBox.warning(self, "No Polylines", "No polylines to save!")
            return
            
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Save Polylines", 
            "Are you sure you want to save the polylines? Make sure your project is saved first.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
            
        layer_name = self.layer_name.text() or "Polylines"
        
        try:
            if self.enable_3d.isChecked():
                geom_type = "LineStringZ"
            else:
                geom_type = "LineString"
                
            project_crs = QgsProject.instance().crs()
            format_idx = self.format_combo.currentIndex()
            
            if format_idx == 0:  # Memory layer
                layer = QgsVectorLayer(f"{geom_type}?crs={project_crs.authid()}", 
                    layer_name, "memory")
            else:
                # Get save path
                if format_idx == 1:  # GeoPackage
                    file_path, _ = QFileDialog.getSaveFileName(
                        self, 
                        "Save GeoPackage", 
                        os.path.join(self.last_directory, f"{layer_name}.gpkg"),
                        "GeoPackage (*.gpkg)"
                    )
                else:  # Shapefile
                    file_path, _ = QFileDialog.getSaveFileName(
                        self, 
                        "Save Shapefile", 
                        os.path.join(self.last_directory, f"{layer_name}.shp"),
                        "Shapefile (*.shp)"
                    )
                    
                if not file_path:
                    return
                    
                self.last_directory = os.path.dirname(file_path)
                self.settings.setValue("lastDirectory", self.last_directory)
                    
                layer = QgsVectorLayer(f"{geom_type}?crs={project_crs.authid()}", 
                    layer_name, "memory")
            
            # Add fields
            provider = layer.dataProvider()
            fields = QgsFields()
            fields.append(QgsField("id", QVariant.Int))
            fields.append(QgsField("name", QVariant.String))
            fields.append(QgsField("length_m", QVariant.Double))
            if self.enable_3d.isChecked():
                fields.append(QgsField("min_z", QVariant.Double))
                fields.append(QgsField("max_z", QVariant.Double))
            provider.addAttributes(fields)
            layer.updateFields()
            
            # Add features
            features = []
            
            progress = QProgressDialog("Saving polylines...", "Cancel", 0, len(self.temp_polylines), self)
            progress.setWindowModality(Qt.WindowModal)
            
            for idx, polyline_data in enumerate(self.temp_polylines):
                progress.setValue(idx)
                if progress.wasCanceled():
                    return
                    
                points = polyline_data['points']
                min_z = 0
                max_z = 0
                
                if self.enable_draping.isChecked() and self.raster_combo.currentData():
                    raster_layer = self.raster_combo.currentData()
                    if isinstance(raster_layer, QgsRasterLayer) and self.enable_3d.isChecked():
                        draped_points = self.drape_polyline_on_raster(points, raster_layer)
                        line = QgsLineString(draped_points)
                        
                        z_values = [p.z() for p in draped_points]
                        if z_values:
                            min_z = min(z_values)
                            max_z = max(z_values)
                    else:
                        line = QgsLineString([QgsPoint(p.x(), p.y()) for p in points])
                else:
                    if self.enable_3d.isChecked():
                        line = QgsLineString([QgsPoint(p.x(), p.y(), 0) for p in points])
                    else:
                        line = QgsLineString([QgsPoint(p.x(), p.y()) for p in points])
                
                feature = QgsFeature()
                feature.setGeometry(QgsGeometry(line))
                
                attrs = [idx + 1, polyline_data['name'], line.length()]
                
                if self.enable_3d.isChecked():
                    attrs.extend([min_z, max_z])
                    
                feature.setAttributes(attrs)
                features.append(feature)
            
            progress.close()
            provider.addFeatures(features)
            
            # Save to file if needed
            if format_idx > 0:
                options = QgsVectorFileWriter.SaveVectorOptions()
                if format_idx == 1:
                    options.driverName = "GPKG"
                else:
                    options.driverName = "ESRI Shapefile"
                options.fileEncoding = "UTF-8"
                
                error = QgsVectorFileWriter.writeAsVectorFormatV2(
                    layer, file_path, QgsProject.instance().transformContext(), options)
                
                if error[0] == QgsVectorFileWriter.NoError:
                    saved_layer = QgsVectorLayer(file_path, layer_name, "ogr")
                    QgsProject.instance().addMapLayer(saved_layer)
                else:
                    raise Exception(f"Error saving file: {error[1]}")
            else:
                QgsProject.instance().addMapLayer(layer)
            
            self.clear_polylines()
            
            QMessageBox.information(self, "Success", 
                f"Saved {len(features)} polyline(s) to layer '{layer_name}'")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save polylines:\n{str(e)}")
            
    def clear_polylines(self):
        """Clear all temporary polylines"""
        self.temp_polylines = []
        self.polyline_list.clear()
        self.save_btn.setEnabled(False)
        if self.map_tool:
            self.map_tool.reset()
            
    def extract_profiles(self):
        """Extract elevation profiles from visible line layers"""
        if not self.output_path_edit.text():
            QMessageBox.warning(self, "No Output Location", "Please select an output location first!")
            return
            
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Extract Profiles", 
            "This will process all visible line layers. Large datasets may take time.\n\nMake sure your project is saved. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
            
        try:
            self._run_profile_extraction()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Profile extraction failed:\n{str(e)}")
            
    def _run_profile_extraction(self):
        """Run the profile extraction process with chunked processing"""
        distance_interval = self.profile_interval.value()
        output_path = self.output_path_edit.text()
        
        # Get visible layers
        root = QgsProject.instance().layerTreeRoot()
        
        line_layers = []
        for layer in QgsProject.instance().mapLayers().values():
            node = root.findLayer(layer.id())
            if (node and node.isVisible() and 
                layer.isValid() and 
                layer.type() == QgsMapLayer.VectorLayer and 
                layer.geometryType() == QgsWkbTypes.LineGeometry):
                line_layers.append(layer)
        
        if not line_layers:
            QMessageBox.warning(self, "No Line Layers", "No visible line layers found!")
            return
        
        dem_layer = None
        for layer in QgsProject.instance().mapLayers().values():
            node = root.findLayer(layer.id())
            if (node and node.isVisible() and 
                layer.type() == QgsMapLayer.RasterLayer):
                dem_layer = layer
                break
        
        if not dem_layer:
            QMessageBox.warning(self, "No DEM Layer", "No visible raster/DEM layer found!")
            return
        
        # Create progress dialog
        progress = QProgressDialog("Processing profiles...", "Cancel", 0, 100, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.show()
        
        # Use context manager for Excel writer
        if self.output_format.currentText().startswith("Excel"):
            with SafeExcelWriter(output_path) as excel_writer:
                self._process_layers(line_layers, dem_layer, distance_interval, 
                                   excel_writer, progress, None)
                excel_writer.save()
        else:
            # For CSV, create directory
            csv_dir = os.path.join(output_path, "elevation_profiles")
            os.makedirs(csv_dir, exist_ok=True)
            self._process_layers(line_layers, dem_layer, distance_interval, 
                               None, progress, csv_dir)
        
        progress.close()
        
        QMessageBox.information(self, "Success", 
            f"Profile extraction complete!\n\nSaved to: {output_path}")
            
    def _process_layers(self, line_layers, dem_layer, distance_interval, 
                       excel_writer, progress, csv_dir):
        """Process layers with chunked feature processing"""
        total_features = sum(layer.featureCount() for layer in line_layers)
        features_processed = 0
        
        CHUNK_SIZE = 100  # Process features in chunks
        
        for line_layer in line_layers:
            # Process features in chunks
            feature_iterator = line_layer.getFeatures()
            feature_count = 0
            
            while True:
                chunk = []
                for _ in range(CHUNK_SIZE):
                    try:
                        feature = next(feature_iterator)
                        chunk.append(feature)
                    except StopIteration:
                        break
                        
                if not chunk:
                    break
                    
                for feature in chunk:
                    percent = (features_processed * 100) // total_features
                    progress.setValue(percent)
                    progress.setLabelText(f"Processing {line_layer.name()} - Feature {feature_count+1}")
                    QCoreApplication.processEvents()
                    
                    if progress.wasCanceled():
                        return
                    
                    try:
                        self._process_single_feature(feature, line_layer, dem_layer, 
                                                   distance_interval, excel_writer, csv_dir)
                    except Exception as e:
                        QgsMessageLog.logMessage(
                            f"Error processing feature {feature.id()}: {str(e)}",
                            "IntegratedProfileAnalyzer", Qgis.Warning
                        )
                    
                    features_processed += 1
                    feature_count += 1
                
                # Periodic garbage collection
                if feature_count % 500 == 0:
                    gc.collect()
                    
    def _process_single_feature(self, feature, line_layer, dem_layer, 
                              distance_interval, excel_writer, csv_dir):
        """Process a single feature"""
        line_geom = feature.geometry()
        if not line_geom or line_geom.isEmpty():
            return
        
        length = line_geom.length()
        if length <= 0:
            return
        
        # ── CRS transformation: line layer → DEM layer ────────────────────────
        # The DEM (e.g. 1_Second_DEM / SRTM) is often in WGS84 (EPSG:4326)
        # while the line layer is in a projected CRS (e.g. MGA Zone 51).
        # Without this transform every identify() call returns no data.
        line_crs = line_layer.crs()
        dem_crs  = dem_layer.crs()
        needs_transform = (dem_crs.isValid() and line_crs.isValid() and
                           dem_crs.authid() != line_crs.authid())
        if needs_transform:
            crs_transform = QgsCoordinateTransform(
                line_crs, dem_crs, QgsProject.instance()
            )
        
        distances = []
        elevations = []
        
        num_points = int(length / distance_interval) + 1
        
        for point_idx in range(num_points):
            current_distance = point_idx * distance_interval
            if current_distance > length:
                break
            
            point = line_geom.interpolate(current_distance).asPoint()
            
            # Transform to DEM CRS if needed
            if needs_transform:
                try:
                    point = crs_transform.transform(point)
                except Exception:
                    continue
            
            ident = dem_layer.dataProvider().identify(point, QgsRaster.IdentifyFormatValue)
            if ident.isValid():
                results = ident.results()
                if results:
                    try:
                        elevation = results[1]
                        # Use `is not None` — avoids skipping elevation == 0.0
                        if elevation is not None and elevation != -9999:
                            distances.append(current_distance)
                            elevations.append(elevation)
                    except (KeyError, IndexError):
                        try:
                            elevation = next(iter(results.values()))
                            if elevation is not None and elevation != -9999:
                                distances.append(current_distance)
                                elevations.append(elevation)
                        except StopIteration:
                            continue
        
        if distances and elevations:
            # Get feature name — safe attribute access (field may not exist)
            def _safe_attr(feat, field):
                try:
                    val = feat.attribute(field)
                    return str(val) if val is not None and val != NULL else None
                except Exception:
                    return None

            feature_name = (
                _safe_attr(feature, 'Name') or
                _safe_attr(feature, 'name') or
                _safe_attr(feature, 'label') or
                _safe_attr(feature, 'id') or
                f"Feature_{feature.id()}"
            )
            sheet_name = str(feature_name)
            
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            # Build headers and rows based on column configuration
            if hasattr(self, 'column_config'):
                active_columns = [col for col in self.column_config['columns'] if col['active']]
                headers = [col['name'] for col in active_columns]
                
                # Find the first distance and elevation columns only
                first_distance_col = None
                first_elevation_col = None
                
                for i, col in enumerate(active_columns):
                    if col['type'] == 'distance' and first_distance_col is None:
                        first_distance_col = i
                    elif col['type'] == 'elevation' and first_elevation_col is None:
                        first_elevation_col = i
                
                # Prepare data - only populate the first distance and elevation columns
                max_rows = len(distances)
                rows = []
                
                for i in range(max_rows):
                    row = [None] * len(active_columns)  # Initialize all columns as empty
                    
                    # Only populate the first distance and elevation columns with actual data
                    if first_distance_col is not None:
                        row[first_distance_col] = distances[i]
                    if first_elevation_col is not None:
                        row[first_elevation_col] = elevations[i]
                    
                    # All other columns (additional distance/elevation pairs, marker columns, custom columns) 
                    # remain as None/empty for manual entry
                    
                    rows.append(row)
            else:
                # Fallback to default headers - only populate first two columns
                headers = [
                    'Distance_01', 'Height_01', 'Distance_02', 'Height_02',
                    'Distance_03', 'Height_03', 'Distance_04', 'Height_04',
                    'Marker_Distance', 'Marker_Height', 'Marker_Distance_2', 'Marker_Height_2'
                ]
                
                rows = []
                for i in range(len(distances)):
                    # Only populate first two columns (Distance_01, Height_01), leave the rest empty
                    row = [distances[i], elevations[i]] + [None] * 10
                    rows.append(row)
            
            if excel_writer:
                excel_writer.add_sheet(sheet_name, {'headers': headers, 'rows': rows})
            else:
                # Save as individual CSV
                safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                csv_path = os.path.join(csv_dir, f"{safe_name}.csv")
                
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
    
    def show_help(self):
        """Show help dialog with user guidance"""
        help_dialog = QDialog(self)
        help_dialog.setWindowTitle("Help - QGIS section drafter")
        help_dialog.setMinimumSize(600, 500)
        help_dialog.resize(700, 600)
        
        layout = QVBoxLayout()
        
        # Create scrollable text area for help content
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setHtml("""
        <h2>QGIS section drafter - User Guide</h2>
        
        <h3>🔧 Common Display Issues & Solutions:</h3>
        <ul>
            <li><b>Can't see all items in lists/tables:</b> Try resizing the dialog window or individual panels</li>
            <li><b>Content cut off:</b> Look for scroll bars on the right side of widgets</li>
            <li><b>Tables too narrow:</b> Drag column dividers to resize table columns</li>
            <li><b>Dialog too small:</b> Drag the dialog edges to make it larger</li>
        </ul>
        
        <h3>📋 Workflow Overview:</h3>
        <ol>
            <li><b>Create Polylines:</b> Draw lines on the map for analysis</li>
            <li><b>Extract Profiles:</b> Get elevation data along your lines</li>
            <li><b>Find Intersections:</b> Calculate where lines cross other features</li>
            <li><b>Plot Profiles:</b> Create matplotlib graphs of your data</li>
            <li><b>Excel Merge:</b> Combine multiple datasets</li>
        </ol>
        
        <h3>💡 Tips for Better Performance:</h3>
        <ul>
            <li>Save your QGIS project before starting</li>
            <li>Test with small datasets first</li>
            <li>Make sure all layers are loaded and visible</li>
            <li>Check that coordinate systems match between layers</li>
        </ul>
        
        <h3>⚠️ Troubleshooting:</h3>
        <ul>
            <li><b>Tool not responding:</b> Check the QGIS message log (View → Panels → Log Messages)</li>
            <li><b>Empty results:</b> Verify your layer selections and data projections</li>
            <li><b>Memory issues:</b> Close other applications and try smaller datasets</li>
        </ul>
        
        <h3>🎨 Display Customization:</h3>
        <ul>
            <li>Use "Configure Plot Appearance" to change colors and styles</li>
            <li>Resize tables by dragging column headers</li>
            <li>All lists support multi-selection (Ctrl+Click)</li>
            <li>Window size and position are automatically saved</li>
        </ul>
        """)
        
        layout.addWidget(help_text)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(help_dialog.close)
        layout.addWidget(close_btn)
        
        help_dialog.setLayout(layout)
        help_dialog.exec_()
    
    def make_combo_user_friendly(self, combo_box, tooltip_text=""):
        """Make a combo box more user-friendly with better sizing and tooltips"""
        combo_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        combo_box.setMinimumHeight(25)
        combo_box.setMaxVisibleItems(10)
        if tooltip_text:
            combo_box.setToolTip(tooltip_text)
        return combo_box
    
    def make_layout_compact(self, layout):
        """Make any layout more compact by reducing spacing and margins"""
        if hasattr(layout, 'setSpacing'):
            layout.setSpacing(4)
        if hasattr(layout, 'setContentsMargins'):
            layout.setContentsMargins(6, 6, 6, 6)
        return layout
            
    def closeEvent(self, event):
        """Clean up when dialog closes"""
        try:
            if self.map_tool:
                self.iface.mapCanvas().unsetMapTool(self.map_tool)
            
            # Force garbage collection
            gc.collect()
            
            QgsMessageLog.logMessage("Dialog closed", "IntegratedProfileAnalyzer", Qgis.Info)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error during cleanup: {str(e)}", 
                                   "IntegratedProfileAnalyzer", Qgis.Warning)
        event.accept()


# ── Custom additions (structural dip, matplotlib window, section azimuth) ─────
# All logic lives in custom_additions.py which plugin updates never overwrite.
# If this block is removed by an update, just paste these lines back here.
try:
    from .custom_additions import apply_custom_patches
    apply_custom_patches(CombinedGeospatialToolDialog)
except Exception as _e:
    print(f"[QGIS section drafter] custom_additions not loaded: {_e}")